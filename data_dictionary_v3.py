#!/usr/bin/env python3
"""
data_dictionary_v3.py  (AEP Swiss Army Knife)
=============================================
Data Dictionary v3.4. Sucks every XDM schema out of an AEP sandbox, filters down
to the ones that actually matter, and writes a tabbed Excel workbook: a master
field index, one tab per schema (full field list, ready to paste into Claude
for a Mermaid ERD), a Datasets tab mapping every dataset's friendly name to its
SQL table (system) name, and -- with --data-dict -- real field coverage + top-5
example values sampled from ingested data.

The workbook is marked STRICTLY CONFIDENTIAL: with --data-dict it contains
real sampled customer data.

Pick a credential set (keyring-first from the OS vault, or a plaintext creds/
folder where keyring is unavailable) and the tool will:
  1. Authenticate against IMS (client_credentials).
  2. GET /sandboxes -- every sandbox the credential can see.
  3. Prompt you to pick which sandbox(es) to read (Enter = prod by default).
  4. Per chosen sandbox: list all tenant schemas, count datasets per schema
     (Catalog), and pull every descriptor (identities, relationships, and the
     friendly "dual" labels).
  5. FILTER the schemas (see below), print the full list to screen with a
     KEEP / DROP verdict per schema so you can see exactly what was excluded.
  6. For each KEPT schema, resolve its full field list (dot notation + data
     type), join identities / relationships / dual labels, and write
        output/Data Dictionary - <Client> - <YYYY-MM-DD>.xlsx
     (Summary, a master Field Index, a Schemas index, a Datasets / table-name
     map, then one tab per schema).

Filtering (phase 1 -- transparent and reported, tune the constants below):
  KEEP   schema is referenced by >= 1 dataset (the UI "DATASETS" column).
  DROP   no-dataset   nothing ingests into it (ad-hoc audience schemas,
                      unused drafts) -- not interesting yet.
  DROP   adhoc        meta:class is the ad-hoc class (auto-generated ingestion
                      schemas).
  DROP   ajo          Adobe Journey Optimizer-managed schema (title / extends
                      heuristic) -- out of scope for now.

The dual-label (alternateDisplayInfo) count is reported per sandbox so you can
see whether this tenant uses friendly field labels at all. The Friendly Name
column is always present (blank when a field has no alternate label).

Phase 2 -- data dictionary (--data-dict): for each field, sample real ingested
records (Snappy-Parquet) via the Data Access API and tally COVERAGE (% of
sampled rows where the field is populated) and the TOP-5 values
(value(count), pipe-separated). One download covers every field at once -- no
Query Service, no per-field calls. Two extra columns are added to each sampled
schema's tab. Defaults to ALL kept schemas (a bare --data-dict samples every
tab -- expect this to run for a long while); --data-dict=<substr> narrows it to
schemas whose title matches (e.g. --data-dict=profile for the Profile schema).

v3.1 -- Profile coverage fix: a Profile-class schema is the post-merge UNION
(identity-deduped, last-write-wins), so sampling its feeding datasets tallies
pre-merge fragments and reads falsely sparse. For Profile-class schemas the tool
now samples the Profile Snapshot Export dataset belonging to the DEFAULT merge
policy instead (auto-resolved: default merge policy -> the snapshot dataset
tagged with that policy id whose schemaRef is profile__union). Snapshots are
huge, so it downloads the smallest non-empty partition file. Override the dataset
with --profile-snapshot=<datasetId> (e.g. for orgs where the default isn't the
one you want).

v3.2 -- Two things ship under this label:
  * Bundled Luma demo dataset (demo/luma/): Adobe's public platform-utils sample
    data, organised + tenant-normalised for offline demos and tests.
  * SQL table (system) names. Each dataset carries an AEP Query Service table name
    in tags['adobe/pqs/table'] (e.g. "acme_cja_acme_order_event_dataset") -- the
    normalized SYSTEM name you SELECT ... FROM, distinct from the friendly dataset
    name. The workbook now surfaces it so SQL can be formed against the right table
    straight from this file: a new "Datasets" tab maps every dataset's friendly
    name -> table name -> schema; the Schemas index and Field Index each gain a
    "Table name(s)" column; and every schema tab's header block lists its SQL
    table name(s) next to the field list.
  * Profile column on the Datasets tab (tags.unifiedProfile): flags every dataset
    enabled for Unified Profile and the Profile Snapshot Export(s) -- so you know
    which tables to query for whole profiles. Profile-related rows sort to the top.
  * Data-completeness warnings so a partial dictionary is never read as gospel:
    any schema whose --data-dict coverage is missing/partial/empty is listed in a
    Summary "DATA COMPLETENESS" block and flagged with a red banner on its own tab
    (coverage MISSING != 0%). The Profile Snapshot Export is sampled once per run:
    if its (huge) file manifest 504s under load, the failure is cached so the
    other Profile schemas fail fast instead of each re-hitting the dead snapshot.

v3.3 -- Readability and provenance:
  * An Audiences tab: every audience with tags, owner, last-modified and its
    segmentation rule (PQL) rendered readable.
  * A provenance line on every tab (when it ran, which script version, which
    commit, and a loud warning when the working tree was dirty).
  * A How to Use tab, filters + frozen headers everywhere, and the credential
    name dropped from the FILENAME in favour of the sandbox.

v3.4 -- The credential name is gone from the workbook itself. The titles inside
the file used to fall back to the credential's service name, so a workbook read
with a key stored as "<name>" was headed "Data Dictionary - <Name>" -- the name
of an entry in our own vault, printed where the reader expects the subject of
the report.
Titles now use the client name ONLY when the credential record carries an
explicit "client" key, and otherwise identify the workbook by SANDBOX, matching
what the filename has said since v3.3. Every tab also carries a link to the
release notes, so a dictionary found months later can be traced to what the
version that produced it actually did.

Full version history: RELEASE_NOTES.md (link in the workbook's How to Use tab).

openpyxl is needed for the XLSX; pyarrow + tzdata for --data-dict (all optional,
pip install -r requirements.txt).

Usage:
    python data_dictionary_v3.py                       # interactive menus
    python data_dictionary_v3.py "acme-beta"          # pick creds by service name
    python data_dictionary_v3.py "acme-beta" --sandbox=prod,dev1
    python data_dictionary_v3.py "acme-beta" --sandbox=all
    python data_dictionary_v3.py "acme-k" --sandbox=prod --data-dict       # ALL schemas
    python data_dictionary_v3.py "acme-k" --sandbox=prod --data-dict --dd-rows=2000
    python data_dictionary_v3.py "acme-k" --sandbox=prod --data-dict=profile  # one schema
    python data_dictionary_v3.py "acme-k" --sandbox=prod --data-dict \
        --profile-snapshot=67dc0539eaafb02aeeb92ed7   # override snapshot dataset

Credentials come from the OS keyring (Windows Credential Manager) via aep_creds,
falling back to a plaintext creds/ folder where keyring is unavailable; manage
them with credential_validator_v2.py. Service names are the keys you stored --
run `credential_validator_v2.py list` to see them.
"""

from __future__ import annotations

import json
import logging
import re
import ssl
import sys
import time
import traceback
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import aep_creds  # keyring-backed credential store (replaces creds/*.json)

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------
SCRIPT_NAME    = "data_dictionary_v3"
SCRIPT_VERSION = "3.4.0"
SCRIPT_DATE    = "2026-08-14"
SCRIPT_AUTHOR  = "Barry Mann (barrymann.com)"
AUTHOR_SITE     = "https://barrymann.com"
AUTHOR_LINKEDIN = "https://www.linkedin.com/in/barrymann/"

# Where the version history lives. Surfaced in the workbook (provenance line +
# How to Use tab) so whoever opens a dictionary months from now can find out
# what the version that produced it actually did.
RELEASE_NOTES_URL = ("https://github.com/mannbarry2/aep-swiss-army-knife"
                     "/blob/main/RELEASE_NOTES.md")

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"

IMS_URL = "https://ims-na1.adobelogin.com/ims/token"
PLATFORM = "https://platform.adobe.io"
SANDBOX_LIST_URL = f"{PLATFORM}/data/foundation/sandbox-management/sandboxes"
SCHEMAS_URL = f"{PLATFORM}/data/foundation/schemaregistry/tenant/schemas"
DESCRIPTORS_URL = f"{PLATFORM}/data/foundation/schemaregistry/tenant/descriptors"
DATASETS_URL = f"{PLATFORM}/data/foundation/catalog/dataSets"
CATALOG_BATCHES_URL = f"{PLATFORM}/data/foundation/catalog/batches"
EXPORT_URL = f"{PLATFORM}/data/foundation/export"   # Data Access (batch files)
UPS_MERGE_POLICIES_URL = f"{PLATFORM}/data/core/ups/config/mergePolicies"

# Profile coverage (v3.1): a Profile-class schema's coverage must be sampled from
# the merged UNION (Profile Snapshot Export), not its pre-merge feeding datasets.
PROFILE_CLASS = "https://ns.adobe.com/xdm/context/profile"
PROFILE_UNION = "https://ns.adobe.com/xdm/context/profile__union"

# Worksheet tab colour for Profile-class schemas (they stand out from the
# event/lookup tabs; their coverage is sampled from the Profile Snapshot Export
# union, not pre-merge feeds). Explained in the Summary tab's colour key.
PROFILE_TAB_COLOR = "7030A0"   # purple

# Data dictionary (phase 2): default schema scope when --data-dict is given
# with no value, and the default sample size. Default is "all" -- a bare
# --data-dict samples EVERY kept schema (one coverage/top-values pass per tab).
# Narrow it with --data-dict=<substr> (e.g. --data-dict=profile) for one schema.
DD_DEFAULT_SCOPE = "all"
DD_DEFAULT_ROWS = 1000
# Profile Snapshot Export partitions are large; give each file download longer.
DD_SNAPSHOT_TIMEOUT = 120

DEFAULT_SCOPES = (
    "openid,AdobeID,read_organizations,"
    "additional_info.projectedProductContext,session"
)

# Schema Registry media types.
XED_LIST = "application/vnd.adobe.xed+json"            # full(ish) resource list
XED_FULL = "application/vnd.adobe.xed-full+json; version=1"  # allOf resolved
# Descriptors come back grouped by @type, each value a list of full objects.
DESCRIPTOR_ACCEPT = "application/vnd.adobe.xdm+json"

# --- Filtering knobs --------------------------------------------------------
ADHOC_CLASS = "https://ns.adobe.com/xdm/data/adhoc"
# Ad-hoc / auto-generated schemas: the canonical adhoc class, OR a per-schema
# generated class whose id tail is a long hex blob (Adobe Campaign / audience
# / import dumps -- the bulk of a tenant), OR an auto-generated title.
ADHOC_CLASS_HEX = re.compile(r"^[0-9a-fA-F]{24,}$")
ADHOC_TITLE_SUBSTRINGS = [
    "xdm schema for dataset",   # " XDM Schema for dataset xxx. Random identifier:"
]
ADHOC_TITLE_PREFIXES = [
    "adhoc xdm schema",
    "schema for audience",      # auto-created audience schemas (hundreds)
]

# Title markers that flag an Adobe Journey Optimizer-managed schema.
# Case-insensitive. Tune these if a real schema gets caught, or an AJO one
# slips through -- every DROP:ajo is printed so you can check.
AJO_TITLE_PREFIXES = ["ao ", "ajo "]
AJO_TITLE_SUBSTRINGS = [
    "ajo",
    "journey optimizer",
    "secondary recipient",
    "aooutput",
]
# extends ($id) substrings -- specific to AJO namespaces so a customer schema
# that merely uses the word "journey" (e.g. "Acme Custom Journey Schema") is
# NOT mis-flagged.
AJO_EXTENDS_SUBSTRINGS = [
    "customerjourneymanagement",
    "/ajo/",
    "messagefeedback",
    "messageexecution",
]

# Adobe-managed "plumbing" schemas (Offer Decisioning, CJA, Audience Portal,
# Journey Orchestration, Unified Profile segment defs, Adobe Campaign
# Recipients) -- real schemas with datasets, but not the customer's own data
# model, so they're dropped before the ERD. Matched on the standard class tail
# OR a title marker.
SYSTEM_CLASS_TAILS = {"decisionevent", "offeritem", "segmentdefinition",
                      "journey"}
SYSTEM_TITLE_SUBSTRINGS = [
    "audience portal",
    "cja audiences",
    "orchestrated campaign",
    "journey inbound",
    "journey orchestration",
    "experience decisioning",
    "decisionevents",
]
SYSTEM_TITLE_EXACT = {"recipients"}

# Test / placeholder schemas. Tunable -- every DROP:test is printed so you can
# check none of your real schemas were caught.
TEST_TITLE_PATTERNS = [
    r"\btest\b",
    r"test profiles",
    r"exclusionstest",
    r"schema for test",
    r"\bsto \d",
    r"\bpoc\b",
    r"^jai ",
]
_TEST_RE = re.compile("|".join(TEST_TITLE_PATTERNS), re.I)

KNOWN_CLASSES = {
    "https://ns.adobe.com/xdm/context/profile": "XDM Individual Profile",
    "https://ns.adobe.com/xdm/context/experienceevent": "XDM ExperienceEvent",
    "https://ns.adobe.com/xdm/data/adhoc": "Ad Hoc",
    "https://ns.adobe.com/xdm/data/record": "XDM Record",
    "https://ns.adobe.com/xdm/data/time-series": "XDM Time-series",
    "https://ns.adobe.com/xdm/classes/segment-definition": "Segment definition",
}

# ----------------------------------------------------------------------------
# ANSI / logging - matches the house style (credential_validator.py et al.)
# ----------------------------------------------------------------------------
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        h = kernel32.GetStdHandle(-11)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
            kernel32.SetConsoleMode(h, mode.value | 0x0004)
    except Exception:
        pass

ANSI = {
    "reset": "\033[0m", "bold": "\033[1m", "dim": "\033[2m",
    "red": "\033[31m", "green": "\033[32m", "yellow": "\033[33m",
    "blue": "\033[34m", "magenta": "\033[35m", "cyan": "\033[36m",
}
LEVEL_COLOR = {
    "DEBUG": ANSI["dim"], "INFO": ANSI["green"],
    "WARNING": ANSI["yellow"],
    "ERROR": ANSI["red"] + ANSI["bold"],
    "CRITICAL": ANSI["red"] + ANSI["bold"],
}


class ColoredFormatter(logging.Formatter):
    def format(self, record):
        color = LEVEL_COLOR.get(record.levelname, "")
        ts = self.formatTime(record, "%H:%M:%S")
        return (
            f"{ANSI['dim']}{ts}{ANSI['reset']} "
            f"{color}[{record.levelname:<7}]{ANSI['reset']} "
            f"{record.getMessage()}"
        )


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(ColoredFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger("data_dictionary_v3")
SSL_CTX = ssl._create_unverified_context()


# ----------------------------------------------------------------------------
# HTTP / IMS / credential helpers  (shared house style)
# ----------------------------------------------------------------------------
def http(url, method="GET", headers=None, data=None, timeout=60):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read(), dict(r.headers)


def flatten_err(text: str, limit: int = 200) -> str:
    return " ".join((text or "").split())[:limit]


def authenticate(conf):
    payload = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": conf["client_id"],
        "client_secret": conf["client_secret"],
        "scope": conf.get("scopes") or DEFAULT_SCOPES,
    }).encode("utf-8")
    body, _ = http(
        conf.get("oauth_url") or IMS_URL,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data=payload,
    )
    return json.loads(body)


def aep_headers(token, conf, sandbox=None, accept="application/json"):
    headers = {
        "Authorization": f"Bearer {token}",
        "x-api-key": conf.get("api_key") or conf["client_id"],
        "x-gw-ims-org-id": conf["org_id"],
        "Accept": accept,
    }
    if sandbox:
        headers["x-sandbox-name"] = sandbox
    return headers


def list_sandboxes(token, conf):
    try:
        body, _ = http(SANDBOX_LIST_URL, headers=aep_headers(token, conf))
        data = json.loads(body)
        return True, data.get("sandboxes") or []
    except urllib.error.HTTPError as e:
        return False, f"HTTP {e.code}: {flatten_err(e.read().decode(errors='replace'))}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


# ----------------------------------------------------------------------------
# AEP reads -- schemas / datasets / descriptors
# ----------------------------------------------------------------------------
def _abs(href: str) -> str:
    if not href:
        return ""
    if href.startswith("http"):
        return href
    return PLATFORM + href


def _next_url(data):
    """Pull the 'next page' URL out of a registry response. _links.next may be
    an object ({'href': ...}) or a bare href string; _page.next is a string
    cursor. Returns an absolute URL or None."""
    if not isinstance(data, dict):
        return None
    nxt = (data.get("_links") or {}).get("next")
    if isinstance(nxt, str) and nxt:
        return _abs(nxt)
    if isinstance(nxt, dict) and nxt.get("href"):
        return _abs(nxt["href"])
    return None


def get_all_schemas(token, conf, sandbox):
    """Page through GET /tenant/schemas (xed+json) following _links.next.
    Returns a list of raw schema-resource dicts."""
    out = []
    url = f"{SCHEMAS_URL}?limit=300"
    headers = aep_headers(token, conf, sandbox, accept=XED_LIST)
    while url:
        body, _ = http(url, headers=headers)
        data = json.loads(body)
        results = data.get("results") if isinstance(data, dict) else data
        out.extend(r for r in (results or []) if isinstance(r, dict))
        url = _next_url(data)
    return out


def get_full_schema(token, conf, sandbox, ref):
    """GET one schema resolved (allOf merged) so it carries a flat 'properties'
    tree. 'ref' is the meta:altId (preferred) or the $id."""
    url = f"{SCHEMAS_URL}/{urllib.parse.quote(ref, safe='')}"
    headers = aep_headers(token, conf, sandbox, accept=XED_FULL)
    body, _ = http(url, headers=headers)
    return json.loads(body)


def _pqs_table(ds: dict) -> str:
    """The dataset's AEP Query Service table name (tags['adobe/pqs/table']) -- the
    normalized SYSTEM name you SELECT ... FROM, distinct from the friendly 'name'
    (e.g. "Acme CJA Order Event Dataset" -> "acme_cja_acme_order_event_dataset").
    '' when the dataset hasn't been assigned one yet."""
    v = (ds.get("tags") or {}).get("adobe/pqs/table")
    if isinstance(v, list):
        return str(v[0]) if v else ""
    return str(v) if v else ""


def _profile_status(ds: dict) -> str:
    """Whether a dataset participates in Unified Profile, from tags.unifiedProfile
    -- needed to know which tables to query for Profile:
        'Profile-enabled'           contributes to the merged profile (enabled:true)
        'snapshot (merge policy X)'  a Profile/Segment Snapshot Export (the merged
                                     UNION you query for whole profiles)
        ''                          not profile-related
    """
    tags = ds.get("tags") or {}
    up = tags.get("unifiedProfile")
    flat = []
    if isinstance(up, list):
        flat = [str(x) for x in up]
    elif isinstance(up, dict):
        for v in up.values():
            flat += [str(x) for x in (v if isinstance(v, list) else [v])]
    elif up:
        flat = [str(up)]
    if any(t.startswith("ups_snapshot_type:") for t in flat):
        mp = next((t.split(":", 1)[1] for t in flat
                   if t.startswith("mergePolicyId:")), "")
        return f"snapshot (merge policy {mp})" if mp else "snapshot export"
    if any(t.startswith("enabled:true") for t in flat):
        return "Profile-enabled"
    return ""


def get_all_datasets(token, conf, sandbox):
    """Page through Catalog /dataSets. Returns (counts, datasets):
        counts    {schema_id: dataset_count}   -- the UI 'DATASETS' column
        datasets  [ {name, table, id, schema_id, profile}, ... ]  -- EVERY dataset

    'table' is the Query Service table name (see _pqs_table); 'profile' flags
    Profile enablement (see _profile_status). Together they let correct SQL be
    formed against the right table -- including the Profile snapshot + its feeds."""
    counts, datasets = {}, []
    headers = aep_headers(token, conf, sandbox)
    start, limit = 0, 100
    while True:
        url = (f"{DATASETS_URL}?limit={limit}&start={start}"
               f"&properties=name,schemaRef,tags")
        body, _ = http(url, headers=headers)
        data = json.loads(body)
        if not isinstance(data, dict) or not data:
            break
        for dsid, ds in data.items():
            if not isinstance(ds, dict):
                continue
            ref = (ds.get("schemaRef") or {}).get("id")
            if ref:
                counts[ref] = counts.get(ref, 0) + 1
            datasets.append({
                "name": ds.get("name") or dsid,
                "table": _pqs_table(ds),
                "id": dsid,
                "schema_id": ref or "",
                "profile": _profile_status(ds),
            })
        if len(data) < limit:
            break
        start += limit
    return counts, datasets


def get_all_descriptors(token, conf, sandbox):
    """GET every descriptor for the sandbox. Returns three dicts keyed by
    (schema_id, lowercased_dot_path):
        identities      -> {"primary": bool, "namespace": str}
        relationships   -> {"dest_id": str, "dest_prop": str, "cardinality": str}
        labels          -> {"title": str, "description": str}
    """
    identities, relationships, labels = {}, {}, {}
    headers = aep_headers(token, conf, sandbox, accept=DESCRIPTOR_ACCEPT)
    body, _ = http(DESCRIPTORS_URL, headers=headers)
    data = json.loads(body)

    # The endpoint returns either a flat list or a dict grouped by @type
    # ({"xdm:descriptorIdentity": [...], "xdm:descriptorOneToOne": [...], ...}),
    # each value a list of full descriptor objects.
    descriptors = []
    if isinstance(data, dict):
        for v in data.values():
            if isinstance(v, list):
                descriptors.extend(x for x in v if isinstance(x, dict))
    elif isinstance(data, list):
        descriptors = [x for x in data if isinstance(x, dict)]

    for d in descriptors:
        dtype = d.get("@type") or ""
        sid = d.get("xdm:sourceSchema") or ""
        if not sid:
            continue

        if dtype == "xdm:descriptorPrimaryKey":
            # sourceProperty is a list of paths; mark each as a primary key.
            raw_props = d.get("xdm:sourceProperty")
            props = raw_props if isinstance(raw_props, list) else [raw_props]
            for p in props:
                prop = _norm_prop(p)
                if not prop:
                    continue
                cur = identities.setdefault((sid, prop),
                                            {"primary": False, "namespace": ""})
                cur["primary"] = True
            continue

        prop = _norm_prop(d.get("xdm:sourceProperty"))
        if not prop:
            continue
        key = (sid, prop)

        if dtype == "xdm:descriptorIdentity":
            cur = identities.setdefault(key, {"primary": False, "namespace": ""})
            cur["primary"] = cur["primary"] or bool(d.get("xdm:isPrimary"))
            cur["namespace"] = (d.get("xdm:namespace")
                                or d.get("xdm:identityNamespace")
                                or cur["namespace"])
        elif dtype in ("xdm:descriptorRelationship", "xdm:descriptorOneToOne"):
            relationships[key] = {
                "dest_id": d.get("xdm:destinationSchema") or "",
                "dest_prop": _norm_prop(d.get("xdm:destinationProperty")),
                "cardinality": d.get("xdm:cardinality") or "1:1",
                "namespace": "",
            }
        elif dtype == "xdm:descriptorReferenceIdentity":
            # A field that references another entity via an identity namespace.
            relationships.setdefault(key, {
                "dest_id": "", "dest_prop": "", "cardinality": "ref",
                "namespace": (d.get("xdm:identityNamespace")
                              or d.get("xdm:namespace") or ""),
            })
        elif dtype == "xdm:alternateDisplayInfo":
            labels[key] = {
                "title": _localized(d.get("xdm:title")),
                "description": _localized(d.get("xdm:description")),
            }
    return identities, relationships, labels


def _norm_prop(p):
    """'/_tenant/person/name/firstName' -> '_tenant.person.name.firstname'
    (leading slash stripped, slashes -> dots, lowercased for matching)."""
    if not isinstance(p, str):
        return ""
    return p.strip("/").replace("/", ".").lower()


def _localized(v):
    """alternateDisplayInfo titles/descriptions can be a plain string or a
    locale map ({'en_us': '...'}). Pull a single readable string out."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, dict):
        for loc in ("en_us", "en_US", "en"):
            if loc in v:
                return _localized(v[loc])
        for val in v.values():
            return _localized(val)
    return str(v)


# ----------------------------------------------------------------------------
# Field flattening
# ----------------------------------------------------------------------------
def flatten_fields(node, prefix="", depth=0):
    """Walk a resolved xed-full 'properties' tree to leaf dot-notation paths.
    Returns [(path, data_type, title, required_bool), ...], where 'title' is the
    field's display name carried in the schema itself (blank if absent).
    Array-of-object paths get a '[]' marker (e.g. orders[].id)."""
    rows = []
    if depth > 18:
        return rows
    props = node.get("properties")
    if not isinstance(props, dict):
        return rows
    required = set(node.get("required") or [])
    for key, sub in props.items():
        if not isinstance(sub, dict):
            continue
        path = f"{prefix}.{key}" if prefix else key
        req = key in required
        title = sub.get("title") or ""
        t = sub.get("type")
        if t == "object" and isinstance(sub.get("properties"), dict):
            rows.extend(flatten_fields(sub, path, depth + 1))
        elif t == "array":
            items = sub.get("items") or {}
            if (isinstance(items, dict) and items.get("type") == "object"
                    and isinstance(items.get("properties"), dict)):
                rows.extend(flatten_fields(items, path + "[]", depth + 1))
            else:
                itype = (items.get("meta:xdmType") or items.get("format")
                         or items.get("type") or "any")
                rows.append((path + "[]", f"array<{itype}>", title, req))
        else:
            disp = sub.get("meta:xdmType") or sub.get("format") or t or "object"
            rows.append((path, disp, title, req))
    return rows


# ----------------------------------------------------------------------------
# Data dictionary (phase 2): sample real records, tally coverage + top values
#
# AEP has no API that returns a value distribution for a single field, and
# Query Service GROUP BY per field is too slow. Instead we download a sample of
# the schema's actual ingested records (Snappy-Parquet) via the Data Access
# API -- ONE download covers every field at once -- and tally locally.
# ----------------------------------------------------------------------------
# Synthetic / load-test datasets pollute example values and tend to carry huge
# files; skip them when sampling.
_SKIP_DS_RE = re.compile(r"perf[\s_-]?test|load[\s_-]?test", re.I)


def get_dataset_ids_by_schema(token, conf, sandbox):
    """{schema $id: [datasetId, ...]} -- which datasets feed each schema
    (synthetic perf/load-test datasets excluded)."""
    out = {}
    headers = aep_headers(token, conf, sandbox)
    start, limit = 0, 100
    while True:
        url = f"{DATASETS_URL}?limit={limit}&start={start}&properties=name,schemaRef"
        data = json.loads(http(url, headers=headers)[0])
        if not isinstance(data, dict) or not data:
            break
        for dsid, v in data.items():
            if not isinstance(v, dict):
                continue
            if _SKIP_DS_RE.search(v.get("name") or ""):
                continue
            ref = (v.get("schemaRef") or {}).get("id")
            if ref:
                out.setdefault(ref, []).append(dsid)
        if len(data) < limit:
            break
        start += limit
    return out


# ----------------------------------------------------------------------------
# Profile coverage (v3.1): default merge policy -> its snapshot export dataset
# ----------------------------------------------------------------------------
def get_default_merge_policy(token, conf, sandbox):
    """The org's DEFAULT Profile merge policy in this sandbox, or None.

    GET /data/core/ups/config/mergePolicies (cursor-paged at _page.next; results
    live under 'children' or 'mergePolicies'). Returns the policy dict flagged
    default==True, preferring one whose schema.name is _xdm.context.profile."""
    headers = aep_headers(token, conf, sandbox)
    pols, start, pages = [], None, 0
    while pages < 50:
        pages += 1
        params = {"limit": 100}
        if start:
            params["start"] = start
        url = f"{UPS_MERGE_POLICIES_URL}?{urllib.parse.urlencode(params)}"
        data = json.loads(http(url, headers=headers)[0]) or {}
        batch = []
        for key in ("children", "mergePolicies"):
            if isinstance(data.get(key), list):
                batch = data[key]
                break
        pols.extend(p for p in batch if isinstance(p, dict))
        nxt = (data.get("_page") or {}).get("next")
        if not batch or len(batch) < 100 or not nxt:
            break
        start = nxt

    defaults = [p for p in pols if p.get("default") is True]
    if not defaults:
        logger.warning(f"  profile: no merge policy is flagged default in "
                       f"{sandbox} ({len(pols)} policies seen).")
        return None
    # Prefer the one whose schema is the Profile class (a default Identity-map
    # policy can also exist); otherwise the first default.
    for p in defaults:
        if (p.get("schema") or {}).get("name") == "_xdm.context.profile":
            return p
    return defaults[0]


def _ups_snapshot_tags(ds):
    """The dataset's tags.unifiedProfile entries as a flat list of strings.
    Handles the dict-of-lists / flat-list / single-value shapes AEP uses."""
    tags = ds.get("tags") or {}
    up = tags.get("unifiedProfile")
    if up is None:
        return []
    if isinstance(up, list):
        return [str(x) for x in up]
    if isinstance(up, dict):
        out = []
        for v in up.values():
            out.extend(v if isinstance(v, list) else [v])
        return [str(x) for x in out]
    return [str(up)]


def find_profile_snapshot_dataset(token, conf, sandbox, merge_policy_id):
    """(dataset_id, name) of the Profile Snapshot Export dataset for the given
    merge policy, or None. A snapshot export dataset has schemaRef.id ==
    profile__union and a tags.unifiedProfile list carrying a ups_snapshot_type:*
    entry; the owning merge policy id appears among those same tag strings."""
    headers = aep_headers(token, conf, sandbox)
    matches = []
    start, limit = 0, 100
    while True:
        url = (f"{DATASETS_URL}?limit={limit}&start={start}"
               f"&properties=name,schemaRef,tags")
        data = json.loads(http(url, headers=headers)[0])
        if not isinstance(data, dict) or not data:
            break
        for dsid, v in data.items():
            if not isinstance(v, dict):
                continue
            if (v.get("schemaRef") or {}).get("id") != PROFILE_UNION:
                continue
            ups = _ups_snapshot_tags(v)
            if not any(t.startswith("ups_snapshot_type:") for t in ups):
                continue
            if merge_policy_id and not any(merge_policy_id in t for t in ups):
                continue
            matches.append((dsid, v.get("name") or dsid))
        if len(data) < limit:
            break
        start += limit

    if not matches:
        return None
    if len(matches) > 1:
        logger.warning(f"  profile: {len(matches)} snapshot dataset(s) match the "
                       f"default merge policy; using {matches[0][1]!r}.")
    return matches[0]


def _batch_record_count(b):
    """Records in a batch from its catalog metrics. -1 when unknown."""
    m = b.get("metrics") or {}
    for v in (m.get("outputRecordCount"), m.get("inputRecordCount"),
              b.get("recordCount"), m.get("recordCount")):
        if v is not None:
            try:
                return int(v)
            except (TypeError, ValueError):
                return -1
    return -1


def _is_consolidation_batch(b):
    """True for a batch AEP's own compaction job produced by merging others.

    Catalog advertises these as ordinary status=success batches carrying the
    combined record count -- so they sort to the top of our richest-first list
    -- but the Data Access API refuses them outright:

        DTAC-4000  "The queried batch is a consolidation batch, which is not
                    supported by Data Access API now."

    Nothing is lost by skipping them: the batches they absorbed are listed in
    their own replay.predecessors and Catalog still serves those individually.
    """
    tags = b.get("tags") or {}
    workflow = tags.get("acp_workflow") or []
    if isinstance(workflow, str):
        workflow = [workflow]
    if any("consolidation" in str(w).lower() for w in workflow):
        return True
    return "compaction" in str(b.get("createdClient") or "").lower()


def _read_parquet_rows(raw_bytes, limit):
    """Parse only up to `limit` rows (row-group streaming) so a million-row
    file isn't fully materialised into memory just to sample a handful."""
    import pyarrow as pa
    import pyarrow.parquet as pq
    pf = pq.ParquetFile(pa.BufferReader(raw_bytes))
    out = []
    for batch in pf.iter_batches(batch_size=max(1, min(limit, 1024))):
        out.extend(batch.to_pylist())
        if len(out) >= limit:
            break
    return out[:limit]


def _get_json_retry(url, headers, timeout, attempts=3, pause=8, label=""):
    """GET + parse JSON, retrying on gateway 5xx / read timeouts. The Data Access
    file-manifest endpoint cold-504s on the FIRST access to a huge Profile
    Snapshot Export file (it warms the manifest server-side, exceeding the ~60s
    gateway limit); the next attempt hits the warm cache and returns in ~1s."""
    last = None
    for i in range(attempts):
        try:
            return json.loads(http(url, headers=headers, timeout=timeout)[0])
        except urllib.error.HTTPError as e:
            last = e
            if e.code not in (502, 503, 504) or i == attempts - 1:
                raise
        except (urllib.error.URLError, TimeoutError, OSError) as e:
            last = e
            if i == attempts - 1:
                raise
        logger.info(f"      {ANSI['dim']}{label or 'request'} retry "
                    f"{i + 2}/{attempts} (warming after "
                    f"{type(last).__name__}){ANSI['reset']}")
        time.sleep(pause)
    raise last


def _list_batch_parquet_files(headers, bid, max_entries=40, meta_timeout=30,
                              meta_attempts=3, meta_pause=8):
    """Enumerate a batch's physical .parquet files as [(fid, name, length)].
    `length` is the file size in bytes (None if AEP omits it). Bounded to
    max_entries dataSetFileId entries so a snapshot with thousands of partitions
    isn't fully walked just to find a small one. meta_timeout is raised for huge
    Profile Snapshot Export batches whose file listings are slow to return."""
    out = []
    # Page the batch-files listing with an explicit limit. The per-file manifest
    # call (/files/{fid}) is left bare: a ?limit on THAT endpoint forces a full
    # enumeration that cold-504s, whereas the bare call returns the first ~100
    # physical files fast. Both are retried to ride out the cold-start 504.
    url = f"{EXPORT_URL}/batches/{bid}/files?limit={max_entries}"
    files = _get_json_retry(url, headers, meta_timeout, attempts=meta_attempts,
                            pause=meta_pause, label="batch files")
    entries = (files.get("data") if isinstance(files, dict) else files) or []
    for fe in entries[:max_entries]:
        fid = fe.get("dataSetFileId")
        if not fid:
            continue
        meta = _get_json_retry(f"{EXPORT_URL}/files/{fid}", headers, meta_timeout,
                               attempts=meta_attempts, pause=meta_pause,
                               label=f"file manifest {fid[:24]}")
        for phys in (meta.get("data") if isinstance(meta, dict) else meta) or []:
            name = phys.get("name") or ""
            if not name.endswith(".parquet"):
                continue
            length = phys.get("length")
            try:
                length = int(length) if length is not None else None
            except (TypeError, ValueError):
                length = None
            out.append((fid, name, length))
    return out


def _download_batch_rows(headers, bid, need, smallest_first=False,
                         file_timeout=75):
    """Download a SUCCESS batch's Parquet file(s) and return (rows, bytes). Logs
    each physical file it streams (name, size, rows parsed).

    smallest_first: for huge datasets (Profile Snapshot Export -- tens of
    millions of rows split into many partitions), download the SMALLEST partition
    files first. http() pulls a whole file into memory before parsing, so a giant
    partition would time out; the smallest one yields the sample cheaply."""
    rows, total_bytes = [], 0
    # Snapshot listings are large/slow; give the metadata calls the same patience
    # as the file downloads when we're in the smallest-first (snapshot) path.
    # A 504 on the SNAPSHOT manifest is expensive far beyond this one batch: the
    # failure is cached and every remaining Profile-class schema is then skipped
    # as coverage MISSING. Worth being markedly more patient on that path.
    cand = _list_batch_parquet_files(
        headers, bid,
        meta_timeout=file_timeout if smallest_first else 30,
        meta_attempts=6 if smallest_first else 3,
        meta_pause=20 if smallest_first else 8)
    if smallest_first:
        # Unknown sizes (length is None) sort last so we prefer a known-small one.
        cand.sort(key=lambda c: (c[2] is None, c[2] if c[2] is not None else 0))
        sizes = [f"{c[2]/1024/1024:.0f}MB" if c[2] else "?" for c in cand[:5]]
        logger.info(f"      {len(cand)} partition file(s) listed; smallest: "
                    f"{', '.join(sizes)}")
    for fid, name, length in cand:
        if len(rows) >= need:
            break
        t0 = time.perf_counter()
        # Event data files run to tens of MB; snapshot partitions far more.
        # Dead batches are bounded by the consecutive-failure bail-out upstream.
        raw = http(f"{EXPORT_URL}/files/{fid}?path={urllib.parse.quote(name)}",
                   headers=headers, timeout=file_timeout)[0]
        total_bytes += len(raw)
        parsed = _read_parquet_rows(raw, need - len(rows))
        rows.extend(parsed)
        logger.info(f"      file {name[:46]} "
                    f"{ANSI['dim']}{len(raw)/1024:.0f} KB -> "
                    f"{len(parsed)} rows in "
                    f"{time.perf_counter() - t0:.1f}s{ANSI['reset']}")
    return rows[:need], total_bytes


def sample_schema_rows(token, conf, sandbox, dsids, target, max_batches=12,
                       smallest_first=False, file_timeout=75):
    """Pull up to `target` real records from a schema's datasets. The quota is
    spread ACROSS the datasets (newest SUCCESS batches first within each), so a
    schema fed by several datasets -- each populating different fields -- gets
    representative coverage rather than draining one. Returns nested row dicts.

    smallest_first/file_timeout are forwarded to _download_batch_rows -- set for
    the Profile Snapshot Export path (smallest partition, longer timeout)."""
    headers = aep_headers(token, conf, sandbox)
    if not dsids:
        return [], {"available": 0, "failed": 0, "empty_reads": 0,
                    "list_failed": 0}
    per_ds = max(1, -(-target // len(dsids)))      # ceil(target / n)
    rows, total_bytes, total_batches = [], 0, 0
    available = failed = empty_reads = 0     # diagnostics: empty vs unreadable
    list_failed = 0     # datasets whose batch LIST never returned (504/timeout):
                        # we learned nothing about them, so 0 rows != empty
    t_start = time.perf_counter()
    logger.info(f"    sampling across {len(dsids)} dataset(s), "
                f"~{per_ds} rows each (target {target}); streaming Parquet "
                f"through memory...")
    for di, dsid in enumerate(dsids, 1):
        if len(rows) >= target:
            break
        need = min(per_ds, target - len(rows))
        # status=success is filtered SERVER-side, and it matters far beyond
        # tidiness: on a continuously-landing feed (an Adobe Analytics one, say)
        # asking Catalog to sort every batch it has ever held blows the gateway
        # timeout, and the dataset reads as permanently unavailable. Filtering
        # first turns a reliable 504 into a ~16s response. We still re-check
        # status below, since this only narrows what we have to sort.
        # ...and when even that times out, ask for FEWER batches. Adobe says as
        # much in the 504 body ("This request timed out on database, please
        # optimize your request"): it is the PAGE SIZE that decides whether the
        # query returns, not the sort. Measured on the heaviest datasets --
        # limit=20 dies, limit=5 answers in ~5s, limit=1 always answers. Fewer
        # batches means a smaller sample pool, which is a real cost, but a
        # narrower sample beats no coverage at all. Earlier rungs retry only
        # twice: dropping to a smaller page beats waiting on a page size this
        # dataset has already outgrown.
        base = f"{CATALOG_BATCHES_URL}?dataSet={dsid}&status=success&orderBy=desc:created"
        shapes = [(f"{base}&limit=20", "limit 20", 2),
                  (f"{base}&limit=5", "limit 5", 2),
                  (f"{base}&limit=1", "limit 1", 3)]
        batches, last_err = None, None
        for url, shape, tries in shapes:
            try:
                # Retried like the file manifests: Catalog gateway-504s under
                # load, and a bare failure here used to read as "no batches"
                # -> false EMPTY.
                batches = _get_json_retry(url, headers, 60, attempts=tries,
                                          label=f"batch list {dsid[:24]} ({shape})")
                if shape != shapes[0][1]:
                    logger.info(f"      {ANSI['dim']}{shape} succeeded -- sample "
                                f"pool for this dataset is narrower than "
                                f"usual{ANSI['reset']}")
                break
            except Exception as e:
                last_err = e
                if shape != shapes[-1][1]:
                    logger.info(f"      {ANSI['dim']}batch list {shape} failed "
                                f"({type(e).__name__}); asking for fewer "
                                f"batches{ANSI['reset']}")
        if batches is None:
            list_failed += 1
            logger.warning(f"    dataset {di}/{len(dsids)} {dsid}: "
                           f"batch list failed ({last_err}); skipping -- this "
                           f"dataset's contents are UNKNOWN, not empty.")
            continue
        if not isinstance(batches, dict):
            list_failed += 1        # malformed response: also tells us nothing
            continue
        # Success batches that actually carry records (skip empty/control
        # partitions and still-loading batches), RICHEST first -- one big batch
        # fills the per-dataset quota in a single download instead of churning
        # through dozens of tiny 1-row batches.
        succ_items = [(bid, b) for bid, b in batches.items()
                      if isinstance(b, dict) and b.get("status") == "success"
                      and _batch_record_count(b) != 0
                      and not _is_consolidation_batch(b)]
        succ_items.sort(key=lambda kv: _batch_record_count(kv[1]), reverse=True)
        succ = [bid for bid, _ in succ_items]
        available += len(succ)
        logger.info(f"    dataset {di}/{len(dsids)} {dsid}: "
                    f"{len(succ)} non-empty success batch(es) available, "
                    f"want {need} rows")
        got, batches_seen, consec_fail = 0, 0, 0
        for bid in succ:
            if got >= need or batches_seen >= max_batches:
                break
            if consec_fail >= 3:
                logger.info(f"      {ANSI['dim']}3 consecutive failures in this "
                            f"dataset -- moving on.{ANSI['reset']}")
                break
            batches_seen += 1
            total_batches += 1
            try:
                chunk, nbytes = _download_batch_rows(
                    headers, bid, need - got,
                    smallest_first=smallest_first, file_timeout=file_timeout)
            except Exception as e:
                consec_fail += 1
                failed += 1
                logger.warning(f"      batch {bid[:24]}: skipped ({e})")
                continue
            total_bytes += nbytes
            if not chunk:
                consec_fail += 1
                empty_reads += 1
                continue
            consec_fail = 0
            rows.extend(chunk)
            got += len(chunk)
            logger.info(f"      batch {bid[:24]} -> {len(chunk)} rows "
                        f"{ANSI['dim']}(dataset {got}/{need}, "
                        f"overall {len(rows)}/{target}){ANSI['reset']}")
    logger.info(f"    sampled {len(rows)} rows from {total_batches} batch(es), "
                f"{total_bytes/1024/1024:.1f} MB streamed, "
                f"{time.perf_counter() - t_start:.1f}s elapsed.")
    return rows[:target], {"available": available, "failed": failed,
                           "empty_reads": empty_reads,
                           "list_failed": list_failed}


def _extract_values(row, path):
    """All scalar/array leaf values at a dot-notation path (handles nested
    structs and '[]' array markers). Returns a flat list (empty if absent)."""
    cur = [row]
    for seg in path.split("."):
        arr = seg.endswith("[]")
        key = seg[:-2] if arr else seg
        nxt = []
        for c in cur:
            if not isinstance(c, dict):
                continue
            v = c.get(key)
            if v is None:
                continue
            if arr:
                if isinstance(v, list):
                    nxt.extend(x for x in v if x is not None)
            else:
                nxt.append(v)
        cur = nxt
        if not cur:
            break
    return cur


def _fmt_val(v):
    return str(v).replace("|", "/").replace("\n", " ").strip()[:40]


def build_data_dict(rows, fields, top_n=5):
    """{field_path: {"coverage": int_pct, "top": "value(count) | ..."}}.
    Coverage = % of sampled rows where the field is populated."""
    from collections import Counter
    n = len(rows) or 1
    out = {}
    for field in fields:
        fpath = field[0]
        present = 0
        counter = Counter()
        for r in rows:
            vals = _extract_values(r, fpath)
            if vals:
                present += 1
            for v in vals:
                if isinstance(v, (str, int, float, bool)):
                    counter[_fmt_val(v)] += 1
        out[fpath] = {
            "coverage": round(100 * present / n),
            "top": " | ".join(f"{val}({cnt})"
                               for val, cnt in counter.most_common(top_n)),
        }
    return out


# ----------------------------------------------------------------------------
# Schema metadata helpers
# ----------------------------------------------------------------------------
def class_name(class_id: str) -> str:
    if not class_id:
        return "?"
    if class_id in KNOWN_CLASSES:
        return KNOWN_CLASSES[class_id]
    tail = class_id.rstrip("/").rsplit("/", 1)[-1]
    return tail.replace("-", " ").replace("_", " ").title()


def fmt_epoch_ms(ms) -> str:
    try:
        return datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc).strftime(
            "%Y-%m-%d %H:%M")
    except Exception:
        return ""


def last_modified(raw: dict) -> str:
    meta = raw.get("meta:registryMetadata") or {}
    for k in ("repo:lastModifiedDate", "repo:lastModified", "repo:updatedDate"):
        if meta.get(k):
            return fmt_epoch_ms(meta[k])
    return ""


def extends_refs(raw: dict):
    """All class / field-group ids a schema references -- meta:extends plus any
    allOf $refs present in the unresolved resource."""
    refs = list(raw.get("meta:extends") or [])
    for entry in raw.get("allOf") or []:
        if isinstance(entry, dict) and entry.get("$ref"):
            refs.append(entry["$ref"])
    return refs


def is_adhoc(raw: dict) -> bool:
    cls = raw.get("meta:class") or ""
    if cls == ADHOC_CLASS:
        return True
    tail = cls.rstrip("/").rsplit("/", 1)[-1]
    if ADHOC_CLASS_HEX.match(tail):       # generated per-schema class
        return True
    title = (raw.get("title") or "").strip().lower()
    if any(title.startswith(p) for p in ADHOC_TITLE_PREFIXES):
        return True
    return any(s in title for s in ADHOC_TITLE_SUBSTRINGS)


def is_ajo(raw: dict) -> bool:
    title = (raw.get("title") or "").strip().lower()
    if any(title.startswith(p) for p in AJO_TITLE_PREFIXES):
        return True
    if any(s in title for s in AJO_TITLE_SUBSTRINGS):
        return True
    refs = " ".join(extends_refs(raw)).lower()
    return any(s in refs for s in AJO_EXTENDS_SUBSTRINGS)


def is_system(raw: dict) -> bool:
    cls = (raw.get("meta:class") or "").rstrip("/").rsplit("/", 1)[-1].lower()
    if cls in SYSTEM_CLASS_TAILS:
        return True
    title = (raw.get("title") or "").strip().lower()
    if title in SYSTEM_TITLE_EXACT:
        return True
    return any(s in title for s in SYSTEM_TITLE_SUBSTRINGS)


def is_test(raw: dict) -> bool:
    return bool(_TEST_RE.search(raw.get("title") or ""))


# ----------------------------------------------------------------------------
# Credential menu + sandbox picker
# ----------------------------------------------------------------------------
def pretty_name(stem: str) -> str:
    return stem.replace("-", " ").replace("_", " ").title()


def client_label(conf: dict) -> str:
    """Human client name for workbook titles, or "" when none is configured.

    ONLY the explicit "client" key in the credential record is used. It used to
    fall back to the credential's own service name, which put the name of a key
    in our own vault at the top of a document that goes out to the client.
    That name identifies whose credential read the sandbox,
    which is our own key-management artefact and nobody else's business; worse,
    it reads like the subject of the report. Where no client name is configured
    the workbook now identifies itself by SANDBOX instead (see write_xlsx),
    which is what the filename has said since v3.3 and is what the reader
    actually needs to know.
    """
    c = (conf.get("client") or "").strip()
    return c.replace("-", " ").replace("_", " ").strip().title()


def short_type(raw: str) -> str:
    t = (raw or "").lower()
    if "prod" in t:
        return "prod"
    if "dev" in t:
        return "dev"
    return raw or "?"


def menu(services):
    """House-style numbered picker over keyring service names. Returns the
    chosen service name string, or None."""
    print()
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Credential bank{ANSI['reset']}  "
          f"{ANSI['dim']}(OS keyring vault){ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 60 + ANSI["reset"])
    for i, name in enumerate(services, 1):
        print(f"  {ANSI['bold']}[{i}]{ANSI['reset']} "
              f"{ANSI['yellow']}{pretty_name(name):<20}{ANSI['reset']} "
              f"{ANSI['dim']}{name}{ANSI['reset']}")
    print(bar)
    raw = input(
        f"\nChoose a credential set by number "
        f"({ANSI['cyan']}1{ANSI['reset']}-{ANSI['cyan']}{len(services)}{ANSI['reset']}), "
        "blank to quit: "
    ).strip()
    if not raw or not raw.isdigit():
        return None
    idx = int(raw)
    if 1 <= idx <= len(services):
        return services[idx - 1]
    logger.warning(f"Choice {idx} out of range.")
    return None


def _default_prod(sandboxes):
    """The sandbox to use when the user just hits Enter: one literally named
    'prod', else the first production-type sandbox, else the first sandbox."""
    for sb in sandboxes:
        if sb.get("name") == "prod":
            return sb
    for sb in sandboxes:
        if short_type(sb.get("type", "")) == "prod":
            return sb
    return sandboxes[0] if sandboxes else None


def pick_sandboxes(sandboxes):
    """Prompt for which sandbox(es) to read. Enter = prod (default).
    Accepts numbers ('1', '1,3'), 'all', or Enter for prod."""
    default = _default_prod(sandboxes)
    print()
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Sandboxes visible to this credential{ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 60 + ANSI["reset"])
    for i, sb in enumerate(sandboxes, 1):
        name = sb.get("name", "?")
        title = sb.get("title") or name
        env = short_type(sb.get("type", ""))
        star = " (default)" if sb is default else ""
        print(f"  {ANSI['bold']}[{i:>2}]{ANSI['reset']} "
              f"{ANSI['yellow']}{title:<22}{ANSI['reset']} "
              f"{ANSI['dim']}{name:<16}{ANSI['reset']} {env}{star}")
    print(bar)
    dlabel = default.get("name", "?") if default else "?"
    raw = input(
        f"\nPick sandbox(es) "
        f"({ANSI['cyan']}1{ANSI['reset']}, {ANSI['cyan']}1,3{ANSI['reset']}, "
        f"{ANSI['cyan']}all{ANSI['reset']}, or Enter for "
        f"{ANSI['green']}{dlabel}{ANSI['reset']}): "
    ).strip()
    if not raw:
        return [default] if default else []
    if raw.lower() == "all":
        return list(sandboxes)
    chosen = []
    for tok in raw.replace(",", " ").split():
        if tok.isdigit() and 1 <= int(tok) <= len(sandboxes):
            sb = sandboxes[int(tok) - 1]
            if sb not in chosen:
                chosen.append(sb)
        else:
            logger.warning(f"Ignoring invalid choice: {tok}")
    return chosen


def resolve_sandboxes_arg(arg: str, sandboxes):
    """Map a --sandbox=... value (names, comma-separated, or 'all') to dicts."""
    if arg.lower() == "all":
        return list(sandboxes)
    wanted = [t for t in arg.replace(",", " ").split() if t]
    by_name = {sb.get("name"): sb for sb in sandboxes}
    out = []
    for w in wanted:
        sb = by_name.get(w)
        if sb and sb not in out:
            out.append(sb)
        elif not sb:
            logger.warning(f"--sandbox: no sandbox named {w!r}; ignoring.")
    return out


# ----------------------------------------------------------------------------
# Per-sandbox collection + filtering
# ----------------------------------------------------------------------------
def collect_sandbox(token, conf, sb):
    """Read one sandbox, filter, and resolve fields for kept schemas.

    Returns dict:
      title, name, env
      verdicts  -> list of (title, class_name, datasets, last_mod, status)
                   status is KEEP / DROP:no-dataset / DROP:adhoc / DROP:ajo
      kept      -> list of kept schema dicts (see build below)
      labels_n  -> count of alternateDisplayInfo (dual) labels in the sandbox
      stats     -> {total, kept, no_dataset, adhoc, ajo, fields, relationships}
    """
    name = sb.get("name", "?")
    title = sb.get("title") or name
    env = short_type(sb.get("type", ""))

    raws = get_all_schemas(token, conf, name)
    ds_counts, ds_all = get_all_datasets(token, conf, name)
    identities, relationships, labels = get_all_descriptors(token, conf, name)

    # id -> title, for resolving relationship destinations to readable names.
    id_to_title = {r.get("$id"): (r.get("title") or r.get("$id")) for r in raws}

    # SQL table names: every dataset's (friendly name, system table name, id)
    # grouped by the schema it feeds, so each schema knows which table(s) to query.
    tables_by_schema = {}
    for d in ds_all:
        d["schema_title"] = (id_to_title.get(d["schema_id"])
                             or _id_tail(d["schema_id"]) or "(no schema)")
        if d["schema_id"]:
            tables_by_schema.setdefault(d["schema_id"], []).append(
                (d["name"], d["table"], d["id"]))
    for v in tables_by_schema.values():
        v.sort(key=lambda t: t[0].lower())

    verdicts = []
    kept = []
    stats = {"total": len(raws), "kept": 0, "no_dataset": 0, "adhoc": 0,
             "audience": 0, "ajo": 0, "system": 0, "test": 0, "fields": 0,
             "relationships": 0}

    # Sort by title for a stable, readable listing.
    for raw in sorted(raws, key=lambda r: (r.get("title") or "").lower()):
        sid = raw.get("$id") or ""
        stitle = raw.get("title") or raw.get("meta:altId") or sid or "?"
        cls = class_name(raw.get("meta:class") or "")
        n_ds = ds_counts.get(sid, 0)
        lmod = last_modified(raw)

        if is_adhoc(raw):
            status = "DROP:adhoc"
            stats["adhoc"] += 1
            if stitle.strip().lower().startswith("schema for audience"):
                stats["audience"] += 1
        elif is_ajo(raw):
            status = "DROP:ajo"
            stats["ajo"] += 1
        elif is_system(raw):
            status = "DROP:system"
            stats["system"] += 1
        elif is_test(raw):
            status = "DROP:test"
            stats["test"] += 1
        elif n_ds < 1:
            status = "DROP:no-dataset"
            stats["no_dataset"] += 1
        else:
            status = "KEEP"

        verdicts.append((stitle, cls, n_ds, lmod, status))
        if status != "KEEP":
            continue

        # --- Resolve the kept schema's fields + descriptor joins -------------
        ref = raw.get("meta:altId") or sid
        try:
            full = get_full_schema(token, conf, name, ref)
        except Exception as e:
            logger.warning(f"  {stitle}: full-schema fetch failed ({e}); "
                           "fields will be empty.")
            full = {}
        fields = flatten_fields(full)

        field_rows = []
        n_identities = n_rels = n_labels = 0
        for path, dtype, title, req in fields:
            mkey = (sid, path.replace("[]", "").lower())
            ident = identities.get(mkey)
            rel = relationships.get(mkey)
            lab = labels.get(mkey)
            if ident:
                n_identities += 1
                id_disp = ("PRIMARY" if ident["primary"] else "identity")
                if ident["namespace"]:
                    id_disp += f" ({ident['namespace']})"
            else:
                id_disp = ""
            if rel:
                n_rels += 1
                if rel["dest_id"]:
                    dest = id_to_title.get(rel["dest_id"]) or _id_tail(rel["dest_id"])
                    rel_disp = dest + (f".{rel['dest_prop']}" if rel["dest_prop"] else "")
                    if rel["cardinality"]:
                        rel_disp += f"  [{rel['cardinality']}]"
                elif rel.get("namespace"):
                    rel_disp = f"ref-identity ({rel['namespace']})"
                else:
                    rel_disp = "ref-identity"
            else:
                rel_disp = ""
            # Display name lives on the field itself; fall back to an
            # alternateDisplayInfo descriptor only where the field has none.
            friendly = title or (lab["title"] if lab else "")
            if friendly:
                n_labels += 1
            field_rows.append((path, dtype, friendly, "Y" if req else "",
                               id_disp, rel_disp))

        stats["kept"] += 1
        stats["fields"] += len(field_rows)
        stats["relationships"] += n_rels

        kept.append({
            "title": stitle,
            "class": cls,
            "meta_class": raw.get("meta:class") or "",
            "datasets": n_ds,
            "tables": tables_by_schema.get(sid, []),
            "last_mod": lmod,
            "id": sid,
            "fields": field_rows,
            "n_fields": len(field_rows),
            "n_identities": n_identities,
            "n_relationships": n_rels,
            "n_labels": n_labels,
        })

    return {
        "title": title, "name": name, "env": env,
        "verdicts": verdicts, "kept": kept, "datasets": ds_all,
        "labels_n": len(labels), "stats": stats,
    }


def _id_tail(sid: str) -> str:
    return sid.rsplit("/", 1)[-1] if sid else ""


# ----------------------------------------------------------------------------
# Console output
# ----------------------------------------------------------------------------
STATUS_COLOR = {
    "KEEP": ANSI["green"],
    "DROP:no-dataset": ANSI["dim"],
    "DROP:adhoc": ANSI["yellow"],
    "DROP:ajo": ANSI["magenta"],
    "DROP:system": ANSI["blue"],
    "DROP:test": ANSI["cyan"],
}


def print_sandbox(result):
    title = result["title"]
    bar = ANSI["cyan"] + "=" * 78 + ANSI["reset"]
    print()
    print(bar)
    print(f"  {ANSI['bold']}{title}{ANSI['reset']}  "
          f"{ANSI['dim']}({result['name']}, {result['env']}){ANSI['reset']}")
    print(bar)

    header = (f"{'Schema':<44} {'Class':<22} {'DS':>3} "
              f"{'Last Modified':<16} Status")
    print(ANSI["bold"] + header + ANSI["reset"])
    print(ANSI["dim"] + "-" * 100 + ANSI["reset"])
    for stitle, cls, n_ds, lmod, status in result["verdicts"]:
        color = STATUS_COLOR.get(status, "")
        disp_title = stitle if len(stitle) <= 44 else stitle[:41] + "..."
        disp_cls = cls if len(cls) <= 22 else cls[:19] + "..."
        ds_disp = str(n_ds) if n_ds else "-"
        print(f"{disp_title:<44} {ANSI['dim']}{disp_cls:<22}{ANSI['reset']} "
              f"{ds_disp:>3} {lmod:<16} {color}{status}{ANSI['reset']}")

    s = result["stats"]
    print(ANSI["dim"] + "-" * 100 + ANSI["reset"])
    print(f"  {ANSI['bold']}{s['total']}{ANSI['reset']} schemas seen  ->  "
          f"{ANSI['green']}{s['kept']} KEEP{ANSI['reset']}  "
          f"{ANSI['dim']}({s['adhoc']} adhoc, {s['ajo']} ajo, "
          f"{s['system']} system, {s['test']} test, "
          f"{s['no_dataset']} no-dataset dropped){ANSI['reset']}")
    print(f"  {ANSI['dim']}Of the adhoc, {ANSI['reset']}{s['audience']}"
          f"{ANSI['dim']} were auto-created 'Schema for audience...' schemas "
          f"-- filtered out.{ANSI['reset']}")
    dual = result["labels_n"]
    dcolor = ANSI["green"] if dual else ANSI["yellow"]
    print(f"  Dual labels (alternateDisplayInfo) in sandbox: "
          f"{dcolor}{dual}{ANSI['reset']}"
          f"{'' if dual else '  -- this tenant has none'}")
    print(f"  Kept fields: {s['fields']}   relationships: {s['relationships']}")


# ----------------------------------------------------------------------------
# XLSX output
# ----------------------------------------------------------------------------
# Columns for the Schemas index tab (one row per kept schema).
SCHEMA_CSV_COLUMNS = ["Sandbox", "Schema", "Class", "Datasets",
                      "SQL table name(s)", "Fields", "Identities",
                      "Relationships", "Dual-labelled fields", "Last Modified",
                      "Schema $id"]


# Per-schema tab: one field per row (the schema's own metadata lives in the
# tab's title block, so it isn't repeated on every row).
SCHEMA_FIELD_COLUMNS = ["Field (dot notation)", "Data Type", "Friendly Name",
                        "Required", "Identity", "Relationship -> target"]
_HEADER_BG = "1F4E78"


def _safe_sheet_name(name: str, used: set) -> str:
    s = re.sub(r"[\[\]\:\*\?\/\\]", "-", name).strip()[:31] or "sheet"
    base, i = s, 2
    while s.lower() in used:
        suffix = f"-{i}"
        s = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(s.lower())
    return s


def _sheet_label(title: str) -> str:
    """A concise, readable tab name from a schema title: drop the common
    'Acme ' prefix and ' Schema' suffix so the 31-char tab name carries the
    distinctive part. The full title stays in the tab's title block + index."""
    t = (title or "").strip()
    low = t.lower()
    if low.startswith("acme "):
        t = t[6:]
    if t.lower().endswith(" schema"):
        t = t[:-7]
    return t.strip() or (title or "").strip() or "schema"


def _sql_table_names(k) -> str:
    """Comma-joined Query Service table (system) names for a kept schema's
    datasets -- the FROM targets for SQL, blank if none assigned yet."""
    return ", ".join(sys_nm for _, sys_nm, _ in (k.get("tables") or []) if sys_nm)


def _coverage_status(k):
    """(code, label) describing whether --data-dict coverage is trustworthy for a
    kept schema. So a partial/missing dictionary is never mistaken for gospel:
        ok          sampled fine
        partial     sampled, but some batches failed -- coverage understated
        unreadable  data EXISTS but couldn't be sampled (504/timeout) -- MISSING
        empty       genuinely no ingested records
        not_sampled out of --data-dict scope, or no dataset to sample
    """
    if k.get("datadict"):
        st = k.get("dd_sample_stats") or {}
        if st.get("failed") or st.get("empty_reads") or st.get("list_failed"):
            lf = st.get("list_failed", 0)
            return ("partial", f"PARTIAL -- {st.get('failed', 0)} batch(es) failed "
                    f"to read"
                    + (f"; {lf} dataset(s) never listed (contents unknown)"
                       if lf else "")
                    + "; coverage is understated")
        return ("ok", "")
    reason = k.get("dd_empty_reason")
    if reason == "unreadable":
        return ("unreadable", "MISSING -- data exists but could not be sampled "
                "(504/timeout); coverage is unknown, not 0%")
    if reason == "empty":
        return ("empty", "empty -- no records ingested into this schema")
    return ("not_sampled", "not sampled (out of scope or no dataset)")


CONFIDENTIAL = "STRICTLY CONFIDENTIAL"
FIELD_INDEX_COLUMNS = ["Field (dot notation)", "Data Type", "Friendly Name",
                       "Identity", "Schema", "SQL table name(s)", "Tab",
                       "Coverage %", "Top values (count)"]



# ----------------------------------------------------------------------------
# Audiences / segments (lifted from journey_audience_census.py)
# ----------------------------------------------------------------------------
# The dictionary describes the data; this describes what the business does WITH
# it. Same sandbox, same run, so schemas -> datasets -> audiences -> the rule
# behind each audience can be read top to bottom in one workbook.
#
# NOTE: this block is a copy of the same code in journey_audience_census.py, by
# request. A fix made here does NOT reach that script, or vice versa.
AUDIENCES_URL = "https://platform.adobe.io/data/core/ups/audiences"
UNIFIED_TAGS_URL = "https://experience.adobe.io/unifiedtags/tags"

SYSTEM_TAG_PREFIXES = ("audience_portal_",)

def fetch_tag_vocabulary(token, conf, sandbox):
    """{tag_id: tag_name} for the whole org, paged from the Unified Tags service.
    One bulk walk instead of a GET per tag id -- the vocabulary is small (~140)
    while the audiences referencing it run to thousands."""
    vocab, cursor, guard = {}, None, 0
    while guard < 100:
        guard += 1
        url = f"{UNIFIED_TAGS_URL}?limit=100"
        if cursor:
            url += f"&start={urllib.parse.quote(str(cursor))}"
        try:
            body, _ = http(url, headers=aep_headers(token, conf, sandbox),
                           timeout=60)
            page = json.loads(body) or {}
        except Exception as e:
            logger.warning(f"  tag vocabulary page {guard} failed "
                           f"({type(e).__name__}: {e}); names may fall back to ids.")
            break
        items = page.get("tags") or []
        if not items:
            break
        for t in items:
            if t.get("id"):
                vocab[t["id"]] = t.get("name") or t["id"][:8]
        cursor = (page.get("_page") or {}).get("next")
        if not cursor:
            break
    return vocab

USER_DIRECTORY_URL = "https://usermanagement.adobe.io/v2/usermanagement/users"

def fetch_user_directory(token, conf):
    """{ims_user_id: email} for the org, paged from User Management.

    Audiences record who created/modified them as an opaque IMS id
    (D82F225C...@805f1e8e...), which tells a reader nothing. The directory is a
    couple of pages for a few thousand users, so it is fetched once per run and
    used to turn those ids into people. Empty dict on failure -- the columns
    then show raw ids rather than the report dying over a nicety.
    """
    users, page = {}, 0
    while page < 50:
        url = f"{USER_DIRECTORY_URL}/{conf['org_id']}/{page}"
        try:
            # Org-level endpoint: no sandbox header (and urllib rejects a None
            # header value, so it is dropped rather than passed empty).
            hdrs = {k: v for k, v in aep_headers(token, conf, "prod").items()
                    if k != "x-sandbox-name"}
            body, _ = http(url, headers=hdrs, timeout=60)
            r = json.loads(body) or {}
        except Exception as e:
            logger.warning(f"  user directory page {page} failed "
                           f"({type(e).__name__}); creator columns will show "
                           f"raw ids.")
            break
        for u in r.get("users") or []:
            if u.get("id"):
                users[u["id"]] = u.get("email") or u.get("username") or ""
        if r.get("lastPage"):
            break
        page += 1
    return users

def resolve_actor(actor_id, directory):
    """An IMS id turned into something a human can act on.

    Three kinds appear on audiences and only the first is a person:
      * a directory user        -> their email
      * @techacct.adobe.com     -> an API integration, not somebody's doing
      * @AdobeID service names  -> Adobe's own automation (halo refreshes etc)
    """
    if not actor_id:
        return ""
    actor_id = str(actor_id)
    if actor_id in directory and directory[actor_id]:
        return directory[actor_id]
    if actor_id.endswith("@techacct.adobe.com"):
        return f"(API integration {actor_id.split('@')[0][:8]})"
    if actor_id.endswith("@AdobeID"):
        return f"(Adobe service: {actor_id.split('@')[0]})"
    return actor_id

def fetch_all_audiences(token, conf, sandbox):
    """Every audience in a sandbox, paged in full. Returns (audiences, complete).

    `complete` is False when the walk ended before totalCount -- the caller must
    say so rather than presenting a short list as the whole estate.
    """
    out, seen, cursor, guard, total = [], set(), None, 0, None
    while guard < 500:
        guard += 1
        url = f"{AUDIENCES_URL}?limit=100"
        if cursor is not None:
            url += f"&start={urllib.parse.quote(str(cursor))}"
        try:
            body, _ = http(url, headers=aep_headers(token, conf, sandbox),
                           timeout=90)
            page = json.loads(body) or {}
        except Exception as e:
            logger.warning(f"  {sandbox}: audience page {guard} failed "
                           f"({type(e).__name__}: {e}); list is INCOMPLETE.")
            return out, False
        kids = page.get("children") or []
        meta = page.get("_page") or {}
        if total is None:
            total = meta.get("totalCount")
        for c in kids:
            if c.get("id") and c["id"] not in seen:
                seen.add(c["id"])
                out.append(c)
        if not kids:
            break
        cursor = meta.get("next")
        if cursor is None:
            break
    complete = total is None or len(out) >= int(total)
    if not complete:
        logger.warning(f"  {sandbox}: collected {len(out)} of {total} audience(s) "
                       f"-- list is INCOMPLETE.")
    return out, complete

# ----------------------------------------------------------------------------
# PQL rendering
# ----------------------------------------------------------------------------
# An audience's rule comes back as 'pql/json': Adobe's syntax TREE, not the PQL
# you see in the UI. There is no supported way to ask for the text form -- every
# format/expressionFormat parameter still returns json, and the conversion
# endpoint is not available to a read-only credential -- so we render it here.
#
# The grammar is large (15 node types, 34 functions, trees up to 17 deep). We
# render the shapes that make up the overwhelming majority -- boolean logic,
# comparisons, field paths, segment references -- and REFUSE to guess at the
# rest. Event-sequence and time-window nodes emit a visible <...> marker and set
# the "partial" flag, so a half-rendered rule can never be mistaken for a whole
# one. The raw tree is always kept in its own column.
PQL_INFIX = {"and": "and", "or": "or", "=": "=", ">": ">", "<": "<",
             ">=": ">=", "<=": "<=", "!=": "!=", "equals": "=",
             "notEqualTo": "!="}

PQL_WORDY = {"startsWith": "starts with", "endsWith": "ends with",
             "contains": "contains", "doesNotContain": "does not contain",
             "in": "in", "notIn": "not in"}

# Nodes describing event sequences / time windows. Rendering these faithfully is
# a project in itself; misrendering one would quietly change what the rule means.
PQL_COMPLEX = {"chain", "occurs", "timeQualification", "duration", "select",
               "varDecl", "element", "gap", "range"}

XL_CELL_LIMIT = 32000          # Excel's hard ceiling is 32767

def _pql_literal(node):
    v = node.get("value")
    lt = node.get("literalType")
    if lt == "List" and isinstance(v, list):
        return "[" + ", ".join(_pql_literal({"value": x, "literalType": "String"}
                                            if isinstance(x, str) else {"value": x})
                               for x in v) + "]"
    if isinstance(v, str) and lt in ("String", "Timestamp", "TimeDirection",
                                     "TimeUnit", "Comparison", None):
        return f'"{v}"'
    if isinstance(v, bool):
        return "true" if v else "false"
    return str(v)

def _pql_path(node, state):
    """Dotted field path. The base is a parameter/var reference and contributes
    nothing readable, so it is dropped: v0._experience.x -> _experience.x"""
    parts = []
    cur = node
    while isinstance(cur, dict) and cur.get("nodeType") == "fieldLookup":
        parts.append(cur.get("fieldName") or "?")
        cur = cur.get("object")
    if isinstance(cur, dict) and cur.get("nodeType") not in (
            "parameterReference", "varRef", None):
        head = _render_node(cur, state)
        if head:
            parts.append(head)
    return ".".join(reversed(parts))

def _render_node(node, state):
    if node is None:
        return ""
    if isinstance(node, list):
        return ", ".join(_render_node(n, state) for n in node)
    if not isinstance(node, dict):
        return str(node)

    nt = node.get("nodeType")
    if nt in PQL_COMPLEX:
        state["partial"] = True
        state["complex"].add(nt)
        return f"<{nt}: see raw>"
    if nt == "literal":
        return _pql_literal(node)
    if nt == "fieldLookup":
        return _pql_path(node, state)
    if nt in ("parameterReference", "varRef"):
        return ""                       # the implicit profile/event being tested
    if nt == "lambda":
        return _render_node(node.get("body"), state)
    if nt != "fnApply":
        state["partial"] = True
        state["complex"].add(nt or "unknown")
        return f"<{nt or 'unknown'}: see raw>"

    fn = node.get("fnName")
    params = node.get("params") or []

    if fn == "inSegment" and params:
        sid = (params[0] or {}).get("value")
        name = state["segments"].get(sid)
        return f'inSegment("{name}")' if name else f'inSegment({sid})'
    if fn == "not" and len(params) == 1:
        return f"not ({_render_node(params[0], state)})"
    if fn in ("exists", "isNotNull") and params:
        return f"{_render_node(params[0], state)} exists"
    if fn == "isNull" and params:
        return f"{_render_node(params[0], state)} is null"
    # stringCompare NAMES its comparison in the first parameter -- it is
    # stringCompare(<op>, <field>, <value>), not an infix pair. Treating it as
    # infix renders the operator itself as the left-hand operand.
    if fn == "stringCompare" and len(params) >= 3:
        op = (params[0] or {}).get("value") or "compares"
        return (f"{_render_node(params[1], state)} {op} "
                f"{_render_node(params[2], state)}")
    if fn == "get" and len(params) >= 2:
        base = _render_node(params[0], state)
        fld = (params[1] or {}).get("value") or _render_node(params[1], state)
        return f"{base}.{fld}" if base else str(fld)
    if fn in PQL_INFIX and len(params) >= 2:
        op = PQL_INFIX[fn]
        # equals/stringCompare carry a trailing case-sensitivity flag
        operands = params[:2] if fn in ("equals", "stringCompare") else params
        rendered = [_render_node(p, state) for p in operands]
        joined = f" {op} ".join(r for r in rendered if r != "")
        return f"({joined})" if op in ("and", "or") and len(rendered) > 1 else joined
    if fn in PQL_WORDY and len(params) >= 2:
        return (f"{_render_node(params[0], state)} {PQL_WORDY[fn]} "
                f"{_render_node(params[1], state)}")
    args = ", ".join(r for r in (_render_node(p, state) for p in params) if r != "")
    return f"{fn}({args})"

def render_pql(expression, segments):
    """(readable_text, raw_text, partial) for an audience's expression.

    partial=True means at least one node could not be rendered faithfully and
    appears as a <marker>; the row is labelled so nobody reads it as complete.
    """
    if not expression:
        return "", "", False
    raw = expression.get("value")
    raw_text = raw if isinstance(raw, str) else json.dumps(raw)
    if expression.get("format") == "pql/text":
        return str(raw_text), str(raw_text), False
    try:
        tree = json.loads(raw_text) if isinstance(raw_text, str) else raw_text
    except (ValueError, TypeError):
        return "", str(raw_text), True
    state = {"segments": segments, "partial": False, "complex": set()}
    text = _render_node(tree, state)
    if state["partial"] and state["complex"]:
        text = f"[PARTIAL: {', '.join(sorted(state['complex']))}] {text}"
    return text, str(raw_text), state["partial"]

def audience_eval_type(aud: dict) -> str:
    """Streaming / Edge / Batch from the evaluationInfo carried on the LIST
    payload (no per-audience GET needed). '' when the record omits it."""
    ei = aud.get("evaluationInfo") or {}
    if (ei.get("continuous") or {}).get("enabled"):
        return "Streaming"
    if (ei.get("synchronous") or {}).get("enabled"):
        return "Edge"
    if (ei.get("batch") or {}).get("enabled"):
        return "Batch"
    return ""

def split_audience_tags(aud: dict, vocab: dict):
    """(named_tags, system_tags, unresolved_ids) for one audience.

    named  -- resolved against the Unified Tags vocabulary: the human tagging
              this report is actually about.
    system -- machine-written key:value stamps, kept out of the counts.
    unresolved -- UUIDs with no vocabulary entry (deleted tag, or a tag the
              credential cannot see). Reported rather than silently dropped.
    """
    named, system, unresolved = [], [], []
    for t in (aud.get("tags") or []):
        t = str(t)
        if ":" in t or t.startswith(SYSTEM_TAG_PREFIXES):
            system.append(t)
        elif t in vocab:
            named.append(vocab[t])
        else:
            unresolved.append(t)
    return sorted(set(named)), system, unresolved


def attach_audiences(token, conf, res, caches):
    """Read the sandbox's audiences onto res['audiences'].

    Never fatal: the dictionary's job is schemas and fields, so a failure here
    costs the Audiences tab and nothing else. res['audiences_complete'] records
    whether the listing finished, so a short list is never presented as the
    whole estate.
    """
    sandbox = res["name"]
    res["audiences"], res["audiences_complete"] = [], True
    try:
        if caches.get("vocab") is None:
            caches["vocab"] = fetch_tag_vocabulary(token, conf, sandbox)
            logger.info(f"  tag vocabulary: {len(caches['vocab'])} tag(s).")
        if caches.get("directory") is None:
            caches["directory"] = fetch_user_directory(token, conf)
            logger.info(f"  user directory: {len(caches['directory'])} user(s).")
        auds, complete = fetch_all_audiences(token, conf, sandbox)
    except Exception as e:
        logger.warning(f"  {sandbox}: audiences unavailable "
                       f"({type(e).__name__}: {e}); the Audiences tab will be "
                       f"empty for this sandbox.")
        res["audiences_complete"] = False
        return

    vocab = caches["vocab"] or {}
    directory = caches["directory"] or {}
    # inSegment() names another audience by id; resolve so a combined rule reads
    # as prose. Sandbox-scoped, so built from this sandbox's own listing.
    seg_names = {}
    for a in auds:
        for key in ("id", "audienceId"):
            if a.get(key):
                seg_names[a[key]] = a.get("name") or ""

    rows, n_pql, n_partial, tagged = [], 0, 0, 0
    for a in auds:
        named, system, unresolved = split_audience_tags(a, vocab)
        tagged += bool(named)
        pql, pql_raw, partial = render_pql(a.get("expression"), seg_names)
        if pql:
            n_pql += 1
            n_partial += bool(partial)
        rows.append([
            sandbox, a.get("name") or "", a.get("id") or "",
            audience_eval_type(a), a.get("lifecycleState") or "",
            ", ".join(named), len(named), a.get("namespace") or "",
            resolve_actor(a.get("createdBy"), directory),
            resolve_actor(a.get("lastModifiedBy"), directory),
            pql[:XL_CELL_LIMIT],
            "partial" if partial else ("yes" if pql else ""),
            pql_raw[:XL_CELL_LIMIT],
            ", ".join(t[:8] for t in unresolved), len(system),
        ])
    res["audiences"] = rows
    res["audiences_complete"] = complete
    logger.info(f"  audiences: {len(rows)} ({tagged} tagged, {n_pql} with a "
                f"rule, {n_partial} of those only partly renderable).")


def script_provenance() -> dict:
    """Where this workbook came from, for the "who produced this?" conversation.

    The script's own SCRIPT_DATE is hand-maintained and goes stale the moment
    someone edits without touching it, so the last-changed date is taken from
    git -- the commit that last touched THIS file -- and only falls back to the
    file's mtime when git can't answer (no repo, git absent, exported copy).

    `dirty` matters: if the working tree has uncommitted edits then the commit
    id does NOT describe the code that produced the file, and the workbook says
    so rather than implying a provenance it cannot support.
    """
    import subprocess

    me = Path(__file__).resolve()
    prov = {"script": me.name, "version": SCRIPT_VERSION,
            "generated": f"{datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}",
            "changed": "", "commit": "", "dirty": False, "source": "file mtime"}

    def git(*args):
        return subprocess.run(("git", "-C", str(me.parent)) + args,
                              capture_output=True, text=True, timeout=15)

    try:
        r = git("log", "-1", "--format=%cI%x00%h", "--", str(me))
        if r.returncode == 0 and r.stdout.strip():
            iso, _, short = r.stdout.strip().partition("\x00")
            prov["changed"] = iso.strip()[:16].replace("T", " ")
            prov["commit"] = short.strip()
            prov["source"] = "git"
        d = git("status", "--porcelain", "--", str(me))
        prov["dirty"] = bool(d.returncode == 0 and d.stdout.strip())
    except Exception:
        pass                      # any git trouble -> fall through to mtime

    if not prov["changed"]:
        mtime = datetime.fromtimestamp(me.stat().st_mtime, timezone.utc)
        prov["changed"] = f"{mtime:%Y-%m-%d %H:%M UTC}"
    return prov


def provenance_line(prov: dict) -> str:
    """One line naming the script, when it ran, and when it last changed."""
    bits = [f"Generated {prov['generated']}",
            f"by {prov['script']} v{prov['version']}"]
    if prov["commit"]:
        bits.append(f"commit {prov['commit']}")
    bits.append(f"script last changed {prov['changed']}"
                + ("" if prov["source"] == "git" else " (file timestamp)"))
    bits.append(f"release notes: {RELEASE_NOTES_URL}")
    line = "  |  ".join(bits)
    if prov["dirty"]:
        line += ("  |  WARNING: the script had UNCOMMITTED changes when this ran "
                 "-- the commit id above does not describe the code that "
                 "produced this file.")
    return line


def _archive_previous(out_dir: Path, safe_client: str) -> int:
    """Move any prior dictionary for this client into out_dir/archive/ so the
    output folder only ever holds the newest. Returns how many were moved."""
    prev = sorted(out_dir.glob(f"Data Dictionary - {safe_client} - *.xlsx"))
    if not prev:
        return 0
    arch = out_dir / "archive"
    arch.mkdir(exist_ok=True)
    moved = 0
    for p in prev:
        dest = arch / p.name
        i = 2
        while dest.exists():               # never clobber an archived copy
            dest = arch / f"{p.stem} ({i}){p.suffix}"
            i += 1
        try:
            p.rename(dest)
        except OSError as e:
            # Almost always the previous workbook still open in Excel. Tidying
            # is a courtesy; it must never destroy the run that just spent
            # forty minutes sampling data.
            logger.warning(f"Could not archive {p.name} ({e.__class__.__name__})"
                           f" -- it is probably open. Leaving it where it is.")
            continue
        moved += 1
    return moved


def write_xlsx(results, client: str, datestr: str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed -- skipping XLSX "
                       "(pip install -r requirements.txt).")
        return None

    OUTPUT_DIR.mkdir(exist_ok=True)
    # The filename says WHAT this is, WHERE it came from and WHEN -- not which
    # credential happened to read it. The credential label is an artefact of our
    # own key management and means nothing to whoever opens the file.
    sbs = [str(res.get("name") or "") for res in results if res.get("name")]
    where = "-".join(sbs) if 0 < len(sbs) <= 2 else f"{len(sbs)}-sandboxes"
    safe_client = re.sub(r"[^0-9A-Za-z _-]+", "", where).strip() or "sandbox"
    path = OUTPUT_DIR / f"Data Dictionary - {safe_client} - {datestr}.xlsx"

    # ...and the titles INSIDE the workbook say the same thing as the filename.
    # `client` is filled only when the credential record carries an explicit
    # client name; otherwise we identify the workbook by sandbox rather than by
    # whichever key happened to read it.
    label = client or where

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=_HEADER_BG)
    title_font = Font(bold=True, size=14)
    conf_font = Font(bold=True, size=11, color="C00000")
    center = Alignment(horizontal="center")
    any_dd = any(k.get("datadict") for res in results for k in res["kept"])

    def style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = head_font
            cell.fill = head_fill
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 28
        ws.freeze_panes = ws.cell(row=row + 1, column=1)
        # Remember where the table starts so filters can be applied once every
        # row has been written -- the range needs the LAST row, which isn't
        # known yet here. Applied in one sweep just before saving, so no sheet
        # can be forgotten.
        ws._dd_table = (row, ncols)

    def autofit(ws, widths):
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    def confidential(ws, cell="A1"):
        # Visible banner + a print header on every page.
        ws[cell] = CONFIDENTIAL
        ws[cell].font = conf_font
        ws.oddHeader.center.text = f'&"-,Bold"&12&KC00000{CONFIDENTIAL}'
        ws.evenHeader.center.text = ws.oddHeader.center.text

    wb = Workbook()

    # Unique worksheet name per kept schema (Excel: <=31 chars, unique). Built
    # up front so the Field Index and Schemas index can name each schema's tab.
    used = {"summary", "how to use", "schemas", "field index", "datasets",
            "audiences"}
    tabbed = []  # (res, k, sheet_name)
    for res in results:
        for k in res["kept"]:
            tabbed.append((res, k, _safe_sheet_name(_sheet_label(k["title"]),
                                                    used)))

    # ---- Summary tab --------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    confidential(ws)
    ws["A2"] = f"Data Dictionary v{SCRIPT_VERSION}  -  {label}"
    ws["A2"].font = title_font
    prov = script_provenance()
    ws["A3"] = provenance_line(prov)
    ws["A3"].font = Font(italic=True,
                         color="C00000" if prov["dirty"] else "666666")
    ws["A4"] = (
        "WHAT THIS IS: a top-to-bottom picture of this Adobe Experience Platform "
        "sandbox, read straight from the platform -- the SCHEMAS that define the "
        "data, the DATASETS that hold it (and the SQL table name to query each "
        "one), and the AUDIENCES built on top of it with the rule behind each. "
        "Kept = a schema referenced by at least one dataset and not ad-hoc, AJO, "
        "system or test. See the How to Use tab if you're not sure where to "
        "start.")
    ws["A4"].font = Font(italic=True, color="666666")
    ws["A4"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[4].height = 46

    # Tab colour key: a coloured swatch cell + its meaning, so readers know what
    # the purple schema tabs signify without having to ask.
    ws["A5"] = "Tab colour key:"
    ws["A5"].font = Font(italic=True, bold=True, color="666666")
    ws["B5"].fill = PatternFill("solid", fgColor=PROFILE_TAB_COLOR)
    ws["C5"] = ("Purple tab = XDM Individual Profile (Profile-class) schema. "
                "Coverage for these is sampled from the Profile Snapshot Export "
                "union, not the pre-merge feeding datasets.")
    ws["C5"].font = Font(italic=True, color="666666")

    hdr = ["Sandbox", "Env", "Schemas seen", "Kept", "Dropped: no-dataset",
           "Dropped: adhoc", "Dropped: AJO", "Dropped: system",
           "Dropped: test", "Kept fields", "Relationships", "Dual labels"]
    r = 6
    for c, nm in enumerate(hdr, 1):
        ws.cell(r, c, nm)
    style_header(ws, len(hdr), row=r)
    r += 1
    for res in results:
        s = res["stats"]
        ws.cell(r, 1, res["title"])
        ws.cell(r, 2, res["env"]).alignment = center
        for c, val in enumerate([s["total"], s["kept"], s["no_dataset"],
                                 s["adhoc"], s["ajo"], s["system"], s["test"],
                                 s["fields"], s["relationships"],
                                 res["labels_n"]], 3):
            ws.cell(r, c, val).alignment = center
        r += 1
    autofit(ws, [24, 6, 13, 7, 19, 15, 13, 15, 13, 12, 14, 12])

    # ---- Data completeness: what's MISSING ----------------------------------
    # On a --data-dict run, list every schema whose coverage is not trustworthy
    # (missing / partial / empty) so the dictionary is never taken as gospel.
    if any_dd:
        r += 2
        ws.cell(r, 1, "DATA COMPLETENESS -- coverage is NOT complete; "
                "the schemas below are missing or partial:").font = Font(
                    bold=True, color="C00000")
        r += 1
        gap_hdr = ["Schema", "Tab", "Sandbox", "Coverage status"]
        for c, nm in enumerate(gap_hdr, 1):
            cell = ws.cell(r, c, nm)
            cell.font = head_font
            cell.fill = head_fill
        r += 1
        listed = 0
        for res, k, nm in tabbed:
            code, label = _coverage_status(k)
            if code == "ok":
                continue
            ws.cell(r, 1, k["title"])
            ws.cell(r, 2, nm)
            ws.cell(r, 3, res["title"])
            cell = ws.cell(r, 4, label)
            if code in ("unreadable", "partial"):
                cell.font = Font(color="C00000", bold=True)
            r += 1
            listed += 1
        if not listed:
            ws.cell(r, 1, "(every sampled schema returned coverage)").font = Font(
                italic=True, color="666666")
        r += 1

    # Author / provenance footer, so the file can be traced back to a person as
    # well as to a script.
    r += 1
    ws.cell(r, 1, f"Produced by {SCRIPT_AUTHOR}").font = Font(bold=True,
                                                              color="666666")
    for link in (AUTHOR_SITE, AUTHOR_LINKEDIN):
        r += 1
        c = ws.cell(r, 1, link)
        c.font = Font(color="0563C1", underline="single")
        c.hyperlink = link

    # ---- How to Use: for the reader who has never seen this before ----------
    # Second tab deliberately, so it sits next to the Summary rather than being
    # buried behind thirty schema tabs.
    hu = wb.create_sheet("How to Use")
    confidential(hu)
    hu["A2"] = "How to use this workbook"
    hu["A2"].font = title_font
    hu["A3"] = ("Everything here was read directly from Adobe Experience "
                "Platform. Nothing is hand-maintained, so it is as current as "
                "the date on the Summary tab -- and no more.")
    hu["A3"].font = Font(italic=True, color="666666")

    HOW_TO = [
        ("START HERE", "", ""),
        ("I want to...", "Go to", "What you'll find"),
        ("...see what data exists at all",
         "Schemas",
         "Every schema kept, its class, how many datasets feed it, and which "
         "tab describes it."),
        ("...find one particular field",
         "Field Index",
         "Every field across every schema in one list. Filter the Field column "
         "to search; it tells you which schema and tab it belongs to."),
        ("...write SQL against the data",
         "Datasets",
         "Maps each dataset's friendly name to its SQL table name. Use the "
         "TABLE NAME in FROM, not the friendly name -- they differ. The "
         "Profile column flags what feeds Real-Time Customer Profile."),
        ("...understand one schema in detail",
         "Its own tab",
         "Every field with its type, friendly label, whether it is an "
         "identity, and (when the coverage pass ran) how often it is actually "
         "populated plus its five commonest values."),
        ("...see who is being targeted, and how",
         "Audiences",
         "Every audience with its tags, who built it, who last changed it, and "
         "the rule behind it in readable form."),
        ("", "", ""),
        ("THINGS THAT WILL CATCH YOU OUT", "", ""),
        ("Coverage % is a SAMPLE",
         "schema tabs",
         "Taken from up to a thousand real records, not the whole dataset. "
         "Treat it as 'roughly how often this is filled in', not an exact "
         "figure."),
        ("MISSING is not the same as 0%",
         "Summary",
         "0% means we looked and the field was empty. MISSING means we could "
         "not read the data at all. Any schema in that state is listed in the "
         "DATA COMPLETENESS block on the Summary tab and banner-flagged on its "
         "own tab."),
        ("Some rules are only PARTLY shown",
         "Audiences",
         "Rules marked 'partial' contain an event-sequence or time-window "
         "clause that is shown as <...>. What IS displayed is correct, but "
         "there is more to the rule -- the raw column holds the full "
         "definition."),
        ("Not every audience has a rule",
         "Audiences",
         "Only Origin=AEPSegments audiences are rule-based. Uploads, Data "
         "Distiller and Audience Orchestration compositions are built "
         "elsewhere, so their rule column is legitimately blank."),
        ("'Last modified by' is often a robot",
         "Audiences",
         "Entries like (Adobe service: pathos) are Adobe's own automation, not "
         "a colleague. Only email addresses are people."),
        ("", "", ""),
        ("EVERY TAB", "", ""),
        ("Filters and frozen headers are on",
         "all tabs",
         "Click any header arrow to filter. The header row stays put as you "
         "scroll."),
        ("Confidential",
         "all tabs",
         "With the coverage pass, this workbook contains real sampled customer "
         "data. Treat it accordingly."),
    ]
    hr = 5
    for c, nm in enumerate(("I want to...", "Go to", "What you'll find"), 1):
        hu.cell(hr, c, nm)
    style_header(hu, 3, row=hr)
    rr = hr + 1
    section_fill = PatternFill("solid", fgColor="DCE6F1")
    for left, mid, right in HOW_TO[1:]:
        if left in ("I want to...",):
            continue
        a = hu.cell(rr, 1, left)
        b = hu.cell(rr, 2, mid)
        c_ = hu.cell(rr, 3, right)
        if left and not mid and not right:            # a section heading row
            for cell in (a, b, c_):
                cell.fill = section_fill
            a.font = Font(bold=True, color="1F4E78")
        else:
            a.font = Font(bold=True)
            b.font = Font(color="1F4E78")
        for cell in (a, b, c_):
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        hu.row_dimensions[rr].height = 30 if right else 18
        rr += 1
    autofit(hu, [34, 18, 86])
    hu._dd_table = None          # prose, not a table -- no filter dropdowns
    hu.cell(rr + 1, 1, f"Generated by {SCRIPT_NAME}.py  -  {SCRIPT_AUTHOR}")
    hu.cell(rr + 1, 1).font = Font(italic=True, color="666666")
    hu.cell(rr + 2, 1, AUTHOR_SITE).font = Font(color="0563C1", underline="single")
    hu.cell(rr + 2, 1).hyperlink = AUTHOR_SITE
    hu.cell(rr + 3, 1, AUTHOR_LINKEDIN).font = Font(color="0563C1",
                                                    underline="single")
    hu.cell(rr + 3, 1).hyperlink = AUTHOR_LINKEDIN
    # What changed in this version, and every version before it.
    hu.cell(rr + 5, 1, f"What's new in v{SCRIPT_VERSION} -- release notes:").font = (
        Font(bold=True))
    hu.cell(rr + 6, 1, RELEASE_NOTES_URL).font = Font(color="0563C1",
                                                      underline="single")
    hu.cell(rr + 6, 1).hyperlink = RELEASE_NOTES_URL

    # ---- Master Field Index (every field across all schemas, for lookup) ----
    fi = wb.create_sheet("Field Index")
    confidential(fi)
    fi["A2"] = f"Field Index  -  {label}"
    fi["A2"].font = title_font
    fi["A3"] = ("Every field across all schemas. Look up an exact dot-notation "
                "path (Ctrl-F); the Tab column says which sheet it lives on.")
    fi["A3"].font = Font(italic=True, color="666666")
    hr = 5
    for c, nm in enumerate(FIELD_INDEX_COLUMNS, 1):
        fi.cell(hr, c, nm)
    style_header(fi, len(FIELD_INDEX_COLUMNS), row=hr)
    index_rows = []
    for res, k, name in tabbed:
        dd = k.get("datadict") or {}
        tnames = _sql_table_names(k)
        for (fpath, dtype, friendly, req, ident, rel) in k["fields"]:
            info = dd.get(fpath, {})
            index_rows.append((fpath, dtype, friendly, ident, k["title"], tnames,
                               name, info.get("coverage"), info.get("top")))
    index_rows.sort(key=lambda x: (x[0].lower(), x[4].lower()))
    ridx = hr + 1
    for (fpath, dtype, friendly, ident, sch, tnames, tab, cov, top) in index_rows:
        fi.cell(ridx, 1, fpath)
        fi.cell(ridx, 2, dtype)
        fi.cell(ridx, 3, friendly)
        fi.cell(ridx, 4, ident)
        fi.cell(ridx, 5, sch)
        fi.cell(ridx, 6, tnames)
        fi.cell(ridx, 7, tab)
        if cov is not None:
            cc = fi.cell(ridx, 8, cov)
            cc.alignment = center
            cc.number_format = '0"%"'
        fi.cell(ridx, 9, top)
        ridx += 1
    autofit(fi, [52, 18, 28, 20, 38, 34, 24, 11, 60])

    # ---- Schemas index tab (one row per kept schema) ------------------------
    sh = wb.create_sheet("Schemas")
    confidential(sh)
    index_cols = ["Tab"] + SCHEMA_CSV_COLUMNS
    hr = 3
    for c, nm in enumerate(index_cols, 1):
        sh.cell(hr, c, nm)
    style_header(sh, len(index_cols), row=hr)
    rr = hr + 1
    for res, k, name in tabbed:
        row = [name, res["title"], k["title"], k["class"], k["datasets"],
               _sql_table_names(k), k["n_fields"], k["n_identities"],
               k["n_relationships"], k["n_labels"], k["last_mod"], k["id"]]
        for c, val in enumerate(row, 1):
            sh.cell(rr, c, val)
        rr += 1
    autofit(sh, [26, 18, 38, 22, 9, 40, 7, 10, 13, 16, 16, 58])

    # ---- Datasets tab: friendly name -> SQL table (system) name -------------
    # EVERY dataset in each sandbox, so a query can be aimed at the right table.
    # The table name (tags['adobe/pqs/table']) is what Query Service uses in
    # FROM -- distinct from the friendly dataset name shown in the UI.
    dt = wb.create_sheet("Datasets")
    confidential(dt)
    dt["A2"] = f"Datasets / SQL table names  -  {label}"
    dt["A2"].font = title_font
    dt["A3"] = ("Every dataset and its AEP Query Service table name. SQL uses the "
                "SYSTEM table name, not the friendly name: SELECT ... FROM "
                "<Table Name>. The Profile column flags datasets enabled for "
                "Unified Profile and the Profile Snapshot Export (the merged union "
                "to query for whole profiles).")
    dt["A3"].font = Font(italic=True, color="666666")
    DATASET_COLUMNS = ["Sandbox", "Schema", "Friendly Name (dataset)",
                       "Table Name (SQL / system)", "Profile", "Dataset ID"]
    hr = 5
    for c, nm in enumerate(DATASET_COLUMNS, 1):
        dt.cell(hr, c, nm)
    style_header(dt, len(DATASET_COLUMNS), row=hr)
    rr = hr + 1
    # Profile-related datasets first (snapshot, then enabled), then the rest --
    # so the tables you need for Profile queries sit at the top.
    def _prof_rank(p):
        return 0 if p.startswith("snapshot") else (1 if p else 2)
    for res in results:
        for d in sorted(res.get("datasets") or [],
                        key=lambda x: (_prof_rank(x.get("profile") or ""),
                                       (x.get("schema_title") or "~").lower(),
                                       (x.get("name") or "").lower())):
            dt.cell(rr, 1, res["title"])
            dt.cell(rr, 2, d.get("schema_title"))
            dt.cell(rr, 3, d.get("name"))
            dt.cell(rr, 4, d.get("table"))
            pcell = dt.cell(rr, 5, d.get("profile"))
            if (d.get("profile") or "").startswith("snapshot"):
                pcell.font = Font(bold=True, color="7030A0")
            elif d.get("profile"):
                pcell.font = Font(color="2E7D32")
            dt.cell(rr, 6, d.get("id"))
            rr += 1
    autofit(dt, [18, 40, 42, 44, 26, 34])

    # ---- Audiences tab: what the business does WITH the data ----------------
    # Completes the chain: schema -> dataset -> audience -> the rule behind it.
    # Only the rule-based audiences (Origin = AEPSegments) carry PQL; uploads,
    # Data Distiller and Audience Orchestration compositions have none, which is
    # why that column is legitimately blank for a good number of rows.
    aud_rows = [r for res in results for r in (res.get("audiences") or [])]
    aud_incomplete = [res["title"] for res in results
                      if not res.get("audiences_complete", True)]
    if aud_rows or aud_incomplete:
        at = wb.create_sheet("Audiences")
        confidential(at)
        at["A2"] = f"Audiences / segments  -  {label}"
        at["A2"].font = title_font
        note = ("Every audience in the sandbox, with its tags, who created and "
                "last changed it, and its segmentation rule. PQL is rendered "
                "from Adobe's syntax tree: rows marked 'partial' contain an "
                "event-sequence or time-window clause shown as <...> -- the raw "
                "column holds the complete definition. Only Origin=AEPSegments "
                "audiences have a rule at all.")
        if aud_incomplete:
            note += (f"  WARNING: the listing was INCOMPLETE for "
                     f"{', '.join(aud_incomplete)} -- those counts understate "
                     f"the estate.")
        at["A3"] = note
        at["A3"].font = Font(italic=True,
                             color="C00000" if aud_incomplete else "666666")
        AUDIENCE_COLUMNS = ["Sandbox", "Audience name", "Audience id",
                            "Evaluation", "Lifecycle", "Tags", "Tag count",
                            "Origin", "Created by", "Last modified by",
                            "PQL (readable)", "PQL rendered",
                            "PQL (raw pql/json)", "Unresolved tag ids",
                            "System tags"]
        hr = 5
        for c, nm in enumerate(AUDIENCE_COLUMNS, 1):
            at.cell(hr, c, nm)
        style_header(at, len(AUDIENCE_COLUMNS), row=hr)
        rr = hr + 1
        # Tagged and rule-bearing audiences first: the ones anyone came to read.
        for row in sorted(aud_rows,
                          key=lambda r: (0 if r[5] else 1, 0 if r[10] else 1,
                                         str(r[1]).lower())):
            for c, val in enumerate(row, 1):
                at.cell(rr, c, val)
            rr += 1
        autofit(at, [16, 46, 36, 12, 14, 34, 10, 20, 32, 32, 70, 13, 50, 22, 12])
        at.freeze_panes = at.cell(hr + 1, 1).coordinate
        at.auto_filter.ref = (f"A{hr}:"
                              f"{get_column_letter(len(AUDIENCE_COLUMNS))}{rr - 1}")

    # ---- One tab per schema, listing its individual fields ------------------
    for res, k, name in tabbed:
        sheet = wb.create_sheet(name)
        # Profile-class schema (the post-merge union) -> coloured tab so it
        # stands out from the event/lookup schemas; its coverage is sampled
        # differently (Profile Snapshot Export, not the pre-merge feeds). The
        # colour is explained in the Summary tab's colour key.
        if k.get("meta_class") == PROFILE_CLASS:
            sheet.sheet_properties.tabColor = PROFILE_TAB_COLOR
        confidential(sheet)
        dd = k.get("datadict")
        sheet["A2"] = k["title"]
        sheet["A2"].font = title_font
        a3 = (f'{k["class"]}   |   {k["datasets"]} dataset(s)   |   '
              f'{k["n_fields"]} fields   |   {k["n_identities"]} identity'
              f'   |   {k["n_relationships"]} relationship'
              f'   |   modified {k["last_mod"]}   |   sandbox: {res["title"]}')
        if dd:
            a3 += f'   |   data dictionary: sampled {k.get("dd_rows", 0)} records'
            if k.get("dd_source"):
                a3 += f'   |   source: {k["dd_source"]}'
        # Loud flag when coverage is missing/partial so this tab is never read as
        # gospel: data exists but couldn't be sampled, or only some batches read.
        a3_color = "666666"
        if any_dd:
            code, label = _coverage_status(k)
            if code != "ok":
                a3 += f'   |   /!\\ COVERAGE {label}'
                if code in ("unreadable", "partial"):
                    a3_color = "C00000"
        sheet["A3"] = a3
        sheet["A3"].font = Font(italic=True, color=a3_color,
                                bold=(a3_color == "C00000"))
        sheet["A4"] = k["id"]
        sheet["A4"].font = Font(italic=True, color="999999", size=9)
        # SQL table (system) name(s) for this schema's dataset(s) -- the FROM
        # target(s), right next to the field list so a query can be formed here.
        tnames = _sql_table_names(k)
        if tnames:
            sheet["A5"] = f"SQL table name(s):  {tnames}"
            sheet["A5"].font = Font(italic=True, bold=True, color="C55A11")

        cols = list(SCHEMA_FIELD_COLUMNS)
        if dd:
            cols += ["Coverage %", "Top values (count)"]
        hr = 6
        for c, nm in enumerate(cols, 1):
            sheet.cell(hr, c, nm)
        style_header(sheet, len(cols), row=hr)
        ridx = hr + 1
        for (fpath, dtype, friendly, req, ident, rel) in k["fields"]:
            sheet.cell(ridx, 1, fpath)
            sheet.cell(ridx, 2, dtype)
            sheet.cell(ridx, 3, friendly)
            sheet.cell(ridx, 4, req).alignment = center
            sheet.cell(ridx, 5, ident)
            sheet.cell(ridx, 6, rel)
            if dd:
                info = dd.get(fpath, {})
                cov = sheet.cell(ridx, 7, info.get("coverage"))
                cov.alignment = center
                cov.number_format = '0"%"'
                sheet.cell(ridx, 8, info.get("top"))
            ridx += 1
        autofit(sheet, [50, 20, 28, 9, 22, 42] + ([11, 70] if dd else []))

    # ---- Filters on every table, in one sweep -------------------------------
    # Done here rather than per sheet so a tab added later can't quietly ship
    # without them.
    for sheet in wb.worksheets:
        table = getattr(sheet, "_dd_table", None)
        if not table:
            continue
        hrow, ncols = table
        if sheet.max_row > hrow:
            sheet.auto_filter.ref = (f"A{hrow}:"
                                     f"{get_column_letter(ncols)}{sheet.max_row}")

    archived = _archive_previous(OUTPUT_DIR, safe_client)
    if archived:
        logger.info(f"Archived {archived} previous dictionary file(s) -> "
                    f"{OUTPUT_DIR / 'archive'}")
    try:
        wb.save(path)
    except PermissionError:
        # The file is open in Excel. A long sampling run must not be thrown away
        # over that -- write alongside it and say so plainly.
        alt = path.with_name(f"{path.stem} (new){path.suffix}")
        wb.save(alt)
        logger.warning(f"{path.name} is locked (open in Excel?) -- wrote "
                       f"{alt.name} instead.")
        return alt
    return path


# ----------------------------------------------------------------------------
# Driver
# ----------------------------------------------------------------------------
def _in_dd_scope(title: str, scope: str) -> bool:
    return scope == "all" or scope in (title or "").lower()


def _resolve_profile_snapshot(token, conf, sandbox, profile_snapshot):
    """Decide which dataset to sample for a Profile-class schema's coverage.
    Returns (dataset_id, source_label) or None to signal a fallback to feeding
    datasets. An explicit --profile-snapshot=<id> override wins; otherwise the
    default merge policy -> its snapshot export dataset."""
    if profile_snapshot:
        logger.info(f"  profile: using --profile-snapshot override dataset "
                    f"{profile_snapshot}.")
        return profile_snapshot, f"Profile Snapshot Export (override {profile_snapshot})"
    mp = get_default_merge_policy(token, conf, sandbox)
    if not mp:
        return None
    mp_id, mp_name = str(mp.get("id")), (mp.get("name") or str(mp.get("id")))
    logger.info(f"  profile: default merge policy '{mp_name}' ({mp_id}); "
                f"locating its Profile Snapshot Export dataset...")
    found = find_profile_snapshot_dataset(token, conf, sandbox, mp_id)
    if not found:
        logger.warning(f"  profile: no Profile Snapshot Export dataset matches the "
                       f"default merge policy ({mp_id}).")
        return None
    snap_dsid, snap_name = found
    logger.info(f"  profile: snapshot dataset '{snap_name}' ({snap_dsid}).")
    return snap_dsid, f"Profile Snapshot Export (default merge policy '{mp_name}')"


def add_data_dictionary(token, conf, results, dd_scope, dd_rows,
                        profile_snapshot=None):
    """Sample real records and attach k['datadict'] + k['dd_rows'] to each kept
    schema in scope. Mutates `results` in place.

    Profile-class schemas (the post-merge union) are sampled from the Profile
    Snapshot Export dataset of the default merge policy rather than their
    pre-merge feeding datasets -- see _resolve_profile_snapshot."""
    try:
        import pyarrow  # noqa: F401  (fail early if the reader is missing)
    except ImportError:
        logger.error("--data-dict needs pyarrow (pip install pyarrow). "
                     "Skipping the data dictionary; the rest still ran.")
        return

    for res in results:
        in_scope = [k for k in res["kept"] if _in_dd_scope(k["title"], dd_scope)]
        if not in_scope:
            logger.warning(f"  data dict: no kept schema in {res['name']} matches "
                           f"scope '{dd_scope}'. Kept titles include: "
                           f"{', '.join(k['title'] for k in res['kept'][:3])}...")
            continue
        logger.info(f"  data dict: {len(in_scope)} schema(s) in scope for "
                    f"{res['name']}: {', '.join(k['title'] for k in in_scope)}")
        ds_index = get_dataset_ids_by_schema(token, conf, res["name"])
        snap_resolved = None     # per-sandbox cache: (dsid, label) | False
        snap_unreadable = False  # snapshot 504'd once -> don't retry it per schema
        n_scope = len(in_scope)
        t_dd_start = time.perf_counter()         # wall clock for the whole sweep
        for i, k in enumerate(in_scope, 1):
            # Overall progress so an overnight all-schemas run is easy to track:
            # how far through the sandbox we are and how long we've been going.
            elapsed = time.perf_counter() - t_dd_start
            eta = (elapsed / (i - 1) * (n_scope - (i - 1))) if i > 1 else 0
            logger.info(f"  data dict: {ANSI['bold']}schema {i}/{n_scope}"
                        f"{ANSI['reset']} in {res['name']}  "
                        f"{ANSI['dim']}({elapsed/60:.1f} min elapsed"
                        f"{f', ~{eta/60:.0f} min left' if eta else ''})"
                        f"{ANSI['reset']}")
            # Profile-class coverage must come from the merged union snapshot, not
            # the pre-merge feeds. Falls back to feeding datasets if unresolvable.
            smallest_first, file_timeout = False, 75
            if k.get("meta_class") == PROFILE_CLASS:
                if snap_resolved is None:
                    snap_resolved = _resolve_profile_snapshot(
                        token, conf, res["name"], profile_snapshot) or False
                if snap_resolved:
                    snap_dsid, source_label = snap_resolved
                    k["dd_source"] = source_label
                    # The snapshot already 504'd this run -- its manifest is too big
                    # for the gateway. Don't waste ~3 min per Profile schema redoing
                    # it; flag coverage MISSING straight away.
                    if snap_unreadable:
                        k["dd_empty_reason"] = "unreadable"
                        logger.warning(f"  data dict [{i}/{n_scope}]: {k['title']} "
                                       f"-- snapshot unreadable this run (504); "
                                       f"skipping, coverage MISSING.")
                        continue
                    dsids = [snap_dsid]
                    smallest_first, file_timeout = True, DD_SNAPSHOT_TIMEOUT
                else:
                    dsids = ds_index.get(k["id"], [])
                    k["dd_source"] = ("feeding datasets (PRE-MERGE -- snapshot not "
                                      "found; coverage may read sparse)")
                    logger.warning(f"  data dict [{i}/{n_scope}]: {k['title']} is "
                                   f"Profile-class but no snapshot resolved; "
                                   f"falling back to feeding datasets.")
            else:
                dsids = ds_index.get(k["id"], [])
            if not dsids:
                logger.warning(f"  data dict [{i}/{n_scope}]: {k['title']} has no "
                               "(non-perf) datasets to sample; skipping.")
                continue
            logger.info(f"  data dict [{i}/{n_scope}]: '{k['title']}' -- "
                        f"{len(dsids)} dataset(s), {k['n_fields']} fields; "
                        f"sampling up to {dd_rows} records ...")
            try:
                rows, sstats = sample_schema_rows(
                    token, conf, res["name"], dsids, dd_rows,
                    smallest_first=smallest_first, file_timeout=file_timeout)
            except Exception as e:
                logger.warning(f"  data dict: sampling failed for {k['title']}: "
                               f"{type(e).__name__}: {e}")
                logger.debug(traceback.format_exc())
                continue
            k["dd_sample_stats"] = sstats
            if not rows:
                # Distinguish "genuinely empty" from "data exists but unreadable"
                # so a 0% schema is never silently mistaken for an empty one.
                if sstats.get("list_failed"):
                    # The batch LIST never returned for some dataset(s) -- we never
                    # learned whether they carry records, so this is MISSING, not 0%.
                    logger.warning(f"  data dict [{i}/{n_scope}]: {k['title']} -- "
                                   f"0 records, but the batch list failed for "
                                   f"{sstats['list_failed']} of {len(dsids)} "
                                   f"dataset(s) (gateway/timeout). NOT confirmed "
                                   f"empty -- contents unknown, coverage MISSING.")
                    k["dd_empty_reason"] = "unreadable"
                    if smallest_first:      # the snapshot itself failed -- cache it
                        snap_unreadable = True
                elif sstats["available"] == 0:
                    logger.warning(f"  data dict [{i}/{n_scope}]: {k['title']} -- "
                                   f"0 records: EMPTY (no batch in any of its "
                                   f"{len(dsids)} dataset(s) carries records). "
                                   f"Genuinely not populated.")
                    k["dd_empty_reason"] = "empty"
                else:
                    logger.warning(f"  data dict [{i}/{n_scope}]: {k['title']} -- "
                                   f"0 records but {sstats['available']} batch(es) "
                                   f"HAD records; {sstats['failed']} read(s) failed,"
                                   f" {sstats['empty_reads']} parsed empty. NOT "
                                   f"confirmed empty -- data exists but couldn't be "
                                   f"sampled (timeouts/parse). Re-run or raise "
                                   f"timeouts.")
                    k["dd_empty_reason"] = "unreadable"
                    if smallest_first:      # the snapshot itself failed -- cache it
                        snap_unreadable = True
                continue
            # Sampled OK, but flag if some batches failed (partial read).
            if sstats["failed"]:
                logger.warning(f"  data dict [{i}/{n_scope}]: {k['title']} -- "
                               f"note: {sstats['failed']} batch(es) failed to read; "
                               f"coverage is from the {len(rows)} rows that "
                               f"succeeded.")
            t0 = time.perf_counter()
            k["datadict"] = build_data_dict(rows, k["fields"])
            k["dd_rows"] = len(rows)
            populated = sum(1 for v in k["datadict"].values() if v["coverage"] > 0)
            logger.info(f"  data dict [{i}/{n_scope}]: '{k['title']}' done -- "
                        f"{len(rows)} records, {populated}/{len(k['fields'])} "
                        f"fields populated (>0% coverage); "
                        f"tally {time.perf_counter() - t0:.1f}s.")


def prompt_data_dict(dd_rows: int = DD_DEFAULT_ROWS) -> str | None:
    """Naked-run opt-in: ask whether to also run the slow data-dictionary pass
    (coverage % + top values). Returns the scope string ('all') on yes, else
    None. Defaults to No so a stray Enter never kicks off the long download."""
    print()
    bar = ANSI["yellow"] + "-" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Full data dictionary?{ANSI['reset']}")
    print(f"  {ANSI['dim']}Samples up to {dd_rows} real records per kept schema "
          f"(downloads + parses\n  Parquet) to add COVERAGE % and TOP-5 values "
          f"per field.\n  This can take several minutes -- much longer on big "
          f"tenants.{ANSI['reset']}")
    print(bar)
    try:
        raw = input(
            f"\nRun the full data dictionary now? "
            f"({ANSI['cyan']}y{ANSI['reset']}/{ANSI['cyan']}N{ANSI['reset']}, "
            f"default No): "
        ).strip().lower()
    except EOFError:
        return None
    if raw in ("y", "yes"):
        return DD_DEFAULT_SCOPE
    logger.info(f"{ANSI['dim']}Skipping data dictionary -- metadata only. "
                f"(Pass --data-dict next time to skip this prompt.){ANSI['reset']}")
    return None


def run(service: str, sandbox_arg: str | None,
        dd_scope: str | None = None, dd_rows: int = DD_DEFAULT_ROWS,
        profile_snapshot: str | None = None):
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print()
    print(bar)
    print(f"  {ANSI['bold']}Data Dictionary v{SCRIPT_VERSION} for "
          f"{ANSI['yellow']}{pretty_name(service)}{ANSI['reset']}  "
          f"{ANSI['dim']}({service}){ANSI['reset']}")
    print(bar)

    try:
        conf = aep_creds.load_creds(service)
    except aep_creds.CredsError as e:
        logger.error(f"Failed to load credentials for {service!r}: {e}")
        return

    try:
        resp = authenticate(conf)
    except urllib.error.HTTPError as e:
        logger.error(f"IMS auth FAILED: HTTP {e.code} "
                     f"{e.read().decode(errors='replace')[:300]}")
        return
    except Exception as e:
        logger.error(f"IMS auth FAILED: {type(e).__name__}: {e}")
        return
    token = resp["access_token"]
    logger.info(f"IMS authenticated (org {conf['org_id']}).")

    ok, result = list_sandboxes(token, conf)
    if not ok:
        logger.error(f"GET /sandboxes failed: {result}")
        return
    if not result:
        logger.warning("Authenticated, but 0 sandboxes visible.")
        return

    if sandbox_arg:
        chosen = resolve_sandboxes_arg(sandbox_arg, result)
    else:
        chosen = pick_sandboxes(result)
    if not chosen:
        logger.info("No sandboxes chosen. Exiting.")
        return
    logger.info(f"Reading {len(chosen)} sandbox(es): "
                f"{', '.join(sb.get('name', '?') for sb in chosen)}")

    results = []
    # Org-level lookups shared across sandboxes: the tag vocabulary and the user
    # directory are the same wherever you stand, so they are fetched once.
    aud_caches = {"vocab": None, "directory": None}
    for sb in chosen:
        logger.info(f"Collecting {sb.get('name', '?')} ...")
        try:
            res = collect_sandbox(token, conf, sb)
        except urllib.error.HTTPError as e:
            logger.error(f"  {sb.get('name')}: HTTP {e.code} "
                         f"{flatten_err(e.read().decode(errors='replace'))}")
            continue
        except Exception as e:
            logger.error(f"  {sb.get('name')}: {type(e).__name__}: {e}")
            logger.debug(traceback.format_exc())
            continue
        attach_audiences(token, conf, res, aud_caches)
        print_sandbox(res)
        results.append(res)

    if not results:
        logger.warning("No sandbox data collected. Nothing to write.")
        return

    # Naked run (no --data-dict flag): offer the slow coverage pass rather than
    # silently skipping it. Only when interactive -- piped/automated runs keep
    # the old behaviour and are never blocked on input().
    if dd_scope is None and sys.stdin.isatty():
        dd_scope = prompt_data_dict(dd_rows)

    if dd_scope is not None:
        logger.info(f"{ANSI['bold']}Data dictionary ENABLED{ANSI['reset']} "
                    f"(scope: '{dd_scope}', up to {dd_rows} records per schema). "
                    f"This downloads and parses Parquet -- expect minutes.")
        add_data_dictionary(token, conf, results, dd_scope, dd_rows,
                             profile_snapshot=profile_snapshot)
    elif not sys.stdin.isatty():
        # Non-interactive run that wasn't offered the prompt above.
        logger.info(f"{ANSI['dim']}Data dictionary not requested -- pass "
                    f"--data-dict to sample real values + coverage "
                    f"(adds minutes).{ANSI['reset']}")

    datestr = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    client = client_label(conf)
    xlsx_path = write_xlsx(results, client, datestr)
    if xlsx_path:
        n_tabs = sum(len(r["kept"]) for r in results)
        logger.info(f"XLSX written: {xlsx_path}  "
                    f"({n_tabs} schema tab(s) + Summary + Field Index + "
                    f"Schemas index + Datasets)")

    total_kept = sum(r["stats"]["kept"] for r in results)
    total_seen = sum(r["stats"]["total"] for r in results)
    logger.info(f"Done. {total_kept} kept of {total_seen} schemas across "
                f"{len(results)} sandbox(es).")


def print_banner() -> None:
    bar = ANSI["cyan"] + "=" * 72 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}{SCRIPT_NAME} v{SCRIPT_VERSION}{ANSI['reset']}   ({SCRIPT_DATE})")
    print(f"  by {SCRIPT_AUTHOR}")
    print(f"  {ANSI['dim']}Extracts AEP XDM schemas into a tabbed Excel data dictionary "
          f"with field coverage + sample values.{ANSI['reset']}")
    print(bar)
    print(aep_creds.source_banner())


def main():
    print_banner()
    args = sys.argv[1:]
    sandbox_arg = None
    dd_scope = None
    dd_rows = DD_DEFAULT_ROWS
    profile_snapshot = None
    positional = []
    for a in args:
        if a.startswith("--sandbox="):
            sandbox_arg = a.split("=", 1)[1]
        elif a in ("--data-dict", "--dd"):
            dd_scope = DD_DEFAULT_SCOPE          # ALL kept schemas, by default
        elif a.startswith("--data-dict="):
            dd_scope = a.split("=", 1)[1].strip().lower()
        elif a.startswith("--profile-snapshot="):
            profile_snapshot = a.split("=", 1)[1].strip() or None
        elif a.startswith("--dd-rows="):
            try:
                dd_rows = max(1, int(a.split("=", 1)[1]))
            except ValueError:
                logger.warning(f"Ignoring bad --dd-rows value in {a!r}.")
        elif a in ("--debug", "-v"):
            logger.setLevel(logging.DEBUG)
        elif a.startswith("-"):
            continue
        else:
            positional.append(a)

    services = aep_creds.list_services()
    if not services:
        logger.error("No credentials found in the keyring vault or the creds/ "
                     "folder. Add one with credential_validator_v2.py store "
                     "(or migrate), or drop a <service>.json in creds/.")
        return

    if positional:
        try:
            chosen = aep_creds.pick_service(positional[0])
        except aep_creds.CredsError as e:
            logger.error(str(e))
            return
    else:
        chosen = menu(services)

    if not chosen:
        logger.info("Nothing chosen. Exiting.")
        return

    run(chosen, sandbox_arg, dd_scope, dd_rows,
        profile_snapshot=profile_snapshot)


if __name__ == "__main__":
    main()
