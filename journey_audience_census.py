#!/usr/bin/env python3
"""
journey_audience_census.py  (AEP Swiss Army Knife)
==================================================
A census of the two things that drive customer contact in this org: JOURNEYS
and AUDIENCES (the latter formerly known as segments). Two independent reports
live here, because they answer different questions:

  --list        every Adobe Journey Optimizer journey, and the AUDIENCE behind
                each one -- whether the journey READS an audience (a
                read-audience node) or is triggered by AUDIENCE QUALIFICATION
                (a profile qualifying for a segment). Gives a journey ->
                audience table, and from it a per-audience view of which
                journeys consume each audience.

  --audiences   every audience in each sandbox, with its tags, evaluation type
                and segmentation rule (PQL). Note this is the WHOLE audience
                estate -- the journey report above only ever sees the audiences
                some journey happens to reference.

Credentials come from the shared aep_creds layer: the OS keyring vault first,
falling back to a plaintext creds/ folder where keyring is unavailable. Run
with no --creds to pick a service interactively; manage services with
credential_validator_v2.py.

------------------------------------------------------------------------------
THE HYBRID CREDENTIAL (important -- this is why it works)
------------------------------------------------------------------------------
An AJO journey GET on the platform.adobe.io/ajo gateway checks TWO things
independently, and they can come from DIFFERENT credentials:

  * the Bearer TOKEN  -> supplies the identity/permission. Its technical
    account must be granted AEP sandbox + AJO journey access.
  * the x-api-key     -> identifies the integration to the AJO PRODUCT. This
    key must be SUBSCRIBED to AJO, or you get 403 "Api Key is invalid".

In this org neither single credential has both halves:
  - 'acme alpha'  (an internal test account): tech account HAS journey permission, but
                   its api-key is NOT subscribed to AJO.
  - 'acme beta'  (a sibling integration): api-key IS subscribed to AJO, but
                   its tech account has no permission.

So the working call mints the token from ALPHA's credential and sends BETA's
AJO-subscribed api-key. Verified: that combo returns the full journey JSON
(200), while either credential on its own fails.

You express the hybrid with two values:
  --creds   : the credential whose TOKEN to mint (needs journey permission).
              Omit it to pick from the credential bank interactively.
  --api-key : the AJO-SUBSCRIBED api-key for the x-api-key header.
The api-key can also live in the creds JSON as "api_key" (so you don't repeat
it). When neither is set, the creds file's own client_id is used (the
non-hybrid path, for a credential that already has both halves).

Once an admin properly enables ONE credential for AJO, drop the hybrid and just
point --creds at it.

------------------------------------------------------------------------------
USAGE
------------------------------------------------------------------------------
  python journey_audience_census.py                       # interactive picker, then --list
  python journey_audience_census.py --creds "acme alpha" --api-key <ajo-api-key> --list
  python journey_audience_census.py --creds "acme alpha" --api-key <key> <journeyId> [...]
  python journey_audience_census.py --creds "acme alpha" --api-key <key> --sandbox dev <journeyId>

--list pages the ENTIRE journey estate (~1,300), not just the recently-modified
default. The per-journey GETs (and the Streaming/Batch + tag enrichment) run in
parallel -- --workers N (default 16) -- so the full estate takes minutes, not an
hour. --status live,published,paused filters to those statuses BEFORE the
per-journey GETs (the list payload carries status), so you don't fire a GET per
journey when you only want some. --limit N samples the first N. --no-eval drops
the Streaming/Batch + tag columns for a quick id/name/status-only run.

------------------------------------------------------------------------------
AUDIENCE TAG INVENTORY (--audiences)
------------------------------------------------------------------------------
A separate report over the AUDIENCE estate rather than the journey->audience
view above: every audience in each sandbox, with its tags resolved to names.
Used to track how far a tagging convention has actually been rolled out.

  python journey_audience_census.py --creds "<svc>" --audiences \
      --sandbox "prod,roi-prod,sk-prod,hu-prod,cz-prod" --highlight HST

--sandbox takes a comma list here (or 'all-prod' to discover every production
sandbox from the sandbox-management API). --highlight <text> flags matching tags
in the console and shades them in the workbook.

Two wrinkles this handles, both of which would otherwise distort the counts:
  * Audience tags[] mixes UUID references into the org's Unified Tags vocabulary
    with machine-written "key:value" system stamps (Audience Portal halo
    refreshes). Only the former is human tagging; the latter is counted
    separately in a "System tags" column.
  * The vocabulary is fetched ONCE in bulk (~140 tags) rather than a lookup per
    audience, so thousands of audiences resolve in seconds.

Output: output/audiences_<sandbox>_<date>.xlsx -- one row per audience, with its
tags, evaluation type and segmentation rule (PQL). A "Data completeness" tab
appears only when paging failed for a sandbox, so a short list is never mistaken
for the whole estate. The filename carries no credential name: which credential
read the data says nothing about the data.

Read-only: every call is a GET. Nothing is written to AEP/AJO. Stdlib only.
"""
from __future__ import annotations

import concurrent.futures
import json
import logging
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

import aep_creds  # keyring-first credential store, plaintext creds/ fallback

SCRIPT_NAME    = "journey_audience_census"
SCRIPT_VERSION = "1.1.0"
SCRIPT_DATE    = "2026-07-24"
SCRIPT_AUTHOR  = "Barry Mann (barrymann.com)"

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"

IMS_URL = "https://ims-na1.adobelogin.com/ims/token"
AJO_GATEWAY = "https://platform.adobe.io/ajo"
DEFAULT_SCOPES = (
    "openid,AdobeID,read_organizations,"
    "additional_info.projectedProductContext,session"
)
# The journey LIST endpoint is the SINGULAR /ajo/journey (the plural
# /ajo/journeys is a 404 red herring). It returns {filter, pagination, results}
# where pagination = {page, pageSize, totalCount}.
#
# CRITICAL: with no filter the gateway applies a DEFAULT one
# (metadata.lastModifiedAt > ~7 days ago), silently returning only recently-
# modified journeys (~60). We pass an explicit wide date filter to get the FULL
# estate, then page through with page=0,1,2,... pageSize is a fixed 100 (server
# ignores size/limit params), so a ~1,300-journey estate is ~13 pages.
JOURNEY_LIST_URL = f"{AJO_GATEWAY}/journey"
JOURNEY_LIST_FILTER = "metadata.lastModifiedAt>2000-01-01"  # widen past the default
MAX_JOURNEY_PAGES = 500  # safety cap (~50k journeys) against an infinite paging loop

# ----------------------------------------------------------------------------
# ANSI / logging - matches credential_validator.py style
# ----------------------------------------------------------------------------
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        h = kernel32.GetStdHandle(-11)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
            kernel32.SetConsoleMode(h, mode.value | 0x0004)
    except Exception:
        pass

ANSI = {
    "reset": "\033[0m", "bold": "\033[1m", "dim": "\033[2m",
    "red": "\033[31m", "green": "\033[32m", "yellow": "\033[33m",
    "blue": "\033[34m", "magenta": "\033[35m", "cyan": "\033[36m",
}
LEVEL_COLOR = {
    "DEBUG": ANSI["dim"], "INFO": ANSI["green"],
    "WARNING": ANSI["yellow"],
    "ERROR": ANSI["red"] + ANSI["bold"],
    "CRITICAL": ANSI["red"] + ANSI["bold"],
}


class ColoredFormatter(logging.Formatter):
    def format(self, record):
        color = LEVEL_COLOR.get(record.levelname, "")
        ts = self.formatTime(record, "%H:%M:%S")
        return (
            f"{ANSI['dim']}{ts}{ANSI['reset']} "
            f"{color}[{record.levelname:<7}]{ANSI['reset']} "
            f"{record.getMessage()}"
        )


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(ColoredFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger("journey_audience_census")
SSL_CTX = ssl._create_unverified_context()


def print_banner() -> None:
    bar = ANSI["cyan"] + "=" * 72 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}{SCRIPT_NAME} v{SCRIPT_VERSION}{ANSI['reset']}   ({SCRIPT_DATE})")
    print(f"  by {SCRIPT_AUTHOR}")
    print(f"  {ANSI['dim']}Map Adobe Journey Optimizer journeys to the audiences "
          f"behind them.{ANSI['reset']}")
    print(bar)


# ----------------------------------------------------------------------------
# HTTP / IMS / creds helpers
# ----------------------------------------------------------------------------
def http(url, method="GET", headers=None, data=None, timeout=30):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read(), dict(r.headers)


def clean_detail(raw, limit=160):
    """One-line, CR-stripped slice of a (possibly HTML) error body, so gateway
    error pages can't reset the terminal cursor and garble a line."""
    text = raw.decode(errors="replace") if isinstance(raw, bytes) else raw
    return " ".join(text.split())[:limit]


def shorten(s, n=12):
    if not s:
        return "?"
    return s if len(s) <= n else f"{s[:n]}..."


def authenticate(conf) -> str:
    """OAuth server-to-server. ALWAYS minted fresh -- AEP/AJO permissions are
    snapshotted into the token at mint time, so a stale token never reflects a
    just-granted product profile. Returns the access token string."""
    payload = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": conf["client_id"],
        "client_secret": conf["client_secret"],
        "scope": conf.get("scopes") or DEFAULT_SCOPES,
    }).encode("utf-8")
    body, _ = http(IMS_URL, method="POST",
                   headers={"Content-Type": "application/x-www-form-urlencoded"},
                   data=payload)
    return json.loads(body)["access_token"]


def ajo_headers(token, api_key, org_id, sandbox):
    return {
        "Authorization": f"Bearer {token}",
        "x-api-key": api_key,
        "x-gw-ims-org-id": org_id,
        "x-sandbox-name": sandbox,
        "Accept": "application/json",
    }


# ----------------------------------------------------------------------------
# Discovery / menu - matches credential_validator.py
# ----------------------------------------------------------------------------
def menu(services):
    """Single-pick credential bank (the token credential). Operates on keyring
    service-name strings (keyring vault + creds/ fallback, via aep_creds)."""
    print()
    bar = ANSI["cyan"] + "=" * 70 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Credential bank{ANSI['reset']}  "
          f"{ANSI['dim']}(OS keyring vault){ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 70 + ANSI["reset"])
    for i, name in enumerate(services, 1):
        print(f"  {ANSI['bold']}{i:>2}{ANSI['reset']}  "
              f"{ANSI['yellow']}{name:<20}{ANSI['reset']} "
              f"{ANSI['dim']}{name}{ANSI['reset']}")
    print(bar)
    raw = input(
        f"\nPick the token credential by number "
        f"({ANSI['cyan']}1{ANSI['reset']}), blank to quit: "
    ).strip()
    if not raw:
        return None
    tok = raw.replace(",", " ").split()[0]
    if tok.isdigit() and 1 <= int(tok) <= len(services):
        return services[int(tok) - 1]
    logger.warning(f"Invalid choice: {tok}")
    return None


def resolve_service(services, name):
    """Match a --creds arg to a service name: exact, else unique substring."""
    if name in services:
        return name
    term = name.lower()
    hits = [s for s in services if term in s.lower()]
    return hits[0] if len(hits) == 1 else None


def api_key_source(api_key, services, org_id=None):
    """Which credential set supplies this x-api-key? Matches the key against
    each set's client_id / api_key so we can name it in the header instead of
    showing a bare hex prefix. The same client_id can appear in more than one
    org, so a match in the SAME org as the token wins (avoids naming a
    different-org set with a shared key). Uses peek_creds so scanning the whole
    bank doesn't spew a source/security line per service."""
    fallback = None
    for name in services:
        c = aep_creds.peek_creds(name)
        if not c:
            continue
        if api_key in (c.get("client_id"), c.get("api_key")):
            if org_id and c.get("org_id") == org_id:
                return name
            fallback = fallback or name
    return fallback


def pick_api_key(services, token_service):
    """Interactively choose the AJO-subscribed x-api-key (the hybrid). AJO
    checks the api-key is subscribed to the product, and that key is often a
    DIFFERENT credential than the one whose token you mint. Returns an api-key
    string, or None to fall back to the token credential's own key."""
    others = [s for s in services if s != token_service]
    print()
    print(f"  {ANSI['bold']}AJO needs an AJO-subscribed api-key{ANSI['reset']} "
          f"{ANSI['dim']}(the x-api-key; can differ from the token credential)"
          f"{ANSI['reset']}")
    for i, name in enumerate(others, 1):
        print(f"  {ANSI['bold']}{i:>2}{ANSI['reset']}  "
              f"{ANSI['yellow']}{name:<20}{ANSI['reset']} "
              f"{ANSI['dim']}use this set's api-key{ANSI['reset']}")
    raw = input(
        f"\nPick the AJO-subscribed set by number, paste a key, or Enter to use "
        f"{ANSI['yellow']}{token_service}{ANSI['reset']}'s own key: "
    ).strip()
    if not raw:
        return None
    if raw.isdigit() and 1 <= int(raw) <= len(others):
        c = aep_creds.peek_creds(others[int(raw) - 1])
        if c:
            return c.get("api_key") or c.get("client_id")
        logger.warning("Could not load that set.")
        return None
    return raw  # treat anything else as a pasted api-key


# ----------------------------------------------------------------------------
# Journey fetch + audience extraction
# ----------------------------------------------------------------------------
AUDIENCES_URL = "https://platform.adobe.io/data/core/ups/audiences"
# Tag names live in the Unified Tags service on a DIFFERENT host
# (experience.adobe.io); a journey only carries tag ids with name=null.
UNIFIED_TAGS_URL = "https://experience.adobe.io/unifiedtags/tags"


def get_journey(token, api_key, conf, sandbox, journey_id):
    """GET one journey by id. Returns (ok, journey_dict_or_error_string)."""
    url = f"{AJO_GATEWAY}/journey/{journey_id}"
    try:
        body, _ = http(url, headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=30)
        return True, json.loads(body)
    except urllib.error.HTTPError as e:
        return False, f"HTTP {e.code}: {clean_detail(e.read())}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def audience_type(token, api_key, conf, sandbox, audience_id, cache):
    """Streaming / Batch / Edge for one audience, from AEP's audiences API
    (evaluationInfo). Cached -- audiences repeat across journeys. Returns a
    short label; failures degrade to a marker rather than raising."""
    if audience_id in cache:
        return cache[audience_id]
    url = f"{AUDIENCES_URL}/{audience_id}"
    label = "?"
    try:
        body, _ = http(url, headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=20)
        ei = (json.loads(body) or {}).get("evaluationInfo") or {}
        if ei.get("continuous", {}).get("enabled"):
            label = "Streaming"
        elif ei.get("synchronous", {}).get("enabled"):
            label = "Edge"
        elif ei.get("batch", {}).get("enabled"):
            label = "Batch"
        else:
            label = "unknown"
    except urllib.error.HTTPError as e:
        label = "not-found" if e.code == 404 else ("no-access" if e.code == 403 else f"err{e.code}")
    except Exception:
        label = "err"
    cache[audience_id] = label
    return label


def tag_name(token, api_key, conf, sandbox, tag_id, cache):
    """Resolve a journey tag id to its name via the Unified Tags service.
    Cached -- tags repeat across journeys. Falls back to the short id."""
    if tag_id in cache:
        return cache[tag_id]
    name = tag_id[:8]
    try:
        body, _ = http(f"{UNIFIED_TAGS_URL}/{tag_id}",
                       headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=20)
        name = (json.loads(body) or {}).get("name") or tag_id[:8]
    except Exception:
        name = tag_id[:8]
    cache[tag_id] = name
    return name


def journey_tags(token, api_key, conf, sandbox, journey, cache):
    """Comma-joined tag NAMES for a journey (its tags[] carries ids only)."""
    ids = [t.get("id") for t in (journey.get("tags") or [])
           if isinstance(t, dict) and t.get("id")]
    return ", ".join(tag_name(token, api_key, conf, sandbox, t, cache) for t in ids)


# ----------------------------------------------------------------------------
# Audience TAG inventory (--audiences)
# ----------------------------------------------------------------------------
# Audiences carry tags in two different shapes on the SAME tags[] list:
#   * a bare UUID          -> a reference into the org's Unified Tags vocabulary;
#                             the human name only exists in that service.
#   * a "key:value" string -> a machine-written system tag (Audience Portal's
#                             halo refresh stamps, etc). Not user tagging, and it
#                             would swamp the report if counted as such.
# We resolve the former and quarantine the latter.
SANDBOX_LIST_URL = ("https://platform.adobe.io/data/foundation/"
                    "sandbox-management/sandboxes")
SYSTEM_TAG_PREFIXES = ("audience_portal_",)


def list_sandboxes(token, api_key, conf, production_only=True):
    """Sandbox names visible to the credential. production_only keeps type=production."""
    body, _ = http(SANDBOX_LIST_URL,
                   headers=ajo_headers(token, api_key, conf["org_id"], "prod"),
                   timeout=60)
    sbs = (json.loads(body) or {}).get("sandboxes") or []
    return [s.get("name") for s in sbs
            if s.get("name") and s.get("state") == "active"
            and (not production_only or s.get("type") == "production")]


def fetch_tag_vocabulary(token, api_key, conf, sandbox):
    """{tag_id: tag_name} for the whole org, paged from the Unified Tags service.
    One bulk walk instead of a GET per tag id -- the vocabulary is small (~140)
    while the audiences referencing it run to thousands."""
    vocab, cursor, guard = {}, None, 0
    while guard < 100:
        guard += 1
        url = f"{UNIFIED_TAGS_URL}?limit=100"
        if cursor:
            url += f"&start={urllib.parse.quote(str(cursor))}"
        try:
            body, _ = http(url, headers=ajo_headers(token, api_key,
                                                    conf["org_id"], sandbox),
                           timeout=60)
            page = json.loads(body) or {}
        except Exception as e:
            logger.warning(f"  tag vocabulary page {guard} failed "
                           f"({type(e).__name__}: {e}); names may fall back to ids.")
            break
        items = page.get("tags") or []
        if not items:
            break
        for t in items:
            if t.get("id"):
                vocab[t["id"]] = t.get("name") or t["id"][:8]
        cursor = (page.get("_page") or {}).get("next")
        if not cursor:
            break
    return vocab


USER_DIRECTORY_URL = "https://usermanagement.adobe.io/v2/usermanagement/users"


def fetch_user_directory(token, api_key, conf):
    """{ims_user_id: email} for the org, paged from User Management.

    Audiences record who created/modified them as an opaque IMS id
    (D82F225C...@805f1e8e...), which tells a reader nothing. The directory is a
    couple of pages for a few thousand users, so it is fetched once per run and
    used to turn those ids into people. Empty dict on failure -- the columns
    then show raw ids rather than the report dying over a nicety.
    """
    users, page = {}, 0
    while page < 50:
        url = f"{USER_DIRECTORY_URL}/{conf['org_id']}/{page}"
        try:
            # Org-level endpoint: no sandbox header (and urllib rejects a None
            # header value, so it is dropped rather than passed empty).
            hdrs = {k: v for k, v in ajo_headers(token, api_key, conf["org_id"],
                                                 "prod").items()
                    if k != "x-sandbox-name"}
            body, _ = http(url, headers=hdrs, timeout=60)
            r = json.loads(body) or {}
        except Exception as e:
            logger.warning(f"  user directory page {page} failed "
                           f"({type(e).__name__}); creator columns will show "
                           f"raw ids.")
            break
        for u in r.get("users") or []:
            if u.get("id"):
                users[u["id"]] = u.get("email") or u.get("username") or ""
        if r.get("lastPage"):
            break
        page += 1
    return users


def resolve_actor(actor_id, directory):
    """An IMS id turned into something a human can act on.

    Three kinds appear on audiences and only the first is a person:
      * a directory user        -> their email
      * @techacct.adobe.com     -> an API integration, not somebody's doing
      * @AdobeID service names  -> Adobe's own automation (halo refreshes etc)
    """
    if not actor_id:
        return ""
    actor_id = str(actor_id)
    if actor_id in directory and directory[actor_id]:
        return directory[actor_id]
    if actor_id.endswith("@techacct.adobe.com"):
        return f"(API integration {actor_id.split('@')[0][:8]})"
    if actor_id.endswith("@AdobeID"):
        return f"(Adobe service: {actor_id.split('@')[0]})"
    return actor_id


def fetch_all_audiences(token, api_key, conf, sandbox):
    """Every audience in a sandbox, paged in full. Returns (audiences, complete).

    `complete` is False when the walk ended before totalCount -- the caller must
    say so rather than presenting a short list as the whole estate.
    """
    out, seen, cursor, guard, total = [], set(), None, 0, None
    while guard < 500:
        guard += 1
        url = f"{AUDIENCES_URL}?limit=100"
        if cursor is not None:
            url += f"&start={urllib.parse.quote(str(cursor))}"
        try:
            body, _ = http(url, headers=ajo_headers(token, api_key,
                                                    conf["org_id"], sandbox),
                           timeout=90)
            page = json.loads(body) or {}
        except Exception as e:
            logger.warning(f"  {sandbox}: audience page {guard} failed "
                           f"({type(e).__name__}: {e}); list is INCOMPLETE.")
            return out, False
        kids = page.get("children") or []
        meta = page.get("_page") or {}
        if total is None:
            total = meta.get("totalCount")
        for c in kids:
            if c.get("id") and c["id"] not in seen:
                seen.add(c["id"])
                out.append(c)
        if not kids:
            break
        cursor = meta.get("next")
        if cursor is None:
            break
    complete = total is None or len(out) >= int(total)
    if not complete:
        logger.warning(f"  {sandbox}: collected {len(out)} of {total} audience(s) "
                       f"-- list is INCOMPLETE.")
    return out, complete


# ----------------------------------------------------------------------------
# PQL rendering
# ----------------------------------------------------------------------------
# An audience's rule comes back as 'pql/json': Adobe's syntax TREE, not the PQL
# you see in the UI. There is no supported way to ask for the text form -- every
# format/expressionFormat parameter still returns json, and the conversion
# endpoint is not available to a read-only credential -- so we render it here.
#
# The grammar is large (15 node types, 34 functions, trees up to 17 deep). We
# render the shapes that make up the overwhelming majority -- boolean logic,
# comparisons, field paths, segment references -- and REFUSE to guess at the
# rest. Event-sequence and time-window nodes emit a visible <...> marker and set
# the "partial" flag, so a half-rendered rule can never be mistaken for a whole
# one. The raw tree is always kept in its own column.
PQL_INFIX = {"and": "and", "or": "or", "=": "=", ">": ">", "<": "<",
             ">=": ">=", "<=": "<=", "!=": "!=", "equals": "=",
             "notEqualTo": "!="}
PQL_WORDY = {"startsWith": "starts with", "endsWith": "ends with",
             "contains": "contains", "doesNotContain": "does not contain",
             "in": "in", "notIn": "not in"}
# Nodes describing event sequences / time windows. Rendering these faithfully is
# a project in itself; misrendering one would quietly change what the rule means.
PQL_COMPLEX = {"chain", "occurs", "timeQualification", "duration", "select",
               "varDecl", "element", "gap", "range"}
XL_CELL_LIMIT = 32000          # Excel's hard ceiling is 32767


def _pql_literal(node):
    v = node.get("value")
    lt = node.get("literalType")
    if lt == "List" and isinstance(v, list):
        return "[" + ", ".join(_pql_literal({"value": x, "literalType": "String"}
                                            if isinstance(x, str) else {"value": x})
                               for x in v) + "]"
    if isinstance(v, str) and lt in ("String", "Timestamp", "TimeDirection",
                                     "TimeUnit", "Comparison", None):
        return f'"{v}"'
    if isinstance(v, bool):
        return "true" if v else "false"
    return str(v)


def _pql_path(node, state):
    """Dotted field path. The base is a parameter/var reference and contributes
    nothing readable, so it is dropped: v0._experience.x -> _experience.x"""
    parts = []
    cur = node
    while isinstance(cur, dict) and cur.get("nodeType") == "fieldLookup":
        parts.append(cur.get("fieldName") or "?")
        cur = cur.get("object")
    if isinstance(cur, dict) and cur.get("nodeType") not in (
            "parameterReference", "varRef", None):
        head = _render_node(cur, state)
        if head:
            parts.append(head)
    return ".".join(reversed(parts))


def _render_node(node, state):
    if node is None:
        return ""
    if isinstance(node, list):
        return ", ".join(_render_node(n, state) for n in node)
    if not isinstance(node, dict):
        return str(node)

    nt = node.get("nodeType")
    if nt in PQL_COMPLEX:
        state["partial"] = True
        state["complex"].add(nt)
        return f"<{nt}: see raw>"
    if nt == "literal":
        return _pql_literal(node)
    if nt == "fieldLookup":
        return _pql_path(node, state)
    if nt in ("parameterReference", "varRef"):
        return ""                       # the implicit profile/event being tested
    if nt == "lambda":
        return _render_node(node.get("body"), state)
    if nt != "fnApply":
        state["partial"] = True
        state["complex"].add(nt or "unknown")
        return f"<{nt or 'unknown'}: see raw>"

    fn = node.get("fnName")
    params = node.get("params") or []

    if fn == "inSegment" and params:
        sid = (params[0] or {}).get("value")
        name = state["segments"].get(sid)
        return f'inSegment("{name}")' if name else f'inSegment({sid})'
    if fn == "not" and len(params) == 1:
        return f"not ({_render_node(params[0], state)})"
    if fn in ("exists", "isNotNull") and params:
        return f"{_render_node(params[0], state)} exists"
    if fn == "isNull" and params:
        return f"{_render_node(params[0], state)} is null"
    # stringCompare NAMES its comparison in the first parameter -- it is
    # stringCompare(<op>, <field>, <value>), not an infix pair. Treating it as
    # infix renders the operator itself as the left-hand operand.
    if fn == "stringCompare" and len(params) >= 3:
        op = (params[0] or {}).get("value") or "compares"
        return (f"{_render_node(params[1], state)} {op} "
                f"{_render_node(params[2], state)}")
    if fn == "get" and len(params) >= 2:
        base = _render_node(params[0], state)
        fld = (params[1] or {}).get("value") or _render_node(params[1], state)
        return f"{base}.{fld}" if base else str(fld)
    if fn in PQL_INFIX and len(params) >= 2:
        op = PQL_INFIX[fn]
        # equals/stringCompare carry a trailing case-sensitivity flag
        operands = params[:2] if fn in ("equals", "stringCompare") else params
        rendered = [_render_node(p, state) for p in operands]
        joined = f" {op} ".join(r for r in rendered if r != "")
        return f"({joined})" if op in ("and", "or") and len(rendered) > 1 else joined
    if fn in PQL_WORDY and len(params) >= 2:
        return (f"{_render_node(params[0], state)} {PQL_WORDY[fn]} "
                f"{_render_node(params[1], state)}")
    args = ", ".join(r for r in (_render_node(p, state) for p in params) if r != "")
    return f"{fn}({args})"


def render_pql(expression, segments):
    """(readable_text, raw_text, partial) for an audience's expression.

    partial=True means at least one node could not be rendered faithfully and
    appears as a <marker>; the row is labelled so nobody reads it as complete.
    """
    if not expression:
        return "", "", False
    raw = expression.get("value")
    raw_text = raw if isinstance(raw, str) else json.dumps(raw)
    if expression.get("format") == "pql/text":
        return str(raw_text), str(raw_text), False
    try:
        tree = json.loads(raw_text) if isinstance(raw_text, str) else raw_text
    except (ValueError, TypeError):
        return "", str(raw_text), True
    state = {"segments": segments, "partial": False, "complex": set()}
    text = _render_node(tree, state)
    if state["partial"] and state["complex"]:
        text = f"[PARTIAL: {', '.join(sorted(state['complex']))}] {text}"
    return text, str(raw_text), state["partial"]


def audience_eval_type(aud: dict) -> str:
    """Streaming / Edge / Batch from the evaluationInfo carried on the LIST
    payload (no per-audience GET needed). '' when the record omits it."""
    ei = aud.get("evaluationInfo") or {}
    if (ei.get("continuous") or {}).get("enabled"):
        return "Streaming"
    if (ei.get("synchronous") or {}).get("enabled"):
        return "Edge"
    if (ei.get("batch") or {}).get("enabled"):
        return "Batch"
    return ""


def split_audience_tags(aud: dict, vocab: dict):
    """(named_tags, system_tags, unresolved_ids) for one audience.

    named  -- resolved against the Unified Tags vocabulary: the human tagging
              this report is actually about.
    system -- machine-written key:value stamps, kept out of the counts.
    unresolved -- UUIDs with no vocabulary entry (deleted tag, or a tag the
              credential cannot see). Reported rather than silently dropped.
    """
    named, system, unresolved = [], [], []
    for t in (aud.get("tags") or []):
        t = str(t)
        if ":" in t or t.startswith(SYSTEM_TAG_PREFIXES):
            system.append(t)
        elif t in vocab:
            named.append(vocab[t])
        else:
            unresolved.append(t)
    return sorted(set(named)), system, unresolved


def build_audience_rows(token, api_key, conf, sandboxes, highlight=""):
    """Walk each sandbox and return (rows, tally, incomplete) for the report."""
    rows, tally, incomplete = [], {}, []
    vocab, directory = {}, None
    for sb in sandboxes:
        if not vocab:
            vocab = fetch_tag_vocabulary(token, api_key, conf, sb)
            logger.info(f"  tag vocabulary: {len(vocab)} tag(s) in the org.")
        if directory is None:
            directory = fetch_user_directory(token, api_key, conf)
            logger.info(f"  user directory: {len(directory)} user(s) "
                        f"(for the created/modified-by columns).")
        logger.info(f"  {sb}: listing audiences ...")
        auds, complete = fetch_all_audiences(token, api_key, conf, sb)
        if not complete:
            incomplete.append(sb)
        # inSegment() references another audience by id -- resolve to its name so
        # the rule reads as prose. Built per sandbox: ids are sandbox-scoped.
        seg_names = {}
        for a in auds:
            for key in ("id", "audienceId"):
                if a.get(key):
                    seg_names[a[key]] = a.get("name") or ""
        tagged = n_pql = n_partial = 0
        for a in auds:
            named, system, unresolved = split_audience_tags(a, vocab)
            if named:
                tagged += 1
            for n in named:
                tally.setdefault(n, {}).setdefault(sb, 0)
                tally[n][sb] += 1
            pql, pql_raw, partial = render_pql(a.get("expression"), seg_names)
            if pql:
                n_pql += 1
                n_partial += bool(partial)
            rows.append([
                sb, a.get("name") or "", a.get("id") or "",
                audience_eval_type(a), a.get("lifecycleState") or "",
                ", ".join(named),
                len(named),
                a.get("namespace") or "",
                resolve_actor(a.get("createdBy"), directory or {}),
                resolve_actor(a.get("lastModifiedBy"), directory or {}),
                pql[:XL_CELL_LIMIT],
                "partial" if partial else ("yes" if pql else ""),
                pql_raw[:XL_CELL_LIMIT],
                ", ".join(t[:8] for t in unresolved),
                len(system),
            ])
        logger.info(f"  {sb}: {n_pql} audience(s) carry a rule "
                    f"({n_partial} only partly renderable).")
        hi = ""
        if highlight:
            n_hi = sum(1 for r in rows if r[0] == sb
                       and highlight.lower() in str(r[5]).lower())
            hi = f", {n_hi} matching {highlight!r}"
        pct = (tagged / len(auds) * 100) if auds else 0
        logger.info(f"  {sb}: {len(auds)} audience(s), {tagged} tagged "
                    f"({pct:.0f}%){hi}.")
    return rows, tally, incomplete


def write_audience_xlsx(rows, tally, sandboxes, out_path, incomplete=(),
                        highlight=""):
    """One row per audience: tags, evaluation type and its rule (PQL).

    A "Data completeness" tab is added only when a sandbox's listing failed, so
    a short list is never mistaken for the whole estate. `tally` is still used
    for the console tag summary -- it just doesn't get a tab of its own."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    hit_fill = PatternFill("solid", fgColor="FFF2CC")

    def style(ws, headers, widths):
        for c in ws[1]:
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = hdr_fill
            c.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    wb = Workbook()

    # --- Every audience, with its tags. FIRST sheet on purpose: this is the
    # one people came for, and a workbook that opens on a summary tab reads as
    # though the per-audience tags live somewhere else. ---
    ws2 = wb.active
    ws2.title = "Audiences"
    headers2 = ["Sandbox", "Audience name", "Audience id", "Evaluation",
                "Lifecycle", "Tags", "Tag count", "Origin", "Created by",
                "Last modified by", "PQL (readable)", "PQL rendered",
                "PQL (raw pql/json)", "Unresolved tag ids", "System tags"]
    ws2.append(headers2)
    for r in rows:
        ws2.append(r)
        if highlight and highlight.lower() in str(r[5]).lower():
            for c in ws2[ws2.max_row]:
                c.fill = hit_fill
    style(ws2, headers2,
          (12, 52, 38, 12, 14, 40, 11, 22, 34, 34, 80, 13, 60, 24, 12))
    # Deliberately NOT wrapping the rule column. Wrapped text makes Excel
    # auto-fit each row to the tallest cell, and a 30,000-character PQL turns
    # one row into a screenful -- the grid stops being scannable. Rows stay one
    # line high; the full rule is still in the cell (and in the raw column) for
    # anyone who widens it or clicks in.
    ws2.sheet_format.defaultRowHeight = 15
    pql_col = headers2.index("PQL (readable)") + 1
    for row in ws2.iter_rows(min_row=2, min_col=pql_col, max_col=pql_col):
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=False)

    if incomplete:
        ws3 = wb.create_sheet("Data completeness")
        ws3.append(["Sandbox", "Warning"])
        for sb in incomplete:
            ws3.append([sb, "Audience list INCOMPLETE -- paging failed; counts "
                            "for this sandbox understate the estate."])
        style(ws3, ["Sandbox", "Warning"], (16, 90))

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def pick_sandbox(token, api_key, conf, default="prod"):
    """Ask which sandbox to report on, listing what the credential can see.
    Enter takes the default. Falls back to the default off a terminal."""
    if not sys.stdin.isatty():
        return [default]
    try:
        names = list_sandboxes(token, api_key, conf, production_only=False)
    except Exception as e:
        logger.warning(f"Could not list sandboxes ({type(e).__name__}); "
                       f"using {default}.")
        return [default]
    if not names:
        return [default]
    print(f"\n  {ANSI['bold']}Which sandbox?{ANSI['reset']}")
    for i, n in enumerate(names, 1):
        mark = f"  {ANSI['dim']}(default){ANSI['reset']}" if n == default else ""
        print(f"    {i:>2}. {n}{mark}")
    print(f"    {ANSI['dim']}Enter = {default}; or type a name, a number, "
          f"or a comma list{ANSI['reset']}")
    try:
        raw = input("  > ").strip()
    except (EOFError, KeyboardInterrupt):
        return [default]
    if not raw:
        return [default]
    chosen = []
    for part in raw.split(","):
        part = part.strip()
        if not part:
            continue
        if part.isdigit() and 1 <= int(part) <= len(names):
            chosen.append(names[int(part) - 1])
        elif part in names:
            chosen.append(part)
        else:
            logger.warning(f"No sandbox named {part!r}; ignoring it.")
    return chosen or [default]


def run_audience_report(service, conf, token, api_key, opts):
    """--audiences: tag inventory across one or more sandboxes."""
    sandboxes = opts["sandboxes"]
    if sandboxes == ["all-prod"]:
        sandboxes = list_sandboxes(token, api_key, conf, production_only=True)
        logger.info(f"  production sandboxes: {', '.join(sandboxes)}")
    elif sandboxes is None:
        # No --sandbox given: ask rather than assuming the whole estate.
        sandboxes = pick_sandbox(token, api_key, conf)
        logger.info(f"  sandbox(es): {', '.join(sandboxes)}")
    highlight = opts["highlight"]
    rows, tally, incomplete = build_audience_rows(
        token, api_key, conf, sandboxes, highlight=highlight)

    print()
    logger.info(f"{len(rows)} audience(s) across {len(sandboxes)} sandbox(es); "
                f"{len(tally)} distinct tag(s) in use.")
    if highlight:
        hits = {t: v for t, v in tally.items() if highlight.lower() in t.lower()}
        if hits:
            for t, v in sorted(hits.items()):
                spread = ", ".join(f"{sb}={n}" for sb, n in sorted(v.items()))
                logger.info(f"  {ANSI['green']}{highlight} match:{ANSI['reset']} "
                            f"{t} -- {sum(v.values())} audience(s) ({spread})")
        else:
            logger.warning(f"  no tag matching {highlight!r} is applied to any "
                           f"audience in these sandbox(es).")
    if incomplete:
        logger.warning(f"  INCOMPLETE sandbox(es): {', '.join(incomplete)} -- "
                       f"counts understate the estate.")

    # No credential name in the filename: which credential happened to read the
    # data says nothing about the data. What it IS and when it was taken do.
    stamp = datetime.now().strftime("%Y-%m-%d")
    sb_part = "_".join(sandboxes) if len(sandboxes) <= 2 else f"{len(sandboxes)}sandboxes"
    default = OUTPUT_DIR / f"audiences_{sb_part}_{stamp}.xlsx"
    out_path = Path(opts["out"]) if opts["out"] else default
    try:
        write_audience_xlsx(rows, tally, sandboxes, out_path,
                            incomplete=incomplete, highlight=highlight)
        logger.info(f"Wrote {out_path}")
    except PermissionError:
        alt = out_path.with_name(out_path.stem + "_new" + out_path.suffix)
        write_audience_xlsx(rows, tally, sandboxes, alt,
                            incomplete=incomplete, highlight=highlight)
        logger.warning(f"{out_path.name} is locked (open in Excel?) - wrote "
                       f"{alt.name} instead.")


def extract_audiences(journey: dict) -> list[dict]:
    """Pull every audience/segment reference out of a journey definition,
    regardless of how it's used. Returns a de-duped list of
    {audience_id, audience_name, via} dicts.

    Two shapes are handled, plus a recursive catch-all so we don't miss a
    variant we haven't seen yet:
      * read-audience nodes: a node carrying an "audiences":[{id,name}] list.
      * audience qualification: an entry/event referencing a segment, seen as
        segmentId / audienceId / segment.{id,name} keys.
    """
    found: dict[str, dict] = {}

    def add(aid, name, via):
        if not aid:
            return
        cur = found.get(aid)
        if cur is None:
            found[aid] = {"audience_id": aid, "audience_name": name or "", "via": via}
        else:
            if name and not cur["audience_name"]:
                cur["audience_name"] = name
            if via not in cur["via"]:
                cur["via"] = f"{cur['via']}+{via}"

    def walk(obj, node_type=None):
        if isinstance(obj, dict):
            nt = obj.get("type") or node_type
            auds = obj.get("audiences")
            if isinstance(auds, list):
                for a in auds:
                    if isinstance(a, dict):
                        # Label by the actual node type: a read-audience node vs
                        # an 'audience_qualification' entry/trigger node (confirmed
                        # node type); fall back to a generic tag otherwise.
                        ntl = str(nt).lower() if nt else ""
                        via = ("read_audience" if "read" in ntl
                               else "qualification" if "qualif" in ntl
                               else "audiences")
                        add(a.get("id") or a.get("audienceId"), a.get("name"), via)
            for key in ("segmentId", "audienceId"):
                if obj.get(key):
                    add(obj[key], obj.get("name"), "qualification")
            seg = obj.get("segment") or obj.get("audience")
            if isinstance(seg, dict) and (seg.get("id") or seg.get("segmentId")):
                add(seg.get("id") or seg.get("segmentId"), seg.get("name"), "qualification")
            for v in obj.values():
                walk(v, nt)
        elif isinstance(obj, list):
            for v in obj:
                walk(v, node_type)

    walk(journey)
    return list(found.values())


def list_journeys(token, api_key, conf, sandbox):
    """Page through the ENTIRE AJO journey list. Returns (ids, by_id) where
    by_id maps journey id -> {"name", "status"} taken from the list items, so a
    --status filter can run BEFORE the per-journey GET loop.

    Follows the gateway's page-based pagination (pagination.{page,pageSize,
    totalCount}) with an explicit wide date filter, until a page comes back
    empty or we've collected totalCount (whichever first), capped at
    MAX_JOURNEY_PAGES."""
    print(f"  {ANSI['bold']}Listing journeys{ANSI['reset']} "
          f"{ANSI['dim']}(sandbox '{sandbox}', paging the full estate){ANSI['reset']}")
    by_id: dict[str, dict] = {}
    total = None
    page = 0
    while page < MAX_JOURNEY_PAGES:
        params = {"filter": JOURNEY_LIST_FILTER, "page": page}
        url = f"{JOURNEY_LIST_URL}?{urllib.parse.urlencode(params)}"
        try:
            body, _ = http(url, headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=60)
        except urllib.error.HTTPError as e:
            detail = clean_detail(e.read(), 90)
            print(f"     {ANSI['yellow']}[{e.code}] page {page} -> {detail}{ANSI['reset']}")
            if e.code == 403 and ("api key is invalid" in detail.lower() or "403003" in detail):
                print(f"     {ANSI['dim']}      ^ that x-api-key isn't subscribed to AJO "
                      f"(the hybrid) -- pass an AJO-subscribed --api-key.{ANSI['reset']}")
            break
        except Exception as e:
            print(f"     {ANSI['red']}[ERR] page {page} -> {type(e).__name__}: {e}{ANSI['reset']}")
            break
        try:
            data = json.loads(body)
        except Exception:
            print(f"     {ANSI['yellow']}[200] page {page}: non-JSON body{ANSI['reset']}")
            break
        results = data.get("results") if isinstance(data, dict) else (data if isinstance(data, list) else None)
        if not results:
            break
        for it in results:
            if isinstance(it, dict) and it.get("id"):
                by_id[it["id"]] = {"name": it.get("name", "?"), "status": it.get("status") or "?"}
        pg = (data.get("pagination") or {}) if isinstance(data, dict) else {}
        if pg.get("totalCount") is not None:
            total = pg["totalCount"]
        page += 1
        # progress for a long multi-page walk
        print(f"     {ANSI['dim']}page {page}: {len(by_id)}"
              f"{('/' + str(total)) if total is not None else ''} journeys...{ANSI['reset']}")
        if total is not None and len(by_id) >= total:
            break
    else:
        print(f"     {ANSI['yellow']}(stopped at page cap {MAX_JOURNEY_PAGES}){ANSI['reset']}")
    shown_total = total if total is not None else len(by_id)
    print(f"     {ANSI['green']}[OK] {len(by_id)} journeys across {page} page(s) "
          f"(totalCount {shown_total}){ANSI['reset']}")
    return list(by_id.keys()), by_id


# ----------------------------------------------------------------------------
# Output
# ----------------------------------------------------------------------------
def print_table(rows):
    """rows: list of (journey_id, journey_name, audience_id, audience_name, via, type)."""
    if not rows:
        print(f"  {ANSI['dim']}(no audiences found){ANSI['reset']}")
        return
    headers = ("Journey id", "Journey name", "Status", "Tags",
               "Audience id", "Audience name", "Via", "Type")
    caps = (32, 30, 9, 20, 32, 28, 13, 9)
    colors = ("dim", "yellow", "yellow", "blue", "dim", "cyan", "magenta", "green")
    n = len(headers)
    widths = [min(caps[i], max(len(str(r[i])) for r in [headers] + rows)) for i in range(n)]

    def cell(text, w, color=None):
        s = str(text)[:w].ljust(w)
        return f"{ANSI[color]}{s}{ANSI['reset']}" if color else s

    print("  " + " | ".join(f"{ANSI['bold']}{str(h)[:widths[i]].ljust(widths[i])}{ANSI['reset']}"
                            for i, h in enumerate(headers)))
    print("  " + ANSI["cyan"] + "-+-".join("-" * w for w in widths) + ANSI["reset"])
    for r in rows:
        print("  " + " | ".join(cell(c, widths[i], colors[i]) for i, c in enumerate(r)))


def write_xlsx(rows, out_path, subtitle=""):
    """Write the journey -> audience rows to a single-sheet XLSX (house style:
    bold banded header, frozen top row, autofilter, sized columns)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Journeys"
    headers = ["Journey id", "Journey name", "Status", "Tags",
               "Audience id", "Audience name", "Via", "Audience type"]
    ws.append(headers)
    for r in rows:
        ws.append([str(c) for c in r])

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for i, w in enumerate((38, 46, 12, 26, 38, 42, 16, 16), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if subtitle:
        ws.oddHeader.left.text = subtitle

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def run_checker(service, opts, services):
    bar = ANSI["cyan"] + "=" * 70 + ANSI["reset"]
    print()
    print(bar)
    print(f"  {ANSI['bold']}Journey + Audience Census{ANSI['reset']}  "
          f"{ANSI['yellow']}{service}{ANSI['reset']} "
          f"{ANSI['dim']}({service}){ANSI['reset']}")
    print(bar)

    try:
        conf = aep_creds.load_creds(service)
    except aep_creds.CredsError as e:
        logger.error(f"Failed to load credentials for {service!r}: {e}")
        return

    api_key = opts["api_key"] or conf.get("api_key") or conf["client_id"]
    hybrid = api_key != conf["client_id"]
    key_from = api_key_source(api_key, services, conf["org_id"]) or ("pasted key" if hybrid else service)
    # Spell out the two credentials by NAME so the hybrid isn't confusing.
    print(f"  {ANSI['bold']}Token   from:{ANSI['reset']} {ANSI['yellow']}{service}{ANSI['reset']} "
          f"{ANSI['dim']}(client {shorten(conf['client_id'])} - gives the journey permission){ANSI['reset']}")
    print(f"  {ANSI['bold']}Api-key from:{ANSI['reset']} {ANSI['yellow']}{key_from}{ANSI['reset']} "
          f"{ANSI['dim']}(key {shorten(api_key)} - must be subscribed to AJO){ANSI['reset']}"
          + (f"  {ANSI['cyan']}[hybrid]{ANSI['reset']}" if hybrid else ""))
    print(f"  {ANSI['bold']}Org / sandbox:{ANSI['reset']} {ANSI['magenta']}{conf['org_id']}{ANSI['reset']}"
          f"  {ANSI['dim']}/{ANSI['reset']} {opts['sandbox']}")
    print()

    try:
        token = authenticate(conf)
    except urllib.error.HTTPError as e:
        logger.error(f"IMS auth FAILED: HTTP {e.code} {clean_detail(e.read())}")
        return
    except Exception as e:
        logger.error(f"IMS auth FAILED: {type(e).__name__}: {e}")
        return
    print(f"  {ANSI['green']}[OK] token minted fresh{ANSI['reset']}")
    print()

    # --audiences is a standalone report over the AUDIENCE estate (every audience
    # in each sandbox), not the journey->audience view below.
    if opts["audiences"]:
        run_audience_report(service, conf, token, api_key, opts)
        return

    ids = list(opts["ids"])
    want_status = opts["status"]            # set of lowercased statuses, or None
    status_post_filter = bool(want_status)  # cleared if we filter pre-GET on list data
    if opts["list"]:
        listed, by_id = list_journeys(token, api_key, conf, opts["sandbox"])
        if want_status:
            # The list payload carries status, so filter BEFORE the per-journey
            # GETs -- avoids firing ~1,300 GETs when only some statuses are wanted.
            before = len(listed)
            listed = [j for j in listed
                      if (by_id.get(j, {}).get("status") or "").lower() in want_status]
            status_post_filter = False
            print(f"  {ANSI['yellow']}--status {','.join(sorted(want_status))}: "
                  f"{len(listed)} of {before} journeys match (filtered pre-GET)"
                  f"{ANSI['reset']}")
        for jid in listed:
            if jid not in ids:
                ids.append(jid)
        print()

    if not ids:
        logger.info("No journey ids to check. Pass ids, or use --list.")
        return

    if opts["limit"] and len(ids) > opts["limit"]:
        print(f"  {ANSI['yellow']}Sampling first {opts['limit']} of {len(ids)} "
              f"journey(s) (--limit){ANSI['reset']}")
        ids = ids[:opts["limit"]]

    total = len(ids)
    workers = max(1, opts["workers"]) if opts["eval"] else max(1, min(opts["workers"], 24))
    eval_note = "" if opts["eval"] else f" {ANSI['dim']}(--no-eval: no type/tags){ANSI['reset']}"
    print(f"  {ANSI['bold']}Resolving audiences for {total} journey(s){ANSI['reset']} "
          f"{ANSI['dim']}({workers} parallel workers){ANSI['reset']}{eval_note}")
    if status_post_filter:
        print(f"  {ANSI['yellow']}--status {','.join(sorted(want_status))}: no list "
              f"status available (explicit ids), filtering AFTER each GET"
              f"{ANSI['reset']}")
    rows, no_aud, failed, stopped, status_skipped = [], 0, 0, 0, 0
    aud_cache: dict[str, str] = {}   # shared across worker threads; a duplicate
    tag_cache: dict[str, str] = {}   # cache miss just costs one redundant GET

    def _process(jid):
        """Worker: fetch one journey and (with --eval) enrich it. Only touches
        the shared caches; all counters/rows are updated on the main thread."""
        ok, res = get_journey(token, api_key, conf, opts["sandbox"], jid)
        if not ok:
            return {"jid": jid, "ok": False, "detail": res}
        jstatus = res.get("status") or res.get("state") or "?"
        if status_post_filter and jstatus.lower() not in want_status:
            return {"jid": jid, "ok": True, "skip": True}
        auds = extract_audiences(res)
        if opts["eval"]:
            jtags = journey_tags(token, api_key, conf, opts["sandbox"], res, tag_cache)
            for a in auds:
                a["atype"] = audience_type(token, api_key, conf, opts["sandbox"],
                                           a["audience_id"], aud_cache)
        else:
            jtags = ""
            for a in auds:
                a["atype"] = ""
        return {"jid": jid, "ok": True, "name": res.get("name", "?"),
                "status": jstatus, "tags": jtags, "auds": auds}

    done = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_process, jid) for jid in ids]
        for fut in concurrent.futures.as_completed(futures):
            r = fut.result()
            done += 1
            prefix = f"  {ANSI['dim']}[{done:>4}/{total}]{ANSI['reset']} "
            if not r["ok"]:
                failed += 1
                print(prefix + f"{ANSI['red']}FAIL{ANSI['reset']} {shorten(r['jid'], 36)}  {r['detail']}")
                continue
            if r.get("skip"):
                status_skipped += 1
                continue
            jid, jname, jstatus, jtags, auds = (r["jid"], r["name"], r["status"],
                                                r["tags"], r["auds"])
            is_stopped = jstatus.lower() in ("stopped", "closed", "paused")
            if is_stopped:
                stopped += 1
            stat_col = ANSI["red"] if is_stopped else ANSI["dim"]
            head = (prefix + f"{ANSI['yellow']}{jname[:38]:<38}{ANSI['reset']} "
                    f"{stat_col}{jstatus[:9]:<9}{ANSI['reset']} ")
            if jtags:
                head += f"{ANSI['blue']}{jtags[:22]:<22}{ANSI['reset']} "
            if not auds:
                no_aud += 1
                print(head + f"{ANSI['dim']}-> no audience{ANSI['reset']}")
                rows.append((jid, jname, jstatus, jtags, "", "", "", ""))
            else:
                names = ", ".join((a["audience_name"] or a["audience_id"][:8])
                                  + (f" [{a['atype']}]" if a.get("atype") else "")
                                  for a in auds)
                print(head + f"{ANSI['dim']}->{ANSI['reset']} {ANSI['cyan']}{names[:46]}{ANSI['reset']}")
                for a in auds:
                    rows.append((jid, jname, jstatus, jtags, a["audience_id"],
                                 a["audience_name"], a["via"], a.get("atype", "")))

    # Parallel completion order is arbitrary; sort by journey then audience name
    # so the table/XLSX read tidily.
    rows.sort(key=lambda r: (str(r[1]).lower(), str(r[5]).lower()))

    # The per-journey lines above already show every result; only redraw the
    # full aligned table for small/sample runs where it's a tidy overview.
    if rows and len(rows) <= 15:
        print()
        print_table(rows)
    print()
    links = sum(1 for r in rows if r[4])          # rows with an actual audience
    stop = (f"{ANSI['red']}{stopped} stopped{ANSI['reset']}{ANSI['green']}"
            if stopped else "0 stopped")
    skip_txt = f", {status_skipped} skipped by --status" if status_skipped else ""
    print(f"  {ANSI['green' if links else 'yellow']}=> {len(ids)} journey(s): "
          f"{links} audience link(s), {no_aud} with no audience, {stop}, "
          f"{failed} failed{skip_txt}.{ANSI['reset']}")

    if opts["xlsx"] or opts["out"]:
        stamp = datetime.now().strftime("%Y-%m-%d")
        default = OUTPUT_DIR / f"journeys_{opts['sandbox']}_{stamp}.xlsx"
        out_path = Path(opts["out"]) if opts["out"] else default
        subtitle = f"AJO journeys - {service} - {opts['sandbox']} - {stamp}"
        try:
            write_xlsx(rows, out_path, subtitle=subtitle)
            logger.info(f"Wrote {len(rows)} row(s) -> {out_path}")
        except ImportError:
            logger.error("openpyxl not installed (pip install openpyxl) - XLSX not written.")
        except PermissionError:
            # The target is open in Excel (locks the file). Fall back to a
            # time-suffixed name rather than losing the run.
            alt = out_path.with_name(f"{out_path.stem}_{datetime.now().strftime('%H%M%S')}{out_path.suffix}")
            try:
                write_xlsx(rows, alt, subtitle=subtitle)
                logger.warning(f"{out_path.name} is locked (open in Excel?) - wrote {alt.name} instead.")
            except Exception as e:
                logger.error(f"XLSX write failed: {type(e).__name__}: {e}")
        except Exception as e:
            logger.error(f"XLSX write failed: {type(e).__name__}: {e}")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def parse_args(argv):
    opts = {"creds": None, "api_key": None, "sandbox": "prod", "list": False,
            "limit": 0, "xlsx": False, "out": None, "eval": True,
            "status": None, "workers": 16, "ids": [],
            "audiences": False, "sandboxes": None, "highlight": ""}
    given_sandbox = False
    i = 0
    while i < len(argv):
        a = argv[i]
        if a in ("--creds", "-c") and i + 1 < len(argv):
            opts["creds"] = argv[i + 1]; i += 2; continue
        if a == "--api-key" and i + 1 < len(argv):
            opts["api_key"] = argv[i + 1]; i += 2; continue
        if a in ("--sandbox", "-s") and i + 1 < len(argv):
            opts["sandbox"] = argv[i + 1]; given_sandbox = True; i += 2; continue
        if a == "--audiences":
            opts["audiences"] = True; i += 1; continue
        if a == "--highlight" and i + 1 < len(argv):
            opts["highlight"] = argv[i + 1]; i += 2; continue
        if a == "--list":
            opts["list"] = True; i += 1; continue
        if a in ("--limit", "-n") and i + 1 < len(argv):
            opts["limit"] = int(argv[i + 1]); i += 2; continue
        if a == "--xlsx":
            opts["xlsx"] = True; i += 1; continue
        if a == "--no-eval":
            opts["eval"] = False; i += 1; continue
        if a == "--out" and i + 1 < len(argv):
            opts["out"] = argv[i + 1]; i += 2; continue
        if a == "--status" and i + 1 < len(argv):
            opts["status"] = {s.strip().lower() for s in argv[i + 1].split(",") if s.strip()} or None
            i += 2; continue
        if a in ("--workers", "-w") and i + 1 < len(argv):
            opts["workers"] = max(1, int(argv[i + 1])); i += 2; continue
        if a.startswith("-"):
            i += 1; continue
        opts["ids"].append(a); i += 1
    # --audiences can read several sandboxes in one run, so --sandbox takes a
    # comma list (or 'all-prod' for every production sandbox). Left as None when
    # not passed, which is the signal to ASK rather than assume the estate --
    # the journey path keeps its own "prod" default via opts["sandbox"].
    opts["sandboxes"] = ([s.strip() for s in str(opts["sandbox"]).split(",")
                          if s.strip()] or ["prod"]) if given_sandbox else None
    return opts


def main():
    print_banner()
    print(aep_creds.source_banner())
    opts = parse_args(sys.argv[1:])
    services = aep_creds.list_services()
    if not services:
        logger.error("No credentials found in the keyring vault or the creds/ "
                     "folder. Add one with credential_validator_v2.py store "
                     "(or migrate), or drop a <service>.json in creds/.")
        return

    if opts["creds"]:
        service = resolve_service(services, opts["creds"])
        if not service:
            logger.warning(f"No credential set named {opts['creds']!r} "
                           f"(known: {', '.join(services)})")
            return
    else:
        service = menu(services)

    if not service:
        logger.info("Nothing chosen. Exiting.")
        return

    # Friendliness for a bare run: AJO needs an api-key SUBSCRIBED to AJO, often
    # a DIFFERENT credential than the token (the hybrid). If none was supplied
    # and we're on a terminal, offer to pick one rather than failing with a 403.
    if not opts["api_key"] and sys.stdin.isatty():
        has_field = bool(aep_creds.peek_creds(service).get("api_key"))
        if not has_field:
            opts["api_key"] = pick_api_key(services, service)

    # No ids and no --list from an interactive pick -> default to listing all.
    # (--audiences is its own report and never wants the journey walk.)
    if not opts["ids"] and not opts["list"] and not opts["audiences"]:
        opts["list"] = True

    run_checker(service, opts, services)
    print()
    logger.info("Done.")


if __name__ == "__main__":
    main()
