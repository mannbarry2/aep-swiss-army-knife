#!/usr/bin/env python3
"""
ajo_journey_checker.py  (AEP Swiss Army Knife)
==============================================
Read Adobe Journey Optimizer (AJO) journeys and pull out the AUDIENCE behind
each one -- whether the journey READS an audience (a read-audience node) or is
triggered by AUDIENCE QUALIFICATION (a profile qualifying for a segment). The
end goal is a journey -> audience table, and from it a per-audience view of
which journeys consume each audience.

Shares the credential bank, picker and console style of credential_validator:
run with no --creds to pick interactively from ./creds/.

------------------------------------------------------------------------------
THE HYBRID CREDENTIAL (important -- this is why it works)
------------------------------------------------------------------------------
An AJO journey GET on the platform.adobe.io/ajo gateway checks TWO things
independently, and they can come from DIFFERENT credentials:

  * the Bearer TOKEN  -> supplies the identity/permission. Its technical
    account must be granted AEP sandbox + AJO journey access.
  * the x-api-key     -> identifies the integration to the AJO PRODUCT. This
    key must be SUBSCRIBED to AJO, or you get 403 "Api Key is invalid".

In this org neither single credential has both halves:
  - 'acme alpha'  (an internal test account): tech account HAS journey permission, but
                   its api-key is NOT subscribed to AJO.
  - 'acme beta'  (a sibling integration): api-key IS subscribed to AJO, but
                   its tech account has no permission.

So the working call mints the token from ALPHA's credential and sends BETA's
AJO-subscribed api-key. Verified: that combo returns the full journey JSON
(200), while either credential on its own fails.

You express the hybrid with two values:
  --creds   : the credential whose TOKEN to mint (needs journey permission).
              Omit it to pick from the credential bank interactively.
  --api-key : the AJO-SUBSCRIBED api-key for the x-api-key header.
The api-key can also live in the creds JSON as "api_key" (so you don't repeat
it). When neither is set, the creds file's own client_id is used (the
non-hybrid path, for a credential that already has both halves).

Once an admin properly enables ONE credential for AJO, drop the hybrid and just
point --creds at it.

------------------------------------------------------------------------------
USAGE
------------------------------------------------------------------------------
  python ajo_journey_checker.py                       # interactive picker, then --list
  python ajo_journey_checker.py --creds "acme alpha" --api-key <ajo-api-key> --list
  python ajo_journey_checker.py --creds "acme alpha" --api-key <key> <journeyId> [...]
  python ajo_journey_checker.py --creds "acme alpha" --api-key <key> --sandbox dev <journeyId>

--list pages the ENTIRE journey estate (~1,300), not just the recently-modified
default. The per-journey GETs (and the Streaming/Batch + tag enrichment) run in
parallel -- --workers N (default 16) -- so the full estate takes minutes, not an
hour. --status live,published,paused filters to those statuses BEFORE the
per-journey GETs (the list payload carries status), so you don't fire a GET per
journey when you only want some. --limit N samples the first N. --no-eval drops
the Streaming/Batch + tag columns for a quick id/name/status-only run.

Read-only: every call is a GET. Nothing is written to AEP/AJO. Stdlib only.
"""
from __future__ import annotations

import concurrent.futures
import json
import logging
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime
from pathlib import Path

SCRIPT_NAME    = "ajo_journey_checker"
SCRIPT_VERSION = "1.0.0"
SCRIPT_DATE    = "2026-06-24"
SCRIPT_AUTHOR  = "Barry Mann (barrymann.com)"

SCRIPT_DIR = Path(__file__).resolve().parent
CREDS_DIR = SCRIPT_DIR / "creds"
OUTPUT_DIR = SCRIPT_DIR / "output"

IMS_URL = "https://ims-na1.adobelogin.com/ims/token"
AJO_GATEWAY = "https://platform.adobe.io/ajo"
DEFAULT_SCOPES = (
    "openid,AdobeID,read_organizations,"
    "additional_info.projectedProductContext,session"
)
# The journey LIST endpoint is the SINGULAR /ajo/journey (the plural
# /ajo/journeys is a 404 red herring). It returns {filter, pagination, results}
# where pagination = {page, pageSize, totalCount}.
#
# CRITICAL: with no filter the gateway applies a DEFAULT one
# (metadata.lastModifiedAt > ~7 days ago), silently returning only recently-
# modified journeys (~60). We pass an explicit wide date filter to get the FULL
# estate, then page through with page=0,1,2,... pageSize is a fixed 100 (server
# ignores size/limit params), so a ~1,300-journey estate is ~13 pages.
JOURNEY_LIST_URL = f"{AJO_GATEWAY}/journey"
JOURNEY_LIST_FILTER = "metadata.lastModifiedAt>2000-01-01"  # widen past the default
MAX_JOURNEY_PAGES = 500  # safety cap (~50k journeys) against an infinite paging loop

# ----------------------------------------------------------------------------
# ANSI / logging - matches credential_validator.py style
# ----------------------------------------------------------------------------
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        h = kernel32.GetStdHandle(-11)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
            kernel32.SetConsoleMode(h, mode.value | 0x0004)
    except Exception:
        pass

ANSI = {
    "reset": "\033[0m", "bold": "\033[1m", "dim": "\033[2m",
    "red": "\033[31m", "green": "\033[32m", "yellow": "\033[33m",
    "blue": "\033[34m", "magenta": "\033[35m", "cyan": "\033[36m",
}
LEVEL_COLOR = {
    "DEBUG": ANSI["dim"], "INFO": ANSI["green"],
    "WARNING": ANSI["yellow"],
    "ERROR": ANSI["red"] + ANSI["bold"],
    "CRITICAL": ANSI["red"] + ANSI["bold"],
}


class ColoredFormatter(logging.Formatter):
    def format(self, record):
        color = LEVEL_COLOR.get(record.levelname, "")
        ts = self.formatTime(record, "%H:%M:%S")
        return (
            f"{ANSI['dim']}{ts}{ANSI['reset']} "
            f"{color}[{record.levelname:<7}]{ANSI['reset']} "
            f"{record.getMessage()}"
        )


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(ColoredFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger("ajo_journey_checker")
SSL_CTX = ssl._create_unverified_context()


def print_banner() -> None:
    bar = ANSI["cyan"] + "=" * 72 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}{SCRIPT_NAME} v{SCRIPT_VERSION}{ANSI['reset']}   ({SCRIPT_DATE})")
    print(f"  by {SCRIPT_AUTHOR}")
    print(f"  {ANSI['dim']}Map Adobe Journey Optimizer journeys to the audiences "
          f"behind them.{ANSI['reset']}")
    print(bar)


# ----------------------------------------------------------------------------
# HTTP / IMS / creds helpers
# ----------------------------------------------------------------------------
def http(url, method="GET", headers=None, data=None, timeout=30):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read(), dict(r.headers)


def clean_detail(raw, limit=160):
    """One-line, CR-stripped slice of a (possibly HTML) error body, so gateway
    error pages can't reset the terminal cursor and garble a line."""
    text = raw.decode(errors="replace") if isinstance(raw, bytes) else raw
    return " ".join(text.split())[:limit]


def shorten(s, n=12):
    if not s:
        return "?"
    return s if len(s) <= n else f"{s[:n]}..."


def load_creds(path: Path) -> dict:
    raw = json.loads(path.read_text(encoding="utf-8"))
    conf = {k: (v.strip() if isinstance(v, str) else v)
            for k, v in raw.items() if not k.startswith("_")}
    for key in ("client_id", "client_secret", "org_id"):
        if not conf.get(key):
            raise ValueError(f"Missing required key {key!r} in {path.name}")
    return conf


def authenticate(conf) -> str:
    """OAuth server-to-server. ALWAYS minted fresh -- AEP/AJO permissions are
    snapshotted into the token at mint time, so a stale token never reflects a
    just-granted product profile. Returns the access token string."""
    payload = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": conf["client_id"],
        "client_secret": conf["client_secret"],
        "scope": conf.get("scopes") or DEFAULT_SCOPES,
    }).encode("utf-8")
    body, _ = http(IMS_URL, method="POST",
                   headers={"Content-Type": "application/x-www-form-urlencoded"},
                   data=payload)
    return json.loads(body)["access_token"]


def ajo_headers(token, api_key, org_id, sandbox):
    return {
        "Authorization": f"Bearer {token}",
        "x-api-key": api_key,
        "x-gw-ims-org-id": org_id,
        "x-sandbox-name": sandbox,
        "Accept": "application/json",
    }


# ----------------------------------------------------------------------------
# Discovery / menu - matches credential_validator.py
# ----------------------------------------------------------------------------
def discover_creds():
    """Return ordered list of credential JSON paths."""
    paths = []
    if CREDS_DIR.exists():
        for p in sorted(CREDS_DIR.glob("*.json")):
            if p.stem == "example":
                continue
            paths.append(p)
    return paths


def menu(creds):
    """Single-pick credential bank (the token credential)."""
    print()
    bar = ANSI["cyan"] + "=" * 70 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Credential bank{ANSI['reset']}  "
          f"{ANSI['dim']}({CREDS_DIR}){ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 70 + ANSI["reset"])
    for i, p in enumerate(creds, 1):
        print(f"  {ANSI['bold']}{i:>2}{ANSI['reset']}  "
              f"{ANSI['yellow']}{p.stem:<20}{ANSI['reset']} "
              f"{ANSI['dim']}{p.name}{ANSI['reset']}")
    print(bar)
    raw = input(
        f"\nPick the token credential by number "
        f"({ANSI['cyan']}1{ANSI['reset']}), blank to quit: "
    ).strip()
    if not raw:
        return None
    tok = raw.replace(",", " ").split()[0]
    if tok.isdigit() and 1 <= int(tok) <= len(creds):
        return creds[int(tok) - 1]
    logger.warning(f"Invalid choice: {tok}")
    return None


def resolve_by_stem(creds, name):
    by_stem = {p.stem: p for p in creds}
    if name in by_stem:
        return by_stem[name]
    term = name.lower()
    hits = [p for p in creds if term in p.stem.lower()]
    return hits[0] if len(hits) == 1 else None


def api_key_source(api_key, creds, org_id=None):
    """Which credential file supplies this x-api-key? Matches the key against
    each set's client_id / api_key so we can name it in the header instead of
    showing a bare hex prefix. The same client_id can appear in more than one
    org, so a match in the SAME org as the token wins (avoids naming a
    different-org set with a shared key)."""
    fallback = None
    for p in creds:
        try:
            c = load_creds(p)
        except Exception:
            continue
        if api_key in (c.get("client_id"), c.get("api_key")):
            if org_id and c.get("org_id") == org_id:
                return p.stem
            fallback = fallback or p.stem
    return fallback


def pick_api_key(creds, token_path):
    """Interactively choose the AJO-subscribed x-api-key (the hybrid). AJO
    checks the api-key is subscribed to the product, and that key is often a
    DIFFERENT credential than the one whose token you mint. Returns an api-key
    string, or None to fall back to the token credential's own key."""
    others = [p for p in creds if p != token_path]
    print()
    print(f"  {ANSI['bold']}AJO needs an AJO-subscribed api-key{ANSI['reset']} "
          f"{ANSI['dim']}(the x-api-key; can differ from the token credential)"
          f"{ANSI['reset']}")
    for i, p in enumerate(others, 1):
        print(f"  {ANSI['bold']}{i:>2}{ANSI['reset']}  "
              f"{ANSI['yellow']}{p.stem:<20}{ANSI['reset']} "
              f"{ANSI['dim']}use this set's api-key{ANSI['reset']}")
    raw = input(
        f"\nPick the AJO-subscribed set by number, paste a key, or Enter to use "
        f"{ANSI['yellow']}{token_path.stem}{ANSI['reset']}'s own key: "
    ).strip()
    if not raw:
        return None
    if raw.isdigit() and 1 <= int(raw) <= len(others):
        try:
            c = load_creds(others[int(raw) - 1])
            return c.get("api_key") or c["client_id"]
        except Exception as e:
            logger.warning(f"Could not load that set: {e}")
            return None
    return raw  # treat anything else as a pasted api-key


# ----------------------------------------------------------------------------
# Journey fetch + audience extraction
# ----------------------------------------------------------------------------
AUDIENCES_URL = "https://platform.adobe.io/data/core/ups/audiences"
# Tag names live in the Unified Tags service on a DIFFERENT host
# (experience.adobe.io); a journey only carries tag ids with name=null.
UNIFIED_TAGS_URL = "https://experience.adobe.io/unifiedtags/tags"


def get_journey(token, api_key, conf, sandbox, journey_id):
    """GET one journey by id. Returns (ok, journey_dict_or_error_string)."""
    url = f"{AJO_GATEWAY}/journey/{journey_id}"
    try:
        body, _ = http(url, headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=30)
        return True, json.loads(body)
    except urllib.error.HTTPError as e:
        return False, f"HTTP {e.code}: {clean_detail(e.read())}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def audience_type(token, api_key, conf, sandbox, audience_id, cache):
    """Streaming / Batch / Edge for one audience, from AEP's audiences API
    (evaluationInfo). Cached -- audiences repeat across journeys. Returns a
    short label; failures degrade to a marker rather than raising."""
    if audience_id in cache:
        return cache[audience_id]
    url = f"{AUDIENCES_URL}/{audience_id}"
    label = "?"
    try:
        body, _ = http(url, headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=20)
        ei = (json.loads(body) or {}).get("evaluationInfo") or {}
        if ei.get("continuous", {}).get("enabled"):
            label = "Streaming"
        elif ei.get("synchronous", {}).get("enabled"):
            label = "Edge"
        elif ei.get("batch", {}).get("enabled"):
            label = "Batch"
        else:
            label = "unknown"
    except urllib.error.HTTPError as e:
        label = "not-found" if e.code == 404 else ("no-access" if e.code == 403 else f"err{e.code}")
    except Exception:
        label = "err"
    cache[audience_id] = label
    return label


def tag_name(token, api_key, conf, sandbox, tag_id, cache):
    """Resolve a journey tag id to its name via the Unified Tags service.
    Cached -- tags repeat across journeys. Falls back to the short id."""
    if tag_id in cache:
        return cache[tag_id]
    name = tag_id[:8]
    try:
        body, _ = http(f"{UNIFIED_TAGS_URL}/{tag_id}",
                       headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=20)
        name = (json.loads(body) or {}).get("name") or tag_id[:8]
    except Exception:
        name = tag_id[:8]
    cache[tag_id] = name
    return name


def journey_tags(token, api_key, conf, sandbox, journey, cache):
    """Comma-joined tag NAMES for a journey (its tags[] carries ids only)."""
    ids = [t.get("id") for t in (journey.get("tags") or [])
           if isinstance(t, dict) and t.get("id")]
    return ", ".join(tag_name(token, api_key, conf, sandbox, t, cache) for t in ids)


def extract_audiences(journey: dict) -> list[dict]:
    """Pull every audience/segment reference out of a journey definition,
    regardless of how it's used. Returns a de-duped list of
    {audience_id, audience_name, via} dicts.

    Two shapes are handled, plus a recursive catch-all so we don't miss a
    variant we haven't seen yet:
      * read-audience nodes: a node carrying an "audiences":[{id,name}] list.
      * audience qualification: an entry/event referencing a segment, seen as
        segmentId / audienceId / segment.{id,name} keys.
    """
    found: dict[str, dict] = {}

    def add(aid, name, via):
        if not aid:
            return
        cur = found.get(aid)
        if cur is None:
            found[aid] = {"audience_id": aid, "audience_name": name or "", "via": via}
        else:
            if name and not cur["audience_name"]:
                cur["audience_name"] = name
            if via not in cur["via"]:
                cur["via"] = f"{cur['via']}+{via}"

    def walk(obj, node_type=None):
        if isinstance(obj, dict):
            nt = obj.get("type") or node_type
            auds = obj.get("audiences")
            if isinstance(auds, list):
                for a in auds:
                    if isinstance(a, dict):
                        # Label by the actual node type: a read-audience node vs
                        # an 'audience_qualification' entry/trigger node (confirmed
                        # node type); fall back to a generic tag otherwise.
                        ntl = str(nt).lower() if nt else ""
                        via = ("read_audience" if "read" in ntl
                               else "qualification" if "qualif" in ntl
                               else "audiences")
                        add(a.get("id") or a.get("audienceId"), a.get("name"), via)
            for key in ("segmentId", "audienceId"):
                if obj.get(key):
                    add(obj[key], obj.get("name"), "qualification")
            seg = obj.get("segment") or obj.get("audience")
            if isinstance(seg, dict) and (seg.get("id") or seg.get("segmentId")):
                add(seg.get("id") or seg.get("segmentId"), seg.get("name"), "qualification")
            for v in obj.values():
                walk(v, nt)
        elif isinstance(obj, list):
            for v in obj:
                walk(v, node_type)

    walk(journey)
    return list(found.values())


def list_journeys(token, api_key, conf, sandbox):
    """Page through the ENTIRE AJO journey list. Returns (ids, by_id) where
    by_id maps journey id -> {"name", "status"} taken from the list items, so a
    --status filter can run BEFORE the per-journey GET loop.

    Follows the gateway's page-based pagination (pagination.{page,pageSize,
    totalCount}) with an explicit wide date filter, until a page comes back
    empty or we've collected totalCount (whichever first), capped at
    MAX_JOURNEY_PAGES."""
    print(f"  {ANSI['bold']}Listing journeys{ANSI['reset']} "
          f"{ANSI['dim']}(sandbox '{sandbox}', paging the full estate){ANSI['reset']}")
    by_id: dict[str, dict] = {}
    total = None
    page = 0
    while page < MAX_JOURNEY_PAGES:
        params = {"filter": JOURNEY_LIST_FILTER, "page": page}
        url = f"{JOURNEY_LIST_URL}?{urllib.parse.urlencode(params)}"
        try:
            body, _ = http(url, headers=ajo_headers(token, api_key, conf["org_id"], sandbox), timeout=60)
        except urllib.error.HTTPError as e:
            detail = clean_detail(e.read(), 90)
            print(f"     {ANSI['yellow']}[{e.code}] page {page} -> {detail}{ANSI['reset']}")
            if e.code == 403 and ("api key is invalid" in detail.lower() or "403003" in detail):
                print(f"     {ANSI['dim']}      ^ that x-api-key isn't subscribed to AJO "
                      f"(the hybrid) -- pass an AJO-subscribed --api-key.{ANSI['reset']}")
            break
        except Exception as e:
            print(f"     {ANSI['red']}[ERR] page {page} -> {type(e).__name__}: {e}{ANSI['reset']}")
            break
        try:
            data = json.loads(body)
        except Exception:
            print(f"     {ANSI['yellow']}[200] page {page}: non-JSON body{ANSI['reset']}")
            break
        results = data.get("results") if isinstance(data, dict) else (data if isinstance(data, list) else None)
        if not results:
            break
        for it in results:
            if isinstance(it, dict) and it.get("id"):
                by_id[it["id"]] = {"name": it.get("name", "?"), "status": it.get("status") or "?"}
        pg = (data.get("pagination") or {}) if isinstance(data, dict) else {}
        if pg.get("totalCount") is not None:
            total = pg["totalCount"]
        page += 1
        # progress for a long multi-page walk
        print(f"     {ANSI['dim']}page {page}: {len(by_id)}"
              f"{('/' + str(total)) if total is not None else ''} journeys...{ANSI['reset']}")
        if total is not None and len(by_id) >= total:
            break
    else:
        print(f"     {ANSI['yellow']}(stopped at page cap {MAX_JOURNEY_PAGES}){ANSI['reset']}")
    shown_total = total if total is not None else len(by_id)
    print(f"     {ANSI['green']}[OK] {len(by_id)} journeys across {page} page(s) "
          f"(totalCount {shown_total}){ANSI['reset']}")
    return list(by_id.keys()), by_id


# ----------------------------------------------------------------------------
# Output
# ----------------------------------------------------------------------------
def print_table(rows):
    """rows: list of (journey_id, journey_name, audience_id, audience_name, via, type)."""
    if not rows:
        print(f"  {ANSI['dim']}(no audiences found){ANSI['reset']}")
        return
    headers = ("Journey id", "Journey name", "Status", "Tags",
               "Audience id", "Audience name", "Via", "Type")
    caps = (32, 30, 9, 20, 32, 28, 13, 9)
    colors = ("dim", "yellow", "yellow", "blue", "dim", "cyan", "magenta", "green")
    n = len(headers)
    widths = [min(caps[i], max(len(str(r[i])) for r in [headers] + rows)) for i in range(n)]

    def cell(text, w, color=None):
        s = str(text)[:w].ljust(w)
        return f"{ANSI[color]}{s}{ANSI['reset']}" if color else s

    print("  " + " | ".join(f"{ANSI['bold']}{str(h)[:widths[i]].ljust(widths[i])}{ANSI['reset']}"
                            for i, h in enumerate(headers)))
    print("  " + ANSI["cyan"] + "-+-".join("-" * w for w in widths) + ANSI["reset"])
    for r in rows:
        print("  " + " | ".join(cell(c, widths[i], colors[i]) for i, c in enumerate(r)))


def write_xlsx(rows, out_path, subtitle=""):
    """Write the journey -> audience rows to a single-sheet XLSX (house style:
    bold banded header, frozen top row, autofilter, sized columns)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Journeys"
    headers = ["Journey id", "Journey name", "Status", "Tags",
               "Audience id", "Audience name", "Via", "Audience type"]
    ws.append(headers)
    for r in rows:
        ws.append([str(c) for c in r])

    hdr_fill = PatternFill("solid", fgColor="1F4E78")
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = hdr_fill
        c.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    for i, w in enumerate((38, 46, 12, 26, 38, 42, 16, 16), 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    if subtitle:
        ws.oddHeader.left.text = subtitle

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def run_checker(path, opts, creds):
    bar = ANSI["cyan"] + "=" * 70 + ANSI["reset"]
    print()
    print(bar)
    print(f"  {ANSI['bold']}AJO Journey Checker{ANSI['reset']}  "
          f"{ANSI['yellow']}{path.stem}{ANSI['reset']} "
          f"{ANSI['dim']}({path.name}){ANSI['reset']}")
    print(bar)

    try:
        conf = load_creds(path)
    except Exception as e:
        logger.error(f"Failed to load {path.name}: {e}")
        return

    api_key = opts["api_key"] or conf.get("api_key") or conf["client_id"]
    hybrid = api_key != conf["client_id"]
    key_from = api_key_source(api_key, creds, conf["org_id"]) or ("pasted key" if hybrid else path.stem)
    # Spell out the two credentials by NAME so the hybrid isn't confusing.
    print(f"  {ANSI['bold']}Token   from:{ANSI['reset']} {ANSI['yellow']}{path.stem}{ANSI['reset']} "
          f"{ANSI['dim']}(client {shorten(conf['client_id'])} - gives the journey permission){ANSI['reset']}")
    print(f"  {ANSI['bold']}Api-key from:{ANSI['reset']} {ANSI['yellow']}{key_from}{ANSI['reset']} "
          f"{ANSI['dim']}(key {shorten(api_key)} - must be subscribed to AJO){ANSI['reset']}"
          + (f"  {ANSI['cyan']}[hybrid]{ANSI['reset']}" if hybrid else ""))
    print(f"  {ANSI['bold']}Org / sandbox:{ANSI['reset']} {ANSI['magenta']}{conf['org_id']}{ANSI['reset']}"
          f"  {ANSI['dim']}/{ANSI['reset']} {opts['sandbox']}")
    print()

    try:
        token = authenticate(conf)
    except urllib.error.HTTPError as e:
        logger.error(f"IMS auth FAILED: HTTP {e.code} {clean_detail(e.read())}")
        return
    except Exception as e:
        logger.error(f"IMS auth FAILED: {type(e).__name__}: {e}")
        return
    print(f"  {ANSI['green']}[OK] token minted fresh{ANSI['reset']}")
    print()

    ids = list(opts["ids"])
    want_status = opts["status"]            # set of lowercased statuses, or None
    status_post_filter = bool(want_status)  # cleared if we filter pre-GET on list data
    if opts["list"]:
        listed, by_id = list_journeys(token, api_key, conf, opts["sandbox"])
        if want_status:
            # The list payload carries status, so filter BEFORE the per-journey
            # GETs -- avoids firing ~1,300 GETs when only some statuses are wanted.
            before = len(listed)
            listed = [j for j in listed
                      if (by_id.get(j, {}).get("status") or "").lower() in want_status]
            status_post_filter = False
            print(f"  {ANSI['yellow']}--status {','.join(sorted(want_status))}: "
                  f"{len(listed)} of {before} journeys match (filtered pre-GET)"
                  f"{ANSI['reset']}")
        for jid in listed:
            if jid not in ids:
                ids.append(jid)
        print()

    if not ids:
        logger.info("No journey ids to check. Pass ids, or use --list.")
        return

    if opts["limit"] and len(ids) > opts["limit"]:
        print(f"  {ANSI['yellow']}Sampling first {opts['limit']} of {len(ids)} "
              f"journey(s) (--limit){ANSI['reset']}")
        ids = ids[:opts["limit"]]

    total = len(ids)
    workers = max(1, opts["workers"]) if opts["eval"] else max(1, min(opts["workers"], 24))
    eval_note = "" if opts["eval"] else f" {ANSI['dim']}(--no-eval: no type/tags){ANSI['reset']}"
    print(f"  {ANSI['bold']}Resolving audiences for {total} journey(s){ANSI['reset']} "
          f"{ANSI['dim']}({workers} parallel workers){ANSI['reset']}{eval_note}")
    if status_post_filter:
        print(f"  {ANSI['yellow']}--status {','.join(sorted(want_status))}: no list "
              f"status available (explicit ids), filtering AFTER each GET"
              f"{ANSI['reset']}")
    rows, no_aud, failed, stopped, status_skipped = [], 0, 0, 0, 0
    aud_cache: dict[str, str] = {}   # shared across worker threads; a duplicate
    tag_cache: dict[str, str] = {}   # cache miss just costs one redundant GET

    def _process(jid):
        """Worker: fetch one journey and (with --eval) enrich it. Only touches
        the shared caches; all counters/rows are updated on the main thread."""
        ok, res = get_journey(token, api_key, conf, opts["sandbox"], jid)
        if not ok:
            return {"jid": jid, "ok": False, "detail": res}
        jstatus = res.get("status") or res.get("state") or "?"
        if status_post_filter and jstatus.lower() not in want_status:
            return {"jid": jid, "ok": True, "skip": True}
        auds = extract_audiences(res)
        if opts["eval"]:
            jtags = journey_tags(token, api_key, conf, opts["sandbox"], res, tag_cache)
            for a in auds:
                a["atype"] = audience_type(token, api_key, conf, opts["sandbox"],
                                           a["audience_id"], aud_cache)
        else:
            jtags = ""
            for a in auds:
                a["atype"] = ""
        return {"jid": jid, "ok": True, "name": res.get("name", "?"),
                "status": jstatus, "tags": jtags, "auds": auds}

    done = 0
    with concurrent.futures.ThreadPoolExecutor(max_workers=workers) as ex:
        futures = [ex.submit(_process, jid) for jid in ids]
        for fut in concurrent.futures.as_completed(futures):
            r = fut.result()
            done += 1
            prefix = f"  {ANSI['dim']}[{done:>4}/{total}]{ANSI['reset']} "
            if not r["ok"]:
                failed += 1
                print(prefix + f"{ANSI['red']}FAIL{ANSI['reset']} {shorten(r['jid'], 36)}  {r['detail']}")
                continue
            if r.get("skip"):
                status_skipped += 1
                continue
            jid, jname, jstatus, jtags, auds = (r["jid"], r["name"], r["status"],
                                                r["tags"], r["auds"])
            is_stopped = jstatus.lower() in ("stopped", "closed", "paused")
            if is_stopped:
                stopped += 1
            stat_col = ANSI["red"] if is_stopped else ANSI["dim"]
            head = (prefix + f"{ANSI['yellow']}{jname[:38]:<38}{ANSI['reset']} "
                    f"{stat_col}{jstatus[:9]:<9}{ANSI['reset']} ")
            if jtags:
                head += f"{ANSI['blue']}{jtags[:22]:<22}{ANSI['reset']} "
            if not auds:
                no_aud += 1
                print(head + f"{ANSI['dim']}-> no audience{ANSI['reset']}")
                rows.append((jid, jname, jstatus, jtags, "", "", "", ""))
            else:
                names = ", ".join((a["audience_name"] or a["audience_id"][:8])
                                  + (f" [{a['atype']}]" if a.get("atype") else "")
                                  for a in auds)
                print(head + f"{ANSI['dim']}->{ANSI['reset']} {ANSI['cyan']}{names[:46]}{ANSI['reset']}")
                for a in auds:
                    rows.append((jid, jname, jstatus, jtags, a["audience_id"],
                                 a["audience_name"], a["via"], a.get("atype", "")))

    # Parallel completion order is arbitrary; sort by journey then audience name
    # so the table/XLSX read tidily.
    rows.sort(key=lambda r: (str(r[1]).lower(), str(r[5]).lower()))

    # The per-journey lines above already show every result; only redraw the
    # full aligned table for small/sample runs where it's a tidy overview.
    if rows and len(rows) <= 15:
        print()
        print_table(rows)
    print()
    links = sum(1 for r in rows if r[4])          # rows with an actual audience
    stop = (f"{ANSI['red']}{stopped} stopped{ANSI['reset']}{ANSI['green']}"
            if stopped else "0 stopped")
    skip_txt = f", {status_skipped} skipped by --status" if status_skipped else ""
    print(f"  {ANSI['green' if links else 'yellow']}=> {len(ids)} journey(s): "
          f"{links} audience link(s), {no_aud} with no audience, {stop}, "
          f"{failed} failed{skip_txt}.{ANSI['reset']}")

    if opts["xlsx"] or opts["out"]:
        stamp = datetime.now().strftime("%Y-%m-%d")
        default = OUTPUT_DIR / (f"ajo_journeys_{path.stem.replace(' ', '_')}_"
                                f"{opts['sandbox']}_{stamp}.xlsx")
        out_path = Path(opts["out"]) if opts["out"] else default
        subtitle = f"AJO journeys - {path.stem} - {opts['sandbox']} - {stamp}"
        try:
            write_xlsx(rows, out_path, subtitle=subtitle)
            logger.info(f"Wrote {len(rows)} row(s) -> {out_path}")
        except ImportError:
            logger.error("openpyxl not installed (pip install openpyxl) - XLSX not written.")
        except PermissionError:
            # The target is open in Excel (locks the file). Fall back to a
            # time-suffixed name rather than losing the run.
            alt = out_path.with_name(f"{out_path.stem}_{datetime.now().strftime('%H%M%S')}{out_path.suffix}")
            try:
                write_xlsx(rows, alt, subtitle=subtitle)
                logger.warning(f"{out_path.name} is locked (open in Excel?) - wrote {alt.name} instead.")
            except Exception as e:
                logger.error(f"XLSX write failed: {type(e).__name__}: {e}")
        except Exception as e:
            logger.error(f"XLSX write failed: {type(e).__name__}: {e}")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def parse_args(argv):
    opts = {"creds": None, "api_key": None, "sandbox": "prod", "list": False,
            "limit": 0, "xlsx": False, "out": None, "eval": True,
            "status": None, "workers": 16, "ids": []}
    i = 0
    while i < len(argv):
        a = argv[i]
        if a in ("--creds", "-c") and i + 1 < len(argv):
            opts["creds"] = argv[i + 1]; i += 2; continue
        if a == "--api-key" and i + 1 < len(argv):
            opts["api_key"] = argv[i + 1]; i += 2; continue
        if a in ("--sandbox", "-s") and i + 1 < len(argv):
            opts["sandbox"] = argv[i + 1]; i += 2; continue
        if a == "--list":
            opts["list"] = True; i += 1; continue
        if a in ("--limit", "-n") and i + 1 < len(argv):
            opts["limit"] = int(argv[i + 1]); i += 2; continue
        if a == "--xlsx":
            opts["xlsx"] = True; i += 1; continue
        if a == "--no-eval":
            opts["eval"] = False; i += 1; continue
        if a == "--out" and i + 1 < len(argv):
            opts["out"] = argv[i + 1]; i += 2; continue
        if a == "--status" and i + 1 < len(argv):
            opts["status"] = {s.strip().lower() for s in argv[i + 1].split(",") if s.strip()} or None
            i += 2; continue
        if a in ("--workers", "-w") and i + 1 < len(argv):
            opts["workers"] = max(1, int(argv[i + 1])); i += 2; continue
        if a.startswith("-"):
            i += 1; continue
        opts["ids"].append(a); i += 1
    return opts


def main():
    print_banner()
    opts = parse_args(sys.argv[1:])
    creds = discover_creds()
    if not creds:
        logger.error(f"No credential JSONs found in {CREDS_DIR}. "
                     f"Drop your <tenant>.json files there.")
        return

    if opts["creds"]:
        path = resolve_by_stem(creds, opts["creds"])
        if not path:
            logger.warning(f"No credential set named {opts['creds']!r} (looked in {CREDS_DIR})")
            return
    else:
        path = menu(creds)

    if not path:
        logger.info("Nothing chosen. Exiting.")
        return

    # Friendliness for a bare run: AJO needs an api-key SUBSCRIBED to AJO, often
    # a DIFFERENT credential than the token (the hybrid). If none was supplied
    # and we're on a terminal, offer to pick one rather than failing with a 403.
    if not opts["api_key"] and sys.stdin.isatty():
        try:
            has_field = bool(load_creds(path).get("api_key"))
        except Exception:
            has_field = False
        if not has_field:
            opts["api_key"] = pick_api_key(creds, path)

    # No ids and no --list from an interactive pick -> default to listing all.
    if not opts["ids"] and not opts["list"]:
        opts["list"] = True

    run_checker(path, opts, creds)
    print()
    logger.info("Done.")


if __name__ == "__main__":
    main()
