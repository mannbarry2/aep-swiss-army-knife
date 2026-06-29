#!/usr/bin/env python3
"""
audit_batch_schedules_v2.py
===========================
Enhanced fork of audit_batch_schedules.py. Same credential picker, same
per-sandbox GET /schedules sweep, but every schedule is now *classified* and
*flagged for anomalies*, with a roll-up summary at the end.

Pick a credential set from ./creds/ and the auditor will:
  1. Authenticate against IMS (client_credentials).
  2. GET /sandboxes -- every sandbox the credential can see.
  3. Prompt you to pick which sandbox(es) to audit (one, several, or all)
     before any schedule calls are made.
  4. Per chosen sandbox, GET /schedules (x-sandbox-name header).
  5. Classify each schedule and flag anomalies, print a table + summary, and
     write outputs: a flat CSV and a tabbed XLSX workbook -- one worksheet per
     sandbox plus a Summary tab.

The tabbed-workbook layout (one sandbox per tab + a Summary tab) is the
intended house style for export-producing tools in this repo going forward.

Classification (TYPE) -- precedence top to bottom:
  QUERY        the schedule id carries a human-readable query name
               (e.g. ...acme_order_merge_query...). Named queries win even
               when they run on a cron, because the name is the useful signal.
  CRON         id is UUID-only AND the time is a recurring cron expression
               with no single daily run time (e.g. 0 */2 * * *, 0 * * * *).
  SEGMENTATION id is UUID-only AND the schedule runs at a fixed clock time.

Anomaly FLAGS (comma-separated when several apply):
  ODD_TIME   fixed time is not on a round hour or half hour (e.g. 17:49, 05:41)
  ONCE       the schedule is a one-shot (@once)
  DISABLED   enabled = false
  LATE       SEGMENTATION whose fixed time is >= 05:00 UTC on a prod sandbox

Stdlib only, VDI-friendly. No pip install required.

Usage:
    python audit_batch_schedules_v2.py            # interactive credential menu
    python audit_batch_schedules_v2.py prod       # pick a set by filename stem
"""

from __future__ import annotations

import csv
import json
import logging
import re
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------
SCRIPT_NAME = "audit_batch_schedules_v2"
SCRIPT_VERSION = "2.0.0"
SCRIPT_DATE = "2026-06-24"
SCRIPT_AUTHOR = "Barry Mann (barrymann.com)"

SCRIPT_DIR = Path(__file__).resolve().parent
CREDS_DIR = SCRIPT_DIR / "creds"
OUTPUT_DIR = SCRIPT_DIR / "output"

IMS_URL = "https://ims-na1.adobelogin.com/ims/token"
SANDBOX_LIST_URL = (
    "https://platform.adobe.io/data/foundation/sandbox-management/sandboxes"
)
SCHEDULES_URL = "https://platform.adobe.io/data/foundation/query/schedules"

DEFAULT_SCOPES = (
    "openid,AdobeID,read_organizations,"
    "additional_info.projectedProductContext,session"
)

# ----------------------------------------------------------------------------
# ANSI / logging - matches credential_validator.py style
# ----------------------------------------------------------------------------
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        h = kernel32.GetStdHandle(-11)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
            kernel32.SetConsoleMode(h, mode.value | 0x0004)
    except Exception:
        pass

ANSI = {
    "reset": "\033[0m", "bold": "\033[1m", "dim": "\033[2m",
    "red": "\033[31m", "green": "\033[32m", "yellow": "\033[33m",
    "blue": "\033[34m", "magenta": "\033[35m", "cyan": "\033[36m",
}
LEVEL_COLOR = {
    "DEBUG": ANSI["dim"], "INFO": ANSI["green"],
    "WARNING": ANSI["yellow"],
    "ERROR": ANSI["red"] + ANSI["bold"],
    "CRITICAL": ANSI["red"] + ANSI["bold"],
}


class ColoredFormatter(logging.Formatter):
    def format(self, record):
        color = LEVEL_COLOR.get(record.levelname, "")
        ts = self.formatTime(record, "%H:%M:%S")
        return (
            f"{ANSI['dim']}{ts}{ANSI['reset']} "
            f"{color}[{record.levelname:<7}]{ANSI['reset']} "
            f"{record.getMessage()}"
        )


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(ColoredFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger("audit_batch_schedules_v2")
SSL_CTX = ssl._create_unverified_context()


# ----------------------------------------------------------------------------
# HTTP / IMS / credential helpers  (unchanged from v1)
# ----------------------------------------------------------------------------
def http(url, method="GET", headers=None, data=None, timeout=30):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read(), dict(r.headers)


def flatten(text: str, limit: int = 180) -> str:
    """Collapse a multi-line error body (e.g. an HTML 404 page) to a single
    truncated line so it never breaks the summary table's alignment."""
    return " ".join((text or "").split())[:limit]


def load_creds(path: Path):
    raw = json.loads(path.read_text(encoding="utf-8"))
    conf = {
        k: v.strip() if isinstance(v, str) else v
        for k, v in raw.items()
        if not k.startswith("_")
    }
    for key in ("client_id", "client_secret", "org_id"):
        if not conf.get(key):
            raise ValueError(f"Missing required key {key!r} in {path.name}")
    return conf


def authenticate(conf):
    payload = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": conf["client_id"],
        "client_secret": conf["client_secret"],
        "scope": conf.get("scopes") or DEFAULT_SCOPES,
    }).encode("utf-8")
    body, _ = http(
        conf.get("oauth_url") or IMS_URL,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data=payload,
    )
    return json.loads(body)


def aep_headers(token, conf, sandbox=None):
    headers = {
        "Authorization": f"Bearer {token}",
        "x-api-key": conf.get("api_key") or conf["client_id"],
        "x-gw-ims-org-id": conf["org_id"],
        "Accept": "application/json",
    }
    if sandbox:
        headers["x-sandbox-name"] = sandbox
    return headers


def list_sandboxes(token, conf):
    """GET /sandboxes. Returns (ok, sandboxes_or_error_string)."""
    try:
        body, _ = http(SANDBOX_LIST_URL, headers=aep_headers(token, conf))
        data = json.loads(body)
        return True, data.get("sandboxes") or []
    except urllib.error.HTTPError as e:
        err = flatten(e.read().decode(errors="replace"))
        return False, f"HTTP {e.code}: {err}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


def list_schedules(token, conf, sandbox):
    """GET /schedules for one sandbox (x-sandbox-name header).
    Returns (ok, schedules_or_error_string)."""
    headers = aep_headers(token, conf, sandbox)
    try:
        body, _ = http(SCHEDULES_URL, headers=headers)
        data = json.loads(body)
        if isinstance(data, list):
            return True, data
        return True, data.get("schedules") or data.get("schedule") or []
    except urllib.error.HTTPError as e:
        err = flatten(e.read().decode(errors="replace"))
        return False, f"HTTP {e.code}: {err}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


# ----------------------------------------------------------------------------
# Schedule parsing
# ----------------------------------------------------------------------------
def short_type(raw: str) -> str:
    """Map an AEP sandbox type to a compact prod/dev label."""
    t = (raw or "").lower()
    if "prod" in t:
        return "prod"
    if "dev" in t:
        return "dev"
    return raw or "?"


def cron_to_utc_time(cron):
    """Pull an HH:MM (UTC) out of a cron expression when it pins a single
    daily run time. Handles 6-field Quartz ('sec min hour ...') and 5-field
    standard ('min hour ...') crons. Returns None when the minute/hour aren't
    plain numbers (e.g. '*' or '*/2'), i.e. a recurring/sub-daily cron."""
    if not isinstance(cron, str) or not cron.strip():
        return None
    parts = cron.split()
    if len(parts) >= 6:
        minute, hour = parts[1], parts[2]
    elif len(parts) == 5:
        minute, hour = parts[0], parts[1]
    else:
        return None
    if not (minute.isdigit() and hour.isdigit()):
        return None
    return f"{int(hour):02d}:{int(minute):02d}"


def parse_schedule(sched: dict) -> dict:
    """Reduce one Query Service schedule object to the fields v2 needs.

    Returns a dict:
      enabled      'true' / 'false' / '?'
      expr         the raw recurrence expression (cron string or '@once')
      fixed_time   'HH:MM' when the cron pins one daily time, else None
      time_display what to show in the Time column: HH:MM, the cron, or '@once'
      is_cron      True when the time is a recurring cron (no fixed daily time)
      is_once      True when the schedule is a one-shot (@once)
      sid          the canonical composite schedule id (long, kept whole)
    """
    sid = sched.get("id") or sched.get("scheduleId") or "?"

    state = (sched.get("state") or sched.get("status") or "")
    if state:
        enabled = "true" if state.upper() in ("ENABLED", "ACTIVE") else "false"
    elif isinstance(sched.get("enabled"), bool):
        enabled = "true" if sched["enabled"] else "false"
    else:
        enabled = "?"

    inner = sched.get("schedule")
    if isinstance(inner, dict):
        expr = inner.get("schedule") or inner.get("cron")
    elif isinstance(inner, str):
        expr = inner
    else:
        expr = None
    expr = (expr or sched.get("cron") or "").strip()

    is_once = expr.lower() in ("@once", "once")
    fixed_time = None if is_once else cron_to_utc_time(expr)
    is_cron = bool(expr) and not is_once and fixed_time is None

    if is_once:
        time_display = "@once"
    elif fixed_time:
        time_display = fixed_time
    elif expr:
        time_display = expr            # recurring cron -> show the expression
    else:
        time_display = "-"

    return {
        "enabled": enabled,
        "expr": expr,
        "fixed_time": fixed_time,
        "time_display": time_display,
        "is_cron": is_cron,
        "is_once": is_once,
        "sid": sid,
    }


def id_has_query_name(sid: str) -> bool:
    """True when the schedule id embeds a human-readable query name.

    UUID-only ids are <org_hex>_<uuid>_<uuid>... -- every segment carries a
    digit, so no run of plain words appears. Query-named ids splice in a
    snake_case name like 'acme_order_merge_query'. Heuristic: after dropping
    the org-id prefix, split on '_' and '-' and look for >=3 consecutive
    pure-alphabetic word tokens. Three in a row clears random alpha tails
    (e.g. '..._yi0qtl') while catching every real name (the shortest seen,
    'acme_userid_ecid', is three)."""
    tail = sid.split("_", 1)[1] if "_" in sid else sid
    run = 0
    for tok in re.split(r"[_-]", tail.lower()):
        if re.fullmatch(r"[a-z]{3,}", tok):
            run += 1
            if run >= 3:
                return True
        else:
            run = 0
    return False


def classify(rec: dict) -> str:
    """SEGMENTATION / QUERY / CRON. Named ids win (the name is the signal);
    otherwise a recurring cron is CRON and a fixed-time run is SEGMENTATION."""
    if id_has_query_name(rec["sid"]):
        return "QUERY"
    if rec["is_cron"]:
        return "CRON"
    return "SEGMENTATION"


def anomaly_flags(rec: dict, sched_type: str, env: str) -> list[str]:
    """Return the anomaly flags that apply to this schedule."""
    flags = []
    if rec["enabled"] == "false":
        flags.append("DISABLED")
    if rec["is_once"]:
        flags.append("ONCE")
    ft = rec["fixed_time"]
    if ft:
        minute = int(ft.split(":")[1])
        if minute not in (0, 30):
            flags.append("ODD_TIME")
        # Zero-padded HH:MM compares correctly as a string.
        if sched_type == "SEGMENTATION" and env == "prod" and ft >= "05:00":
            flags.append("LATE")
    return flags


# ----------------------------------------------------------------------------
# Credential menu  (single choice, from v1)
# ----------------------------------------------------------------------------
def discover_creds():
    paths = []
    if CREDS_DIR.exists():
        for p in sorted(CREDS_DIR.glob("*.json")):
            if p.stem == "example":
                continue
            paths.append(p)
    return paths


def pretty_name(stem: str) -> str:
    return stem.replace("-", " ").replace("_", " ").title()


def menu(creds):
    print()
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Credential bank{ANSI['reset']}  "
          f"{ANSI['dim']}({CREDS_DIR}){ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 60 + ANSI["reset"])
    for i, p in enumerate(creds, 1):
        print(f"  {ANSI['bold']}[{i}]{ANSI['reset']} "
              f"{ANSI['yellow']}{pretty_name(p.stem):<20}{ANSI['reset']} "
              f"{ANSI['dim']}{p.name}{ANSI['reset']}")
    print(bar)
    raw = input(
        f"\nChoose a credential set by number "
        f"({ANSI['cyan']}1{ANSI['reset']}-{ANSI['cyan']}{len(creds)}{ANSI['reset']}), "
        "blank to quit: "
    ).strip()
    if not raw or not raw.isdigit():
        return None
    idx = int(raw)
    if 1 <= idx <= len(creds):
        return creds[idx - 1]
    logger.warning(f"Choice {idx} out of range.")
    return None


def pick_sandboxes(sandboxes):
    """Prompt for which sandbox(es) to audit. Returns the chosen sandbox dicts.
    Accepts numbers ('1', '1,3'), 'all', or blank to cancel."""
    print()
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Sandboxes visible to this credential{ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 60 + ANSI["reset"])
    for i, sb in enumerate(sandboxes, 1):
        name = sb.get("name", "?")
        title = sb.get("title") or name
        env = short_type(sb.get("type", ""))
        print(f"  {ANSI['bold']}[{i:>2}]{ANSI['reset']} "
              f"{ANSI['yellow']}{title:<22}{ANSI['reset']} "
              f"{ANSI['dim']}{name:<16}{ANSI['reset']} {env}")
    print(bar)
    raw = input(
        f"\nPick sandbox(es) to audit "
        f"({ANSI['cyan']}1{ANSI['reset']}, {ANSI['cyan']}1,3{ANSI['reset']}, "
        f"or {ANSI['cyan']}all{ANSI['reset']}), blank to cancel: "
    ).strip()
    if not raw:
        return []
    if raw.lower() == "all":
        return list(sandboxes)
    chosen = []
    for tok in raw.replace(",", " ").split():
        if tok.isdigit() and 1 <= int(tok) <= len(sandboxes):
            sb = sandboxes[int(tok) - 1]
            if sb not in chosen:
                chosen.append(sb)
        else:
            logger.warning(f"Ignoring invalid choice: {tok}")
    return chosen


# ----------------------------------------------------------------------------
# Table + CSV + summary
# ----------------------------------------------------------------------------
# row = (sandbox, env, enabled, time, type, flags, schedule_id)
CSV_COLUMNS = ["Sandbox", "Env", "Enabled", "Time (UTC)", "Type", "Flags",
               "Schedule ID"]

TYPE_COLOR = {
    "SEGMENTATION": ANSI["blue"],
    "QUERY": ANSI["magenta"],
    "CRON": ANSI["cyan"],
}


def write_csv(rows, stem: str) -> Path:
    """Write rows to output/batch_schedules_v2_<stem>.csv."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    safe = re.sub(r"[^0-9A-Za-z._-]+", "-", stem).strip("-") or "creds"
    path = OUTPUT_DIR / f"batch_schedules_v2_{safe}.csv"
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(CSV_COLUMNS)
        writer.writerows(rows)
    return path


# Per-sandbox tabs drop the Sandbox column (it's the tab name) but keep Env.
SHEET_COLUMNS = ["Env", "Enabled", "Time (UTC)", "Type", "Flags", "Schedule ID"]
_XLSX_HEX = "CC0000"           # red for anomalies
_XLSX_HEADER_BG = "1F4E78"     # dark blue header band


def _safe_sheet_name(name: str, used: set) -> str:
    """Excel sheet names: <=31 chars, none of []:*?/\\, and unique per book."""
    s = re.sub(r"[\[\]\:\*\?\/\\]", "-", name).strip()[:31] or "sheet"
    base, i = s, 2
    while s.lower() in used:
        suffix = f"-{i}"
        s = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(s.lower())
    return s


def write_xlsx(per_sandbox, summary, stem: str):
    """Write a tabbed workbook -- a Summary tab plus one tab per sandbox -- to
    output/batch_schedules_v2_<stem>.xlsx. Returns the path, or None if openpyxl
    is not installed (the CSV is still written either way)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        logger.warning("openpyxl not installed -- skipping XLSX "
                       "(pip install openpyxl). CSV was still written.")
        return None

    OUTPUT_DIR.mkdir(exist_ok=True)
    safe = re.sub(r"[^0-9A-Za-z._-]+", "-", stem).strip("-") or "creds"
    path = OUTPUT_DIR / f"batch_schedules_v2_{safe}.xlsx"

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=_XLSX_HEADER_BG)
    title_font = Font(bold=True, size=14)
    red_font = Font(color=_XLSX_HEX, bold=True)
    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    def _style_header(ws, ncols, row=1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = head_font
            cell.fill = head_fill
        ws.freeze_panes = ws.cell(row=row + 1, column=1)

    def _autofit(ws, widths):
        from openpyxl.utils import get_column_letter
        for idx, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(idx)].width = w

    wb = Workbook()

    # ---- Summary tab --------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = f"Schedule audit  -  {pretty_name(stem)}"
    ws["A1"].font = title_font
    ws["A2"] = f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}"
    ws["A2"].font = Font(italic=True, color="666666")

    counts = summary["counts"]
    r = 4
    ws.cell(r, 1, "Totals").font = bold
    r += 1
    totals = [
        ("SEGMENTATION", f"{counts['SEGMENTATION']}  "
                         f"({summary['seg_prod']} prod / {summary['seg_dev']} dev)"),
        ("QUERY", str(counts["QUERY"])),
        ("CRON", str(counts["CRON"])),
        ("Anomalies flagged", str(summary["anomalies"])),
    ]
    for label, val in totals:
        ws.cell(r, 1, label).font = bold
        ws.cell(r, 2, val)
        r += 1

    # Per-sandbox breakdown table.
    r += 1
    bd_header = ["Sandbox", "Env", "Total", "SEGMENTATION", "QUERY", "CRON",
                 "Anomalies"]
    header_row = r
    for c, name in enumerate(bd_header, 1):
        ws.cell(header_row, c, name)
    _style_header(ws, len(bd_header), row=header_row)
    r += 1
    for st in summary["sandbox_stats"]:
        ws.cell(r, 1, st["title"])
        ws.cell(r, 2, st["env"]).alignment = center
        ws.cell(r, 3, st["total"]).alignment = center
        ws.cell(r, 4, st["SEGMENTATION"]).alignment = center
        ws.cell(r, 5, st["QUERY"]).alignment = center
        ws.cell(r, 6, st["CRON"]).alignment = center
        an = ws.cell(r, 7, st["anomalies"])
        an.alignment = center
        if st["anomalies"]:
            an.font = red_font
        r += 1
    _autofit(ws, [24, 6, 7, 14, 8, 7, 11])

    # ---- One tab per sandbox ------------------------------------------------
    used = {"summary"}
    for title, rows in per_sandbox.items():
        sheet = wb.create_sheet(_safe_sheet_name(title, used))
        for c, name in enumerate(SHEET_COLUMNS, 1):
            sheet.cell(1, c, name)
        _style_header(sheet, len(SHEET_COLUMNS))
        for ridx, row in enumerate(rows, 2):
            # row = (sandbox, env, enabled, time, type, flags, sid); drop col 0.
            for cidx, value in enumerate(row[1:], 1):
                cell = sheet.cell(ridx, cidx, value)
                # Flags is column 5 here; redden when an anomaly is present.
                if cidx == 5 and value not in ("", "-"):
                    cell.font = red_font
        _autofit(sheet, [6, 9, 16, 14, 22, 90])

    wb.save(path)
    return path


def _short_id(sid: str, n: int = 34) -> str:
    """Drop the constant org-id prefix and truncate for the console; the CSV
    keeps the full id."""
    tail = sid.split("_", 1)[1] if "_" in sid else sid
    return tail if len(tail) <= n else f"{tail[:n]}..."


def print_table(rows):
    header = (f"{'Sandbox':<18} {'Env':<5} {'Enabled':<8} {'Time (UTC)':<14} "
              f"{'Type':<13} {'Flags':<22} Schedule ID")
    rule = (f"{'-' * 11:<18} {'-' * 3:<5} {'-' * 7:<8} {'-' * 10:<14} "
            f"{'-' * 12:<13} {'-' * 5:<22} {'-' * 11}")
    print()
    print(ANSI["bold"] + header + ANSI["reset"])
    print(ANSI["dim"] + rule + ANSI["reset"])
    for sandbox, env, enabled, time_utc, sched_type, flags, sid in rows:
        en_color = (ANSI["green"] if enabled == "true"
                    else ANSI["red"] if enabled == "err"
                    else ANSI["dim"])
        type_color = TYPE_COLOR.get(sched_type, "")
        flag_color = ANSI["red"] if flags not in ("", "-") else ANSI["dim"]
        print(f"{ANSI['yellow']}{sandbox:<18}{ANSI['reset']} "
              f"{env:<5} "
              f"{en_color}{enabled:<8}{ANSI['reset']} "
              f"{time_utc:<14} "
              f"{type_color}{sched_type:<13}{ANSI['reset']} "
              f"{flag_color}{(flags or '-'):<22}{ANSI['reset']} "
              f"{ANSI['dim']}{_short_id(sid)}{ANSI['reset']}")
    print()


def print_summary(counts, seg_prod, seg_dev, anomalies):
    print(f"{ANSI['bold']}Summary{ANSI['reset']}")
    print(ANSI["dim"] + "-" * 7 + ANSI["reset"])
    print(f"{TYPE_COLOR['SEGMENTATION']}SEGMENTATION{ANSI['reset']} schedules: "
          f"{counts['SEGMENTATION']}  ({seg_prod} prod / {seg_dev} dev)")
    print(f"{TYPE_COLOR['QUERY']}QUERY{ANSI['reset']} schedules: "
          f"{counts['QUERY']}")
    print(f"{TYPE_COLOR['CRON']}CRON{ANSI['reset']} schedules: "
          f"{counts['CRON']}")
    color = ANSI["red"] if anomalies else ANSI["green"]
    print(f"{color}Anomalies flagged:{ANSI['reset']} {anomalies}")
    print()


# ----------------------------------------------------------------------------
# Audit
# ----------------------------------------------------------------------------
def audit(path: Path):
    bar = ANSI["cyan"] + "=" * 72 + ANSI["reset"]
    print()
    print(bar)
    print(f"  {ANSI['bold']}{SCRIPT_NAME} v{SCRIPT_VERSION}{ANSI['reset']}   ({SCRIPT_DATE})")
    print(f"  by {SCRIPT_AUTHOR}")
    print(f"  {ANSI['dim']}Classify + flag anomalies across AEP batch query schedules "
          f"per sandbox{ANSI['reset']}")
    print(f"  {ANSI['bold']}Auditing schedules (v2) for "
          f"{ANSI['yellow']}{pretty_name(path.stem)}{ANSI['reset']}  "
          f"{ANSI['dim']}({path.name}){ANSI['reset']}")
    print(bar)

    try:
        conf = load_creds(path)
    except Exception as e:
        logger.error(f"Failed to load {path.name}: {e}")
        return

    try:
        resp = authenticate(conf)
    except urllib.error.HTTPError as e:
        body = e.read().decode(errors="replace")[:500]
        logger.error(f"IMS auth FAILED: HTTP {e.code} {body}")
        return
    except Exception as e:
        logger.error(f"IMS auth FAILED: {type(e).__name__}: {e}")
        return
    token = resp["access_token"]
    logger.info(f"IMS authenticated (org {conf['org_id']}).")

    ok, result = list_sandboxes(token, conf)
    if not ok:
        logger.error(f"GET /sandboxes failed: {result}")
        return
    if not result:
        logger.warning("Authenticated, but 0 sandboxes visible for this "
                       "credential.")
        return

    # Pick which sandbox(es) to audit BEFORE making any schedule calls.
    chosen = pick_sandboxes(result)
    if not chosen:
        logger.info("No sandboxes chosen. Exiting.")
        return
    logger.info(f"Auditing {len(chosen)} sandbox(es). Fetching schedules...")

    per_sandbox = {}   # tab title -> list of full row tuples (incl. sandbox col)
    sandbox_stats = []
    counts = {"SEGMENTATION": 0, "QUERY": 0, "CRON": 0}
    seg_prod = seg_dev = anomalies = 0

    for sb in chosen:
        name = sb.get("name", "?")               # technical name for header
        title = sb.get("title") or name          # friendly name for display
        env = short_type(sb.get("type", ""))
        sb_rows = []
        st = {"title": title, "env": env, "total": 0,
              "SEGMENTATION": 0, "QUERY": 0, "CRON": 0, "anomalies": 0}

        ok_s, scheds = list_schedules(token, conf, name)
        if not ok_s:
            logger.warning(f"  {name}: GET /schedules failed: {scheds}")
            sb_rows.append((title, env, "err", "-", "-", "-", "-"))
        elif not scheds:
            sb_rows.append((title, env, "-", "-", "-", "-", "-"))
        else:
            for sched in scheds:
                rec = parse_schedule(sched)
                sched_type = classify(rec)
                flags = anomaly_flags(rec, sched_type, env)
                flags_str = ",".join(flags) if flags else "-"

                counts[sched_type] += 1
                st[sched_type] += 1
                st["total"] += 1
                if sched_type == "SEGMENTATION":
                    if env == "prod":
                        seg_prod += 1
                    elif env == "dev":
                        seg_dev += 1
                if flags:
                    anomalies += 1
                    st["anomalies"] += 1

                sb_rows.append((title, env, rec["enabled"], rec["time_display"],
                                sched_type, flags_str, rec["sid"]))

        per_sandbox[title] = sb_rows
        sandbox_stats.append(st)

    flat_rows = [row for sb_rows in per_sandbox.values() for row in sb_rows]
    print_table(flat_rows)
    print_summary(counts, seg_prod, seg_dev, anomalies)

    summary = {
        "counts": counts,
        "seg_prod": seg_prod,
        "seg_dev": seg_dev,
        "anomalies": anomalies,
        "sandbox_stats": sandbox_stats,
    }

    csv_path = write_csv(flat_rows, path.stem)
    logger.info(f"CSV written:  {csv_path}")
    xlsx_path = write_xlsx(per_sandbox, summary, path.stem)
    if xlsx_path:
        logger.info(f"XLSX written: {xlsx_path}  "
                    f"({len(per_sandbox)} sandbox tab(s) + Summary)")
    logger.info(f"Done. {sum(counts.values())} schedule(s) across "
                f"{len(chosen)} sandbox(es).")


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    creds = discover_creds()
    if not creds:
        logger.error(f"No credential JSONs found in {CREDS_DIR}. "
                     f"Drop your <tenant>.json files there.")
        return

    if args:
        by_stem = {p.stem: p for p in creds}
        chosen = by_stem.get(args[0])
        if not chosen:
            logger.error(f"No credential set named {args[0]!r} "
                         f"(looked in {CREDS_DIR}).")
            return
    else:
        chosen = menu(creds)

    if not chosen:
        logger.info("Nothing chosen. Exiting.")
        return

    audit(chosen)


if __name__ == "__main__":
    main()
