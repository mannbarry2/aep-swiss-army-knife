#!/usr/bin/env python3
"""
audit_streaming_schedules.py  --  "Audience Cataloger and Inspector"
===================================================================
Pull every STREAMING audience from an AEP sandbox (default: prod) into ONE
inspectable spreadsheet, so the High Throughput Streaming (HTS) upgrade can be
triaged audience-by-audience: refactor / deactivate / let-batch / keep.

Adobe is migrating us to HTS. Streaming audiences that are ineligible auto-
convert to batch at a cut-off date. This catalogue flags, per streaming
audience: is anyone in it (profile count), what tags it carries, and -- the
crucial one for the upgrade plan -- is it orchestrated to a destination
(e.g. Facebook / Google), plus a heuristic eligibility read and a suggested
action.

It is READ-ONLY against AEP metadata. The optional --estimate pass submits
profile *preview* (estimate) jobs to fill in missing counts; those are
compute jobs but create/modify nothing.

Data source is pluggable via --source:
  api    (default) -- live AEP. Reuses the prod-proven IMS auth + cursor
         pagination from batch_eval_timing.py. Inventory + tags + eval method
         from GET /data/core/ups/audiences; merge policies from
         /ups/config/mergePolicies; destinations from Flow Service; counts from
         the audience payload, per-audience detail, and (optional) estimate.
  files  -- offline fallback so it can run before API access is granted.
         Merges an Adobe audience report .xlsx (--report) and a UI/devtools
         JSON export of the audience list (--list-json) on Audience ID.

Output: a single-sheet .xlsx (one row per streaming audience) with a summary
block at the top, the Suggested Action column colour-coded, header frozen,
columns auto-fit, Arial throughout. It lands in ./output/ with the sandbox +
timestamp in the filename. The snapshot is point-in-time, not live -- the
snapshot date is stamped in the sheet.

Stdlib only except openpyxl, which is required for the .xlsx output.
VDI-friendly.

Usage:
    python audit_streaming_schedules.py                 # cred menu, prod, api
    python audit_streaming_schedules.py prod            # creds/prod.json by stem
    python audit_streaming_schedules.py prod --sandbox=prod
    python audit_streaming_schedules.py prod --estimate # fill missing counts
    python audit_streaming_schedules.py --no-flows      # skip destination sweep
    python audit_streaming_schedules.py --source=files \
        --report=creds/audience_report.xlsx --list-json=creds/audiences.json
"""

from __future__ import annotations

import json
import logging
import re
import ssl
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------
SCRIPT_NAME = "audit_streaming_schedules"
SCRIPT_VERSION = "1.0.0"
SCRIPT_DATE = "2026-06-24"
SCRIPT_AUTHOR = "Barry Mann (barrymann.com)"

SCRIPT_DIR = Path(__file__).resolve().parent
CREDS_DIR = SCRIPT_DIR / "creds"
OUTPUT_DIR = SCRIPT_DIR / "output"

IMS_URL = "https://ims-na1.adobelogin.com/ims/token"
SANDBOX_LIST_URL = (
    "https://platform.adobe.io/data/foundation/sandbox-management/sandboxes"
)
# UPS (Unified Profile Service) surfaces -- confirmed in use in this repo
# (batch_eval_timing.py, credential_validator.py).
AUDIENCES_URL = "https://platform.adobe.io/data/core/ups/audiences"
MERGE_POLICIES_URL = "https://platform.adobe.io/data/core/ups/config/mergePolicies"
PREVIEW_URL = "https://platform.adobe.io/data/core/ups/preview"
# Flow Service -- base confirmed in repo; sub-paths are standard AEP and are
# called defensively (a wrong path degrades to a blank destination column, it
# never crashes the run).
FLOWS_URL = "https://platform.adobe.io/data/foundation/flowservice/flows"
TARGET_CONNECTIONS_URL = (
    "https://platform.adobe.io/data/foundation/flowservice/targetConnections"
)
CONNECTION_SPECS_URL = (
    "https://platform.adobe.io/data/foundation/flowservice/connectionSpecs"
)

DEFAULT_SANDBOX = "prod"          # this tool targets prod by default
PAGE_LIMIT = 100
MAX_PAGES = 500                   # cursor-loop backstop (~50k items)
ESTIMATE_POLL_SECONDS = 3
ESTIMATE_MAX_POLLS = 20           # ~1 min per audience worst case

# Name / tag substrings that mark an audience as test / throwaway / duplicate.
TEST_PATTERNS = ("test", "learn", "perftest", "copy", "demo", "poc", "sandbox")

# Hours per recurrence unit, for the >24h event-lookback eligibility heuristic.
UNIT_HOURS = {
    "second": 1 / 3600, "minute": 1 / 60, "hour": 1, "day": 24,
    "week": 168, "month": 730, "year": 8760,
}

DEFAULT_SCOPES = (
    "openid,AdobeID,read_organizations,"
    "additional_info.projectedProductContext,session"
)

SUGGESTED_ACTIONS = (
    "Keep (orchestrated)",
    "Keep (eligible)",
    "Refactor (SoS split)",
    "Deactivate? (dormant)",
    "Deactivate? (test)",
)

# ----------------------------------------------------------------------------
# ANSI / logging  (house style, shared with the other tools in this repo)
# ----------------------------------------------------------------------------
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        h = kernel32.GetStdHandle(-11)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
            kernel32.SetConsoleMode(h, mode.value | 0x0004)
    except Exception:
        pass

ANSI = {
    "reset": "\033[0m", "bold": "\033[1m", "dim": "\033[2m",
    "red": "\033[31m", "green": "\033[32m", "yellow": "\033[33m",
    "blue": "\033[34m", "magenta": "\033[35m", "cyan": "\033[36m",
}
LEVEL_COLOR = {
    "DEBUG": ANSI["dim"], "INFO": ANSI["green"],
    "WARNING": ANSI["yellow"],
    "ERROR": ANSI["red"] + ANSI["bold"],
    "CRITICAL": ANSI["red"] + ANSI["bold"],
}


class ColoredFormatter(logging.Formatter):
    def format(self, record):
        color = LEVEL_COLOR.get(record.levelname, "")
        ts = self.formatTime(record, "%H:%M:%S")
        return (
            f"{ANSI['dim']}{ts}{ANSI['reset']} "
            f"{color}[{record.levelname:<7}]{ANSI['reset']} "
            f"{record.getMessage()}"
        )


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(ColoredFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger("audience_cataloger")
SSL_CTX = ssl._create_unverified_context()


# ----------------------------------------------------------------------------
# HTTP / IMS / credential helpers  (lifted verbatim from batch_eval_timing.py)
# ----------------------------------------------------------------------------
def http(url, method="GET", headers=None, data=None, timeout=60):
    """Stdlib-only HTTP. Returns response bytes; raises HTTPError on 4xx/5xx."""
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read()


def load_creds(path: Path):
    raw = json.loads(path.read_text(encoding="utf-8"))
    conf = {
        k: v.strip() if isinstance(v, str) else v
        for k, v in raw.items()
        if not k.startswith("_")
    }
    for key in ("client_id", "client_secret", "org_id"):
        if not conf.get(key):
            raise ValueError(f"Missing required key {key!r} in {path.name}")
    return conf


def authenticate(conf):
    payload = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": conf["client_id"],
        "client_secret": conf["client_secret"],
        "scope": conf.get("scopes") or DEFAULT_SCOPES,
    }).encode("utf-8")
    body = http(
        conf.get("oauth_url") or IMS_URL,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data=payload,
    )
    return json.loads(body)["access_token"]


def aep_headers(token, conf, sandbox):
    return {
        "Authorization": f"Bearer {token}",
        "x-api-key": conf.get("api_key") or conf["client_id"],
        "x-gw-ims-org-id": conf["org_id"],
        "x-sandbox-name": sandbox,
        "Accept": "application/json",
    }


def discover_creds():
    paths = []
    if CREDS_DIR.exists():
        for p in sorted(CREDS_DIR.glob("*.json")):
            if p.stem == "example":
                continue
            paths.append(p)
    return paths


def menu(creds):
    """Prompt for ONE credential set (this tool targets a single sandbox)."""
    print()
    bar = ANSI["cyan"] + "=" * 70 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Credential bank{ANSI['reset']}  "
          f"{ANSI['dim']}({CREDS_DIR}){ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 70 + ANSI["reset"])
    for i, p in enumerate(creds, 1):
        print(f"  {ANSI['bold']}{i:>2}{ANSI['reset']}  "
              f"{ANSI['yellow']}{p.stem:<20}{ANSI['reset']} "
              f"{ANSI['dim']}{p.name}{ANSI['reset']}")
    print(bar)
    raw = input(f"\nPick a credential set by number "
                f"({ANSI['cyan']}1{ANSI['reset']}), blank to quit: ").strip()
    if not raw:
        return None
    if raw.isdigit() and 1 <= int(raw) <= len(creds):
        return creds[int(raw) - 1]
    logger.warning(f"Invalid choice: {raw}")
    return None


# ----------------------------------------------------------------------------
# Time helpers  (from batch_eval_timing.py)
# ----------------------------------------------------------------------------
def to_dt(value) -> "datetime | None":
    if value in (None, "", 0):
        return None
    if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
        n = float(value)
        if n > 1e12:
            n /= 1000.0
        try:
            return datetime.fromtimestamp(n, tz=timezone.utc)
        except (OverflowError, OSError, ValueError):
            return None
    if isinstance(value, str):
        s = value.replace("Z", "+00:00")
        try:
            dt = datetime.fromisoformat(s)
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except ValueError:
            return None
    return None


def fmt_dt(dt: "datetime | None") -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""


# ----------------------------------------------------------------------------
# Audience field extraction  (eval method / tags / created from the repo;
# the rest added for the cataloguing columns -- all null-safe)
# ----------------------------------------------------------------------------
def evaluation_method(aud: dict) -> str:
    ev = aud.get("evaluationInfo") or {}
    if (ev.get("batch") or {}).get("enabled"):
        return "batch"
    if (ev.get("continuous") or {}).get("enabled"):
        return "streaming"
    if (ev.get("synchronous") or {}).get("enabled"):
        return "edge"
    return "?"


def audience_id(aud: dict) -> str:
    return str(aud.get("id") or aud.get("audienceId") or "")


def audience_created(aud: dict) -> "datetime | None":
    for key in ("createEpoch", "creationTime", "createdAt", "created"):
        dt = to_dt(aud.get(key))
        if dt:
            return dt
    return None


def audience_modified(aud: dict) -> "datetime | None":
    for key in ("updateEpoch", "updateTime", "modifiedAt", "lastModified", "updated"):
        dt = to_dt(aud.get(key))
        if dt:
            return dt
    return None


def audience_tags(aud: dict) -> "list[str]":
    """Collect human-meaningful tags from whichever fields are present. AEP
    exposes folder tags under `tags` (sometimes a dict of name->[values]) and
    governance labels under `labels`."""
    out: "list[str]" = []
    tags = aud.get("tags")
    if isinstance(tags, dict):
        for k, v in tags.items():
            if isinstance(v, list):
                out.extend(str(x) for x in v)
            elif v:
                out.append(str(v))
            else:
                out.append(str(k))
    elif isinstance(tags, list):
        out.extend(str(x) for x in tags)
    return out


def audience_created_by(aud: dict) -> str:
    for key in ("createdBy", "createUserId", "createUser", "owner"):
        v = aud.get(key)
        if v:
            return str(v)
    return ""


def audience_folder(aud: dict) -> str:
    for key in ("folderId", "folder", "folderName"):
        v = aud.get(key)
        if v:
            return str(v)
    return ""


def audience_pql(aud: dict) -> str:
    """Best-effort extraction of the PQL definition text across the field
    shapes AEP uses."""
    for key in ("expression", "profileQueryExpression", "pqlExpression"):
        v = aud.get(key)
        if isinstance(v, dict):
            t = v.get("value") or v.get("expression") or v.get("pql")
            if t:
                return str(t)
        elif isinstance(v, str) and v.strip():
            return v
    return ""


# Profile-count extraction: AEP has exposed counts under several shapes; probe
# the audience object (list payload, then per-audience detail) for any of them.
_COUNT_KEYS = ("profileCount", "totalProfiles", "profiles", "count", "totalRows")


def _dig_count(obj) -> "int | None":
    """Find a plausible profile-count integer in obj. The confirmed AEP shape
    is metrics.data.totalProfiles (current members); we also probe the older
    nest shapes. 'data' is included so metrics.data.totalProfiles is reached."""
    if not isinstance(obj, dict):
        return None
    for k in _COUNT_KEYS:
        v = obj.get(k)
        if isinstance(v, (int, float)) and v >= 0:
            return int(v)
        if isinstance(v, str) and v.isdigit():
            return int(v)
    for nest in ("metrics", "_metrics", "lifecycleMetrics", "audienceMetrics",
                 "data"):
        sub = obj.get(nest)
        if isinstance(sub, dict):
            c = _dig_count(sub)
            if c is not None:
                return c
    return None


def audience_count_from_obj(aud: dict) -> "int | None":
    return _dig_count(aud)


def audience_count_snapshot(aud: dict) -> "datetime | None":
    """When the profile count was last refreshed (metrics.updateEpoch). Counts
    are a point-in-time snapshot from the last segment job, not live -- with
    multiple merge policies/snapshot tables, audiences can refresh at different
    times, so this exposes each count's freshness."""
    metrics = aud.get("metrics")
    if isinstance(metrics, dict):
        dt = to_dt(metrics.get("updateEpoch"))
        if dt:
            return dt
        data = metrics.get("data")
        if isinstance(data, dict):
            return to_dt(data.get("updateEpoch"))
    return None


def audience_breakdown(aud: dict) -> "tuple[int|None,int|None,int|None]":
    """(realized, existing, exited) = (joiners, stable, leavers) when AEP
    exposes the lifecycle breakdown; (None, None, None) otherwise."""
    def _i(x):
        return int(x) if isinstance(x, (int, float)) or (isinstance(x, str) and x.isdigit()) else None

    # Confirmed AEP shape: metrics.data.totalProfilesByStatus
    # {realized, existing, exited}.
    metrics = aud.get("metrics")
    if isinstance(metrics, dict):
        data = metrics.get("data")
        if isinstance(data, dict):
            byst = data.get("totalProfilesByStatus")
            if isinstance(byst, dict):
                return (_i(byst.get("realized")), _i(byst.get("existing")),
                        _i(byst.get("exited")))
    # Older/alternate shapes.
    for nest in ("metrics", "_metrics", "lifecycleMetrics", "audienceMetrics"):
        sub = aud.get(nest)
        if not isinstance(sub, dict):
            continue
        realized = sub.get("realized") or sub.get("newProfiles") or sub.get("realizedProfiles")
        existing = sub.get("existing") or sub.get("existingProfiles")
        exited = sub.get("exited") or sub.get("exitedProfiles") or sub.get("leavers")
        if any(x is not None for x in (realized, existing, exited)):
            return _i(realized), _i(existing), _i(exited)
    return None, None, None


# ----------------------------------------------------------------------------
# AEP fetchers  (api mode)
# ----------------------------------------------------------------------------
def _paged(url, headers, array_keys, label):
    """Cursor pagination shared by audiences/merge policies: ?limit&start with
    the next cursor at _page.next. Stops on empty/short page or no cursor."""
    out: "list[dict]" = []
    start = None
    page = 0
    while page < MAX_PAGES:
        page += 1
        params = {"limit": PAGE_LIMIT}
        if start:
            params["start"] = start
        full = f"{url}?{urllib.parse.urlencode(params)}"
        try:
            body = http(full, headers=headers)
        except urllib.error.HTTPError as e:
            detail = e.read().decode(errors="replace")[:300]
            logger.error(f"{label} list failed: HTTP {e.code} {detail}")
            break
        except Exception as e:
            logger.error(f"{label} list failed: {type(e).__name__}: {e}")
            break
        data = json.loads(body) or {}
        batch = []
        for k in array_keys:
            if isinstance(data.get(k), list):
                batch = data[k]
                break
        out.extend(batch)
        nxt = (data.get("_page") or {}).get("next")
        if not batch or len(batch) < PAGE_LIMIT or not nxt:
            break
        start = nxt
    return out


def fetch_audiences(headers) -> "list[dict]":
    auds = _paged(AUDIENCES_URL, headers, ("children", "segments"), "Audience")
    logger.info(f"  fetched {len(auds)} audience(s) total")
    return auds


def fetch_merge_policies(headers) -> "dict[str, dict]":
    """id -> merge policy object. is-active-on-edge is read off each policy."""
    pols = _paged(MERGE_POLICIES_URL, headers, ("children", "mergePolicies"), "Merge policy")
    return {str(p.get("id")): p for p in pols if p.get("id")}


def merge_policy_edge_active(pol: dict) -> str:
    """Y/N/? for whether a merge policy projects to the edge (active on edge)."""
    if not isinstance(pol, dict):
        return "?"
    for k in ("isActiveOnEdge", "activeOnEdge", "edgeEnabled"):
        if isinstance(pol.get(k), bool):
            return "Y" if pol[k] else "N"
    ig = pol.get("identityGraph") or {}
    # On-edge merge policies use the "none" identity-stitching type.
    if isinstance(ig, dict) and ig.get("type"):
        return "Y" if str(ig["type"]).lower() in ("none",) else "N"
    return "?"


# --- Flow Service: audience -> destination mapping -----------------------
def _flow_paged(url, headers, label):
    """Flow Service pagination follows _links.next.href (continuationToken).
    The next href is relative to the flowservice base (e.g. '/flows?...'), NOT
    the host root, so it is resolved against the base path of `url`."""
    base = url.rsplit("/", 1)[0]          # .../foundation/flowservice
    out: "list[dict]" = []
    full = f"{url}?{urllib.parse.urlencode({'limit': PAGE_LIMIT})}"
    page = 0
    while full and page < MAX_PAGES:
        page += 1
        try:
            body = http(full, headers=headers)
        except urllib.error.HTTPError as e:
            detail = e.read().decode(errors="replace")[:200]
            logger.warning(f"{label} fetch failed: HTTP {e.code} {detail}")
            break
        except Exception as e:
            logger.warning(f"{label} fetch failed: {type(e).__name__}: {e}")
            break
        data = json.loads(body) or {}
        items = data.get("items") or data.get("flows") or data.get("children") or []
        out.extend(items)
        nxt = ((data.get("_links") or {}).get("next") or {}).get("href")
        if not nxt or not items:
            break
        # next href is relative to the flowservice base (e.g. '/flows?...').
        full = nxt if nxt.startswith("http") else f"{base}{nxt}"
    return out


# Connection-spec names that are internal AEP plumbing (ingestion, profile
# store, identity), never marketing destinations. Anything NOT matching these
# is treated as a candidate destination; the segmentSelectors in the flow
# detail then confirm it is an audience-activation flow.
_INTERNAL_SPECS = (
    "datalake", "dwh", "activation-ups", "activation-datalake",
    "entity lookup", "ups-atlas", "uis identities", "uis graphs",
    "dataretention", "ups segments", "ups-snapshot",
)


def _is_internal_spec(name) -> bool:
    n = (name or "").lower()
    return any(s in n for s in _INTERNAL_SPECS)


def _fetch_flow_detail(headers, fid) -> "dict | None":
    """GET one flow's full detail. The /flows LIST omits `transformations`
    (where the activated audiences live), so each activation flow is fetched
    individually."""
    if not fid:
        return None
    try:
        body = http(f"{FLOWS_URL}/{urllib.parse.quote(str(fid), safe='')}",
                    headers=headers, timeout=30)
    except Exception as e:
        logger.warning(f"  flow {fid} detail fetch failed: {type(e).__name__}")
        return None
    try:
        data = json.loads(body) or {}
    except (ValueError, TypeError):
        return None
    items = data.get("items")
    if isinstance(items, list) and items:
        return items[0]
    return data


def _segment_ids_from_flow(detail) -> "list[str]":
    """Audience/segment ids activated by a flow, from
    transformations[].params.segmentSelectors.selectors[].value.id (the
    confirmed AEP activation shape; PLATFORM_SEGMENT and friends)."""
    out: "list[str]" = []
    for tr in detail.get("transformations") or []:
        sels = (((tr.get("params") or {}).get("segmentSelectors")) or {}).get("selectors")
        if not isinstance(sels, list):
            continue
        for s in sels:
            v = s.get("value") if isinstance(s, dict) else None
            if isinstance(v, dict):
                sid = str(v.get("id") or v.get("systemSegmentId") or "")
                if sid:
                    out.append(sid)
    return out


def build_destination_map(headers) -> "tuple[dict, bool]":
    """Return (audience_id -> list[(label, active)], ok).

    Activation flows reference their audiences in the flow DETAIL
    (transformations[].params.segmentSelectors) -- NOT in the /flows list
    payload, and never inside the audience object itself. So: list all flows,
    keep the ones whose target connection is a real destination (anything not
    internal UPS/datalake plumbing), GET each one's detail, and map every
    selector audience-id to that flow's destination. Many-to-many: one flow can
    export many audiences, and one audience can feed many destinations. The
    label resolves targetConnection -> connectionSpec ('Facebook Custom
    Audience', 'Google Ads', ...) plus the flow's own name."""
    flows = _flow_paged(FLOWS_URL, headers, "Flows")
    if not flows:
        return {}, False
    tcs = _flow_paged(TARGET_CONNECTIONS_URL, headers, "Target connections")
    specs = _flow_paged(CONNECTION_SPECS_URL, headers, "Connection specs")
    spec_name = {str(s.get("id")): (s.get("name") or "") for s in specs if s.get("id")}
    tc_map = {}
    for tc in tcs:
        tcid = str(tc.get("id") or "")
        if not tcid:
            continue
        sid = str((tc.get("connectionSpec") or {}).get("id") or "")
        tc_map[tcid] = {"name": tc.get("name") or "", "spec": spec_name.get(sid, "")}

    # Candidate activation flows: at least one target connection that is a real
    # (non-internal) destination. Confirmed below via segmentSelectors.
    candidates = []
    for f in flows:
        for t in f.get("targetConnectionIds") or []:
            tc = tc_map.get(str(t))
            if tc and not _is_internal_spec(tc["spec"]):
                candidates.append(f)
                break
    logger.info(f"  scanning {len(candidates)} candidate destination flow(s) "
                f"of {len(flows)} total...")

    aud_dest: "dict[str, list[tuple[str, bool]]]" = {}
    for f in candidates:
        detail = _fetch_flow_detail(headers, f.get("id"))
        if not detail:
            continue
        seg_ids = _segment_ids_from_flow(detail)
        if not seg_ids:
            continue
        labels = []
        for t in detail.get("targetConnectionIds") or f.get("targetConnectionIds") or []:
            tc = tc_map.get(str(t))
            if tc and not _is_internal_spec(tc["spec"]):
                labels.append(tc["spec"] or tc["name"] or str(t))
        spec_label = ", ".join(dict.fromkeys(l for l in labels if l)) or "(destination)"
        flow_name = detail.get("name") or f.get("name") or ""
        label = f"{spec_label} ({flow_name})" if flow_name else spec_label
        state = str(detail.get("state") or f.get("state") or "").lower()
        active = state in ("enabled", "active", "")  # blank state -> assume active
        for sid in seg_ids:
            aud_dest.setdefault(sid, []).append((label, active))
    logger.info(f"  resolved destinations for {len(aud_dest)} audience(s).")
    return {"_map": aud_dest}, True


def destinations_for(dest_index, aud_id) -> "list[tuple[str,bool]]":
    """[(destination_label, is_active), ...] for every flow that activates this
    audience id, from the index built by build_destination_map."""
    return list((dest_index.get("_map") or {}).get(str(aud_id), []))


def fetch_count_detail(headers, aud_id) -> "int | None":
    """GET one audience's detail and dig for a profile count."""
    if not aud_id:
        return None
    try:
        body = http(f"{AUDIENCES_URL}/{urllib.parse.quote(aud_id, safe='')}",
                    headers=headers, timeout=30)
    except Exception:
        return None
    try:
        return _dig_count(json.loads(body) or {})
    except (ValueError, TypeError):
        return None


def estimate_count(headers, aud) -> "int | None":
    """Submit a profile *preview* (estimate) job for an audience's PQL and poll
    for its profile count. Read-only/compute-only -- creates no audience.
    Returns None on any failure (caller marks the cell MISSING)."""
    pql = audience_pql(aud)
    if not pql:
        return None
    mp = aud.get("mergePolicyId") or ""
    payload = {
        "predicateExpression": pql,
        "predicateType": "pql/text",
        "graphType": "pdg",
    }
    if mp:
        payload["mergePolicyId"] = mp
    post_headers = dict(headers)
    post_headers["Content-Type"] = "application/json"
    try:
        body = http(PREVIEW_URL, method="POST", headers=post_headers,
                    data=json.dumps(payload).encode("utf-8"), timeout=30)
        preview_id = (json.loads(body) or {}).get("previewId")
    except urllib.error.HTTPError as e:
        raise e          # let caller auto-disable estimate on a hard failure
    except Exception:
        return None
    if not preview_id:
        return None
    poll_url = f"{PREVIEW_URL}/{urllib.parse.quote(str(preview_id), safe='')}"
    for _ in range(ESTIMATE_MAX_POLLS):
        time.sleep(ESTIMATE_POLL_SECONDS)
        try:
            body = http(poll_url, headers=headers, timeout=30)
            data = json.loads(body) or {}
        except Exception:
            return None
        state = str(data.get("state") or data.get("status") or "").upper()
        c = _dig_count(data)
        if c is not None:
            return c
        if state in ("RESULT_READY", "SUCCEEDED", "COMPLETED", "DONE"):
            return _dig_count(data)
        if state in ("FAILED", "ERROR", "KILLED"):
            return None
    return None


# ----------------------------------------------------------------------------
# Classification: eligibility / inspector flags / suggested action
# ----------------------------------------------------------------------------
def streaming_eligibility(aud: dict, report_row: "dict | None") -> "tuple[str, str]":
    """(eligible Y/N/?, reason). In files mode, the report's Impact
    Classification + Batch Impacted Attributes are authoritative. In api mode
    it's a transparent heuristic over the PQL: an event lookback > 24h is the
    concrete HTS-ineligibility signal we can detect."""
    if report_row:
        impact = str(report_row.get("impact_classification") or "").strip()
        batch_attrs = str(report_row.get("batch_impacted_attributes") or "").strip()
        using_batch = str(report_row.get("using_batch_data") or "").strip().lower()
        if "batch" in impact.lower() or using_batch in ("true", "yes", "y") or batch_attrs:
            why = impact or "Batch Data Inclusion"
            if batch_attrs:
                why += f" ({batch_attrs})"
            return "N", why
        if impact:
            return "Y", impact
    pql = audience_pql(aud)
    if not pql:
        return "?", "no expression to inspect (heuristic)"
    reasons = []
    for m in re.finditer(
            r'inLast\s*\(\s*(\d+)\s*,\s*["\']?'
            r'(second|minute|hour|day|week|month|year)', pql, re.I):
        n = int(m.group(1))
        unit = m.group(2).lower()
        if n * UNIT_HOURS[unit] > 24:
            reasons.append(f"event lookback {n} {unit}(s) > 24h")
    if reasons:
        return "N", "; ".join(dict.fromkeys(reasons))
    return "Y", "no >24h event lookback found (heuristic)"


def is_test_audience(name: str, tags: "list[str]") -> bool:
    hay = (name or "").lower() + " " + " ".join(tags).lower()
    return any(p in hay for p in TEST_PATTERNS)


def inspector_flags(count, realized, existing, exited, is_test) -> "list[str]":
    flags = []
    if count == 0:
        flags.append("DORMANT")
    if (exited is not None and (realized in (0, None)) and (count in (0, None))
            and (existing in (0, None)) and exited and exited > 0):
        flags.append("LIKELY-DEAD")
    if is_test:
        flags.append("TEST/DUPE")
    return flags


def suggested_action(count, eligible, is_test, orchestrated_active) -> str:
    # A live destination attached means it is in production use -- never
    # recommend deactivating it, even if it is test-named or shows 0 profiles
    # (the count is a point-in-time snapshot and may simply be stale).
    if orchestrated_active:
        if eligible == "N":
            return "Refactor (SoS split)"
        return "Keep (orchestrated)"
    if is_test:
        return "Deactivate? (test)"
    if count == 0:
        return "Deactivate? (dormant)"
    if eligible == "N":
        return "Refactor (SoS split)"
    return "Keep (eligible)"


# ----------------------------------------------------------------------------
# files-mode loaders  (offline fallback)
# ----------------------------------------------------------------------------
def _norm_header(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(s or "").lower()).strip("_")


# Maps a normalised report header -> the row key we store it under.
_REPORT_FIELDS = {
    "impact_classification": "impact_classification",
    "using_batch_data": "using_batch_data",
    "batch_impacted_attributes": "batch_impacted_attributes",
    "ss_q_day": "ss_q_day",
    "ssq_day": "ss_q_day",
    "merge_policy": "merge_policy",
    "audience_id": "id",
    "id": "id",
    "audience": "name",
    "name": "name",
}


def load_report_xlsx(path: Path) -> "dict[str, dict]":
    """Read an Adobe audience report .xlsx into {audience_id -> field dict}.
    Finds the header row, then maps known columns by fuzzy header name."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.error("openpyxl required to read the report .xlsx "
                     "(pip install openpyxl).")
        return {}
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    # Find the header row: the first row that contains an id/name-ish column.
    header_idx = 0
    for i, r in enumerate(rows[:10]):
        norm = [_norm_header(c) for c in r]
        if any(h in _REPORT_FIELDS for h in norm):
            header_idx = i
            break
    headers = [_norm_header(c) for c in rows[header_idx]]
    out: "dict[str, dict]" = {}
    for r in rows[header_idx + 1:]:
        rec = {}
        for h, v in zip(headers, r):
            key = _REPORT_FIELDS.get(h)
            if key:
                rec[key] = v
        aid = str(rec.get("id") or "").strip()
        if aid:
            out[aid] = rec
    logger.info(f"  report: {len(out)} row(s) keyed by audience id")
    return out


def load_list_json(path: Path) -> "list[dict]":
    """Read a UI/devtools JSON export of the audience list into a list of
    audience objects (handles a bare array or the common envelopes)."""
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, list):
        return data
    for k in ("children", "segments", "audiences", "data", "items"):
        if isinstance(data.get(k), list):
            return data[k]
    return []


# ----------------------------------------------------------------------------
# Row assembly
# ----------------------------------------------------------------------------
COLUMNS = [
    "Name", "ID", "Description", "Lifecycle", "Created by", "Created",
    "Modified", "Tags", "Folder", "Profile count", "Realized (joiners)",
    "Existing (stable)", "Exited (leavers)", "Count snapshot", "Merge policy ID",
    "Merge policy", "Active on edge?", "Streaming-eligible? (heuristic)",
    "Eligibility reason", "SS Q/Day flux", "Destination orchestrated?",
    "Destination names", "Feeds AJO journey? (manual)", "Inspector flags",
    "Suggested action",
]
# Numeric columns -- written as numbers (not strings) for sortable XLSX cells.
MISSING = "MISSING"


def build_row(aud, *, count, realized, existing, exited, mp_map, dests,
              report_row):
    name = aud.get("name") or "(unnamed)"
    aid = audience_id(aud)
    tags = audience_tags(aud)
    is_test = is_test_audience(name, tags)
    eligible, reason = streaming_eligibility(aud, report_row)

    mp_id = str(aud.get("mergePolicyId") or (report_row or {}).get("merge_policy") or "")
    mp = mp_map.get(mp_id) if mp_map else None
    mp_name = (mp or {}).get("name") or (str((report_row or {}).get("merge_policy") or "")
                                         if not mp else "")
    edge = merge_policy_edge_active(mp) if mp else "?"

    if dests:
        active = any(a for _, a in dests)
        names = ", ".join(dict.fromkeys(lbl for lbl, _ in dests))
        orchestrated = "Y" if active else "inactive only"
    else:
        active = False
        orchestrated = "N"
        names = ""

    flags = inspector_flags(count, realized, existing, exited, is_test)
    action = suggested_action(count, eligible, is_test, active)

    count_cell = count if isinstance(count, int) else MISSING
    ssq = (report_row or {}).get("ss_q_day", "")

    return {
        "Name": name,
        "ID": aid,
        "Description": aud.get("description") or "",
        "Lifecycle": aud.get("lifecycleState") or aud.get("lifecycle") or "",
        "Created by": audience_created_by(aud),
        "Created": fmt_dt(audience_created(aud)),
        "Modified": fmt_dt(audience_modified(aud)),
        "Tags": ", ".join(tags),
        "Folder": audience_folder(aud),
        "Profile count": count_cell,
        "Realized (joiners)": realized if realized is not None else "",
        "Existing (stable)": existing if existing is not None else "",
        "Exited (leavers)": exited if exited is not None else "",
        "Count snapshot": fmt_dt(audience_count_snapshot(aud)),
        "Merge policy ID": mp_id,
        "Merge policy": mp_name,
        "Active on edge?": edge,
        "Streaming-eligible? (heuristic)": eligible,
        "Eligibility reason": reason,
        "SS Q/Day flux": ssq if ssq is not None else "",
        "Destination orchestrated?": orchestrated,
        "Destination names": names,
        "Feeds AJO journey? (manual)": "",   # not API-derivable -- owner confirms
        "Inspector flags": ", ".join(flags),
        "Suggested action": action,
        "_action": action,                   # internal: for colour + summary
    }


# ----------------------------------------------------------------------------
# Output: single-sheet XLSX
# ----------------------------------------------------------------------------
_ACTION_FILL = {
    "Keep (orchestrated)": "BDD7EE",      # blue -- live destination attached
    "Keep (eligible)": "C6EFCE",          # green
    "Refactor (SoS split)": "FFEB9C",     # amber
    "Deactivate? (dormant)": "D9D9D9",    # grey
    "Deactivate? (test)": "FFC7CE",       # red
}


def write_xlsx(rows, summary, sandbox, stamp, snapshot, outdir) -> "Path | None":
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed -- no output written "
                     "(pip install openpyxl).")
        return None

    outdir.mkdir(parents=True, exist_ok=True)
    path = outdir / f"audience_catalog_{sandbox}_{stamp}.xlsx"

    arial = Font(name="Arial", size=10)
    arial_bold = Font(name="Arial", size=10, bold=True)
    title_font = Font(name="Arial", size=14, bold=True)
    note_font = Font(name="Arial", size=9, italic=True, color="666666")
    head_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E78")
    miss_font = Font(name="Arial", size=10, bold=True, color="CC0000")
    center = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "Streaming audiences"

    def put(r, c, v, font=arial):
        cell = ws.cell(r, c, v)
        cell.font = font
        return cell

    # --- Summary block at the top -------------------------------------------
    put(1, 1, "Audience Cataloger and Inspector", title_font)
    put(2, 1, f"Sandbox: {sandbox}   |   Streaming audiences only", note_font)
    put(3, 1, f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}   |   "
              f"Snapshot (point-in-time, not live): {snapshot}", note_font)

    r = 5
    put(r, 1, "Summary -- counts by suggested action", arial_bold)
    r += 1
    for action in SUGGESTED_ACTIONS:
        put(r, 1, action)
        cell = put(r, 2, summary["by_action"].get(action, 0))
        cell.alignment = center
        fill = _ACTION_FILL.get(action)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        r += 1
    put(r, 1, "Total streaming audiences", arial_bold)
    put(r, 2, summary["total"], arial_bold).alignment = center
    r += 1
    put(r, 1, "Profile counts MISSING", arial_bold)
    mc = put(r, 2, summary["missing_counts"],
             miss_font if summary["missing_counts"] else arial_bold)
    mc.alignment = center
    r += 1
    put(r, 1, "Orchestrated to a destination")
    put(r, 2, summary["orchestrated"]).alignment = center
    r += 2

    # --- Header row (frozen) ------------------------------------------------
    header_row = r
    for c, name in enumerate(COLUMNS, 1):
        cell = put(header_row, c, name, head_font)
        cell.fill = head_fill
    ws.freeze_panes = ws.cell(header_row + 1, 1)
    action_col = COLUMNS.index("Suggested action") + 1
    count_col = COLUMNS.index("Profile count") + 1

    # --- Data rows ----------------------------------------------------------
    for r_off, row in enumerate(rows, header_row + 1):
        for c, name in enumerate(COLUMNS, 1):
            val = row.get(name, "")
            cell = put(r_off, c, val)
            if c == action_col:
                fill = _ACTION_FILL.get(row.get("_action"))
                if fill:
                    cell.fill = PatternFill("solid", fgColor=fill)
            if c == count_col and val == MISSING:
                cell.font = miss_font

    # --- Auto-fit widths ----------------------------------------------------
    for c, name in enumerate(COLUMNS, 1):
        width = len(name)
        for row in rows:
            width = max(width, len(str(row.get(name, ""))))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width + 2, 10), 60)

    wb.save(path)
    return path


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------
def parse_args(argv):
    opts = {
        "name": None, "sandbox": None, "source": "api",
        "report": None, "list_json": None, "output": None,
        "estimate": False, "flows": True, "all_methods": False,
    }
    for a in argv:
        if a.startswith("--sandbox="):
            opts["sandbox"] = a.split("=", 1)[1].strip() or None
        elif a.startswith("--source="):
            opts["source"] = a.split("=", 1)[1].strip().lower() or "api"
        elif a.startswith("--report="):
            opts["report"] = a.split("=", 1)[1].strip() or None
        elif a.startswith("--list-json="):
            opts["list_json"] = a.split("=", 1)[1].strip() or None
        elif a.startswith("--output="):
            opts["output"] = a.split("=", 1)[1].strip() or None
        elif a == "--estimate":
            opts["estimate"] = True
        elif a == "--no-estimate":
            opts["estimate"] = False
        elif a == "--no-flows":
            opts["flows"] = False
        elif a in ("--all-methods", "--all"):
            opts["all_methods"] = True
        elif a.startswith("-"):
            continue
        else:
            opts["name"] = a
    return opts


def banner(sandbox, source):
    bar = ANSI["cyan"] + "=" * 72 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}{SCRIPT_NAME} v{SCRIPT_VERSION}{ANSI['reset']}   ({SCRIPT_DATE})")
    print(f"  by {SCRIPT_AUTHOR}")
    print(f"  {ANSI['dim']}Catalogue + triage streaming audiences for the HTS upgrade "
          f"(read-only){ANSI['reset']}")
    print(f"  {ANSI['bold']}Audience Cataloger and Inspector{ANSI['reset']}  "
          f"{ANSI['dim']}(streaming audiences; read-only){ANSI['reset']}")
    print(f"  {ANSI['bold']}Source:{ANSI['reset']}   {ANSI['yellow']}{source}{ANSI['reset']}")
    print(f"  {ANSI['bold']}Sandbox:{ANSI['reset']}  {ANSI['yellow']}{sandbox}{ANSI['reset']}")
    print(bar)


def summarize(rows):
    by_action = Counter(r["_action"] for r in rows)
    return {
        "total": len(rows),
        "by_action": by_action,
        "missing_counts": sum(1 for r in rows if r["Profile count"] == MISSING),
        "orchestrated": sum(1 for r in rows
                            if r["Destination orchestrated?"].startswith("Y")),
    }


def print_console_summary(rows, summary):
    print()
    print(f"  {ANSI['bold']}Streaming audiences catalogued: "
          f"{summary['total']}{ANSI['reset']}")
    for action in SUGGESTED_ACTIONS:
        n = summary["by_action"].get(action, 0)
        print(f"     {ANSI['cyan']}{action:<24}{ANSI['reset']} {n}")
    mc = summary["missing_counts"]
    col = ANSI["red"] if mc else ANSI["green"]
    print(f"     {col}{'Profile counts MISSING':<24}{ANSI['reset']} {mc}")
    print(f"     {ANSI['cyan']}{'Orchestrated':<24}{ANSI['reset']} "
          f"{summary['orchestrated']}")


# ----------------------------------------------------------------------------
# Modes
# ----------------------------------------------------------------------------
def run_api(opts) -> "tuple[list[dict], str, str] | None":
    creds = discover_creds()
    if not creds:
        logger.error(f"No credential JSONs found in {CREDS_DIR}. "
                     f"Drop your <tenant>.json files there.")
        return None
    if opts["name"]:
        path = {p.stem: p for p in creds}.get(opts["name"])
        if not path:
            logger.error(f"No credential set named {opts['name']!r} in {CREDS_DIR}")
            return None
    else:
        path = menu(creds)
    if not path:
        logger.info("Nothing chosen. Exiting.")
        return None

    try:
        conf = load_creds(path)
    except Exception as e:
        logger.error(f"Failed to load {path.name}: {e}")
        return None

    sandbox = opts["sandbox"] or conf.get("sandbox") or DEFAULT_SANDBOX
    if sandbox == "all":
        sandbox = DEFAULT_SANDBOX
    banner(sandbox, "api")

    try:
        token = authenticate(conf)
    except urllib.error.HTTPError as e:
        logger.error(f"IMS auth FAILED: HTTP {e.code} "
                     f"{e.read().decode(errors='replace')[:300]}")
        return None
    except Exception as e:
        logger.error(f"IMS auth FAILED: {type(e).__name__}: {e}")
        return None
    logger.info("IMS authenticated.")
    headers = aep_headers(token, conf, sandbox)

    logger.info(f"Listing audiences in sandbox '{sandbox}'...")
    audiences = fetch_audiences(headers)
    if not audiences:
        logger.warning("No audiences returned (empty sandbox, or no access).")
        return None

    methods = Counter(evaluation_method(a) for a in audiences)
    logger.info("Evaluation methods: " +
                ", ".join(f"{m}={n}" for m, n in methods.most_common()))

    if opts["all_methods"]:
        streaming = audiences
        logger.info(f"--all-methods: cataloguing all {len(streaming)} audience(s).")
    else:
        streaming = [a for a in audiences if evaluation_method(a) == "streaming"]
        logger.info(f"Filtering to {len(streaming)} STREAMING audience(s).")
    if not streaming:
        logger.warning("No streaming audiences found.")
        return None

    mp_map = fetch_merge_policies(headers)
    logger.info(f"  merge policies: {len(mp_map)}")

    dest_index = {}
    if opts["flows"]:
        logger.info("Mapping audiences -> destinations (Flow Service)...")
        dest_index, ok = build_destination_map(headers)
        if not ok:
            logger.warning("  no flows resolved -- 'Destination' columns will "
                           "be blank/N (re-run with access, or fill manually).")
    else:
        logger.info("--no-flows: skipping destination mapping.")

    rows = []
    estimate_on = opts["estimate"]
    n = len(streaming)
    for i, aud in enumerate(streaming, 1):
        aid = audience_id(aud)
        realized, existing, exited = audience_breakdown(aud)
        count = audience_count_from_obj(aud)
        if count is None:
            count = fetch_count_detail(headers, aid)
        if count is None and estimate_on:
            try:
                count = estimate_count(headers, aud)
            except urllib.error.HTTPError as e:
                logger.warning(f"  estimate endpoint failed (HTTP {e.code}); "
                               f"disabling --estimate for the rest, "
                               f"remaining missing counts -> {MISSING}.")
                estimate_on = False
        dests = destinations_for(dest_index, aid) if dest_index else []
        report_row = None
        rows.append(build_row(aud, count=count, realized=realized,
                              existing=existing, exited=exited, mp_map=mp_map,
                              dests=dests, report_row=report_row))
        if i % 10 == 0 or i == n:
            logger.info(f"  catalogued {i}/{n}")

    snapshot = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC (live API)")
    return rows, sandbox, snapshot


def run_files(opts) -> "tuple[list[dict], str, str] | None":
    banner(opts["sandbox"] or "files", "files")
    if not opts["list_json"]:
        logger.error("files mode needs --list-json=<audience list JSON export>.")
        return None
    list_path = Path(opts["list_json"])
    if not list_path.exists():
        logger.error(f"--list-json not found: {list_path}")
        return None
    audiences = load_list_json(list_path)
    logger.info(f"Loaded {len(audiences)} audience(s) from {list_path.name}")

    report_map = {}
    if opts["report"]:
        rpath = Path(opts["report"])
        if rpath.exists():
            report_map = load_report_xlsx(rpath)
        else:
            logger.warning(f"--report not found: {rpath} (continuing without it)")

    if opts["all_methods"]:
        streaming = audiences
    else:
        streaming = [a for a in audiences if evaluation_method(a) == "streaming"]
    logger.info(f"{len(streaming)} streaming audience(s) "
                f"(of {len(audiences)} total).")
    if not streaming:
        logger.warning("No streaming audiences found in the export. If the "
                       "export lacks evaluationInfo, re-run with --all-methods.")
        return None

    rows = []
    for aud in streaming:
        aid = audience_id(aud)
        realized, existing, exited = audience_breakdown(aud)
        count = audience_count_from_obj(aud)
        report_row = report_map.get(aid)
        # files mode has no Flow Service -- destinations are left for manual fill.
        rows.append(build_row(aud, count=count, realized=realized,
                              existing=existing, exited=exited, mp_map={},
                              dests=[], report_row=report_row))

    snapshot = "from local exports (point-in-time)"
    sandbox = opts["sandbox"] or "files"
    return rows, sandbox, snapshot


# ----------------------------------------------------------------------------
# Main
# ----------------------------------------------------------------------------
def main():
    opts = parse_args(sys.argv[1:])
    outdir = Path(opts["output"]) if opts["output"] else OUTPUT_DIR

    if opts["source"] == "files":
        result = run_files(opts)
    elif opts["source"] == "api":
        result = run_api(opts)
    else:
        logger.error(f"Unknown --source={opts['source']!r} (use api or files).")
        return

    if not result:
        return
    rows, sandbox, snapshot = result

    # Stable order: by suggested action (Keep, Refactor, dormant, test), then
    # largest audiences first within each group.
    order = {a: i for i, a in enumerate(SUGGESTED_ACTIONS)}

    def sort_key(r):
        c = r["Profile count"]
        c = c if isinstance(c, int) else -1
        return (order.get(r["_action"], 99), -c, r["Name"].lower())

    rows.sort(key=sort_key)

    summary = summarize(rows)
    print_console_summary(rows, summary)

    safe_sb = re.sub(r"[^0-9A-Za-z._-]+", "-", sandbox).strip("-") or "sandbox"
    stamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    xlsx_path = write_xlsx(rows, summary, safe_sb, stamp, snapshot, outdir)
    if xlsx_path:
        logger.info(f"XLSX written: {xlsx_path}")
    print()
    logger.info(f"Done. {summary['total']} streaming audience(s) catalogued.")


if __name__ == "__main__":
    main()
