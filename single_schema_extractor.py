#!/usr/bin/env python3
"""
single_schema_extractor.py  (AEP Swiss Army Knife)
==================================================
Single Schema and Data Extractor -- a stripped-back fork of data_dictionary_v3.

Where the full Data Dictionary sucks every schema out of a sandbox and writes a
tabbed workbook, this tool does ONE schema, end to end:

  1. Authenticate against IMS (pick a credential set from the keyring vault,
     or a plaintext creds/ folder where keyring is unavailable).
  2. Pick a sandbox (Enter = prod).
  3. PROMPT you to pick which schema to run (only schemas that have >=1 dataset,
     so there's real data to sample; type part of a name to narrow the list).
  4. Pull that schema's full field list (dot notation + type + friendly name +
     identity + relationship), dial into its dataset(s) and grab ONE real
     example value per field.
  5. Print a one-pager to the console AND write a single-sheet workbook:
        output/Schema Extract - <Client> - <Sandbox> - <YYYY-MM-DD HHMMSS>.xlsx
     (the time in the name means re-running never overwrites a prior export)

The output is marked STRICTLY CONFIDENTIAL -- it carries real sampled values.

This is deliberately simple. There is no per-schema filtering heuristic, no
coverage %, no top-5 tally, and no Profile Snapshot Export special-casing -- for
whole-profile coverage use data_dictionary_v3.py --data-dict=profile. Here the
example values are pooled from the schema's feeding datasets: enough to see what
each field looks like.

openpyxl is needed for the XLSX; pyarrow + tzdata for the example values (both
optional -- without pyarrow you still get the field list, just no examples).

Usage:
    python single_schema_extractor.py                     # interactive menus
    python single_schema_extractor.py "acme-beta"        # pick creds by service name
    python single_schema_extractor.py "acme-k" --sandbox=prod
    python single_schema_extractor.py "acme-k" --sandbox=prod --rows=100

Credentials come from the OS keyring (Windows Credential Manager) via
aep_creds; manage them with credential_validator_v2.py. Service names are the
keys you stored -- run `credential_validator_v2.py list` to see them.
"""

from __future__ import annotations

import json
import logging
import re
import ssl
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone
from pathlib import Path

import aep_creds  # keyring-backed credential store (replaces creds/*.json)

# ----------------------------------------------------------------------------
# Constants
# ----------------------------------------------------------------------------
SCRIPT_NAME    = "single_schema_extractor"
SCRIPT_VERSION = "1.1.0"
SCRIPT_DATE    = "2026-07-24"
SCRIPT_AUTHOR  = "Barry Mann (barrymann.com)"

SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"

IMS_URL = "https://ims-na1.adobelogin.com/ims/token"
PLATFORM = "https://platform.adobe.io"
SANDBOX_LIST_URL = f"{PLATFORM}/data/foundation/sandbox-management/sandboxes"
SCHEMAS_URL = f"{PLATFORM}/data/foundation/schemaregistry/tenant/schemas"
DESCRIPTORS_URL = f"{PLATFORM}/data/foundation/schemaregistry/tenant/descriptors"
DATASETS_URL = f"{PLATFORM}/data/foundation/catalog/dataSets"
CATALOG_BATCHES_URL = f"{PLATFORM}/data/foundation/catalog/batches"
EXPORT_URL = f"{PLATFORM}/data/foundation/export"   # Data Access (batch files)

# Default number of real records to sample when hunting for example values. A
# handful is plenty: we only want the first populated value per field.
DEFAULT_ROWS = 60

DEFAULT_SCOPES = (
    "openid,AdobeID,read_organizations,"
    "additional_info.projectedProductContext,session"
)

# Schema Registry media types.
XED_LIST = "application/vnd.adobe.xed+json"            # full(ish) resource list
XED_FULL = "application/vnd.adobe.xed-full+json; version=1"  # allOf resolved
DESCRIPTOR_ACCEPT = "application/vnd.adobe.xdm+json"

# Synthetic / load-test datasets pollute example values and carry huge files;
# skip them when sampling.
_SKIP_DS_RE = re.compile(r"perf[\s_-]?test|load[\s_-]?test", re.I)

KNOWN_CLASSES = {
    "https://ns.adobe.com/xdm/context/profile": "XDM Individual Profile",
    "https://ns.adobe.com/xdm/context/experienceevent": "XDM ExperienceEvent",
    "https://ns.adobe.com/xdm/data/adhoc": "Ad Hoc",
    "https://ns.adobe.com/xdm/data/record": "XDM Record",
    "https://ns.adobe.com/xdm/data/time-series": "XDM Time-series",
    "https://ns.adobe.com/xdm/classes/segment-definition": "Segment definition",
}

# ----------------------------------------------------------------------------
# ANSI / logging - matches the house style (data_dictionary_v3.py et al.)
# ----------------------------------------------------------------------------
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        h = kernel32.GetStdHandle(-11)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
            kernel32.SetConsoleMode(h, mode.value | 0x0004)
    except Exception:
        pass

ANSI = {
    "reset": "\033[0m", "bold": "\033[1m", "dim": "\033[2m",
    "red": "\033[31m", "green": "\033[32m", "yellow": "\033[33m",
    "blue": "\033[34m", "magenta": "\033[35m", "cyan": "\033[36m",
}
LEVEL_COLOR = {
    "DEBUG": ANSI["dim"], "INFO": ANSI["green"],
    "WARNING": ANSI["yellow"],
    "ERROR": ANSI["red"] + ANSI["bold"],
    "CRITICAL": ANSI["red"] + ANSI["bold"],
}


class ColoredFormatter(logging.Formatter):
    def format(self, record):
        color = LEVEL_COLOR.get(record.levelname, "")
        ts = self.formatTime(record, "%H:%M:%S")
        return (
            f"{ANSI['dim']}{ts}{ANSI['reset']} "
            f"{color}[{record.levelname:<7}]{ANSI['reset']} "
            f"{record.getMessage()}"
        )


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(ColoredFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger(SCRIPT_NAME)
SSL_CTX = ssl._create_unverified_context()


# ----------------------------------------------------------------------------
# HTTP / IMS / credential helpers  (shared house style)
# ----------------------------------------------------------------------------
def http(url, method="GET", headers=None, data=None, timeout=60):
    req = urllib.request.Request(url, data=data, headers=headers or {}, method=method)
    with urllib.request.urlopen(req, context=SSL_CTX, timeout=timeout) as r:
        return r.read(), dict(r.headers)


def flatten_err(text: str, limit: int = 200) -> str:
    return " ".join((text or "").split())[:limit]


def authenticate(conf):
    payload = urllib.parse.urlencode({
        "grant_type": "client_credentials",
        "client_id": conf["client_id"],
        "client_secret": conf["client_secret"],
        "scope": conf.get("scopes") or DEFAULT_SCOPES,
    }).encode("utf-8")
    body, _ = http(
        conf.get("oauth_url") or IMS_URL,
        method="POST",
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        data=payload,
    )
    return json.loads(body)


def aep_headers(token, conf, sandbox=None, accept="application/json"):
    headers = {
        "Authorization": f"Bearer {token}",
        "x-api-key": conf.get("api_key") or conf["client_id"],
        "x-gw-ims-org-id": conf["org_id"],
        "Accept": accept,
    }
    if sandbox:
        headers["x-sandbox-name"] = sandbox
    return headers


def list_sandboxes(token, conf):
    try:
        body, _ = http(SANDBOX_LIST_URL, headers=aep_headers(token, conf))
        data = json.loads(body)
        return True, data.get("sandboxes") or []
    except urllib.error.HTTPError as e:
        return False, f"HTTP {e.code}: {flatten_err(e.read().decode(errors='replace'))}"
    except Exception as e:
        return False, f"{type(e).__name__}: {e}"


# ----------------------------------------------------------------------------
# AEP reads -- schemas / datasets / descriptors
# ----------------------------------------------------------------------------
def _abs(href: str) -> str:
    if not href:
        return ""
    return href if href.startswith("http") else PLATFORM + href


def _next_url(data):
    if not isinstance(data, dict):
        return None
    nxt = (data.get("_links") or {}).get("next")
    if isinstance(nxt, str) and nxt:
        return _abs(nxt)
    if isinstance(nxt, dict) and nxt.get("href"):
        return _abs(nxt["href"])
    return None


def get_all_schemas(token, conf, sandbox):
    """Page through GET /tenant/schemas (xed+json). Returns raw schema dicts."""
    out = []
    url = f"{SCHEMAS_URL}?limit=300"
    headers = aep_headers(token, conf, sandbox, accept=XED_LIST)
    while url:
        body, _ = http(url, headers=headers)
        data = json.loads(body)
        results = data.get("results") if isinstance(data, dict) else data
        out.extend(r for r in (results or []) if isinstance(r, dict))
        url = _next_url(data)
    return out


def get_full_schema(token, conf, sandbox, ref):
    """GET one schema resolved (allOf merged) so it carries a flat properties
    tree. 'ref' is the meta:altId (preferred) or the $id."""
    url = f"{SCHEMAS_URL}/{urllib.parse.quote(ref, safe='')}"
    headers = aep_headers(token, conf, sandbox, accept=XED_FULL)
    body, _ = http(url, headers=headers)
    return json.loads(body)


def _pqs_table(ds: dict) -> str:
    """The dataset's AEP Query Service table name (tags['adobe/pqs/table']) --
    the SYSTEM name you SELECT ... FROM, distinct from the friendly name."""
    v = (ds.get("tags") or {}).get("adobe/pqs/table")
    if isinstance(v, list):
        return str(v[0]) if v else ""
    return str(v) if v else ""


def _dataset_tags_flat(ds: dict):
    """A dataset's Catalog tags flattened to lowercased match tokens: every tag
    KEY (e.g. 'bam') plus every 'key:value'. The UI 'TAGS' chips are tag keys."""
    tags = ds.get("tags") or {}
    toks = []
    if isinstance(tags, dict):
        for k, v in tags.items():
            toks.append(str(k).lower())
            for x in (v if isinstance(v, list) else [v]):
                toks.append(f"{k}:{x}".lower())
                toks.append(str(x).lower())
    return toks


def get_all_datasets(token, conf, sandbox):
    """Page through Catalog /dataSets. Returns a list of dataset dicts:
        {id, name, schema_id, table, tags(raw), tag_tokens(flattened)}."""
    out = []
    headers = aep_headers(token, conf, sandbox)
    start, limit = 0, 100
    while True:
        url = (f"{DATASETS_URL}?limit={limit}&start={start}"
               f"&properties=name,schemaRef,tags")
        body, _ = http(url, headers=headers)
        data = json.loads(body)
        if not isinstance(data, dict) or not data:
            break
        for dsid, ds in data.items():
            if not isinstance(ds, dict):
                continue
            out.append({
                "id": dsid,
                "name": ds.get("name") or dsid,
                "schema_id": (ds.get("schemaRef") or {}).get("id") or "",
                "table": _pqs_table(ds),
                "tags": ds.get("tags") or {},
                "tag_tokens": _dataset_tags_flat(ds),
            })
        if len(data) < limit:
            break
        start += limit
    return out


def derive_schema_maps(datasets):
    """From the flat dataset list, build the per-schema lookups:
        counts   {schema_id: dataset_count}
        ids      {schema_id: [datasetId, ...]}   (perf/load-test excluded)
        tables   {schema_id: [(friendly, table, id)]}"""
    counts, ids, tables = {}, {}, {}
    for d in datasets:
        ref = d["schema_id"]
        if not ref:
            continue
        counts[ref] = counts.get(ref, 0) + 1
        tables.setdefault(ref, []).append((d["name"], d["table"], d["id"]))
        if not _SKIP_DS_RE.search(d["name"]):
            ids.setdefault(ref, []).append(d["id"])
    for v in tables.values():
        v.sort(key=lambda t: t[0].lower())
    return counts, ids, tables


def get_all_descriptors(token, conf, sandbox):
    """GET every descriptor for the sandbox. Returns three dicts keyed by
    (schema_id, lowercased_dot_path): identities, relationships, labels."""
    identities, relationships, labels = {}, {}, {}
    headers = aep_headers(token, conf, sandbox, accept=DESCRIPTOR_ACCEPT)
    body, _ = http(DESCRIPTORS_URL, headers=headers)
    data = json.loads(body)

    descriptors = []
    if isinstance(data, dict):
        for v in data.values():
            if isinstance(v, list):
                descriptors.extend(x for x in v if isinstance(x, dict))
    elif isinstance(data, list):
        descriptors = [x for x in data if isinstance(x, dict)]

    for d in descriptors:
        dtype = d.get("@type") or ""
        sid = d.get("xdm:sourceSchema") or ""
        if not sid:
            continue

        if dtype == "xdm:descriptorPrimaryKey":
            raw_props = d.get("xdm:sourceProperty")
            props = raw_props if isinstance(raw_props, list) else [raw_props]
            for p in props:
                prop = _norm_prop(p)
                if not prop:
                    continue
                cur = identities.setdefault((sid, prop),
                                            {"primary": False, "namespace": ""})
                cur["primary"] = True
            continue

        prop = _norm_prop(d.get("xdm:sourceProperty"))
        if not prop:
            continue
        key = (sid, prop)

        if dtype == "xdm:descriptorIdentity":
            cur = identities.setdefault(key, {"primary": False, "namespace": ""})
            cur["primary"] = cur["primary"] or bool(d.get("xdm:isPrimary"))
            cur["namespace"] = (d.get("xdm:namespace")
                                or d.get("xdm:identityNamespace")
                                or cur["namespace"])
        elif dtype in ("xdm:descriptorRelationship", "xdm:descriptorOneToOne"):
            relationships[key] = {
                "dest_id": d.get("xdm:destinationSchema") or "",
                "dest_prop": _norm_prop(d.get("xdm:destinationProperty")),
                "cardinality": d.get("xdm:cardinality") or "1:1",
                "namespace": "",
            }
        elif dtype == "xdm:descriptorReferenceIdentity":
            relationships.setdefault(key, {
                "dest_id": "", "dest_prop": "", "cardinality": "ref",
                "namespace": (d.get("xdm:identityNamespace")
                              or d.get("xdm:namespace") or ""),
            })
        elif dtype == "xdm:alternateDisplayInfo":
            labels[key] = {
                "title": _localized(d.get("xdm:title")),
                "description": _localized(d.get("xdm:description")),
            }
    return identities, relationships, labels


def _norm_prop(p):
    """'/_tenant/person/name/firstName' -> '_tenant.person.name.firstname'."""
    if not isinstance(p, str):
        return ""
    return p.strip("/").replace("/", ".").lower()


def _localized(v):
    """alternateDisplayInfo titles/descriptions: plain string or locale map."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if isinstance(v, dict):
        for loc in ("en_us", "en_US", "en"):
            if loc in v:
                return _localized(v[loc])
        for val in v.values():
            return _localized(val)
    return str(v)


# ----------------------------------------------------------------------------
# Field flattening
# ----------------------------------------------------------------------------
def flatten_fields(node, prefix="", depth=0):
    """Walk a resolved xed-full properties tree to leaf dot-notation paths.
    Returns [(path, data_type, title, required_bool), ...]."""
    rows = []
    if depth > 18:
        return rows
    props = node.get("properties")
    if not isinstance(props, dict):
        return rows
    required = set(node.get("required") or [])
    for key, sub in props.items():
        if not isinstance(sub, dict):
            continue
        path = f"{prefix}.{key}" if prefix else key
        req = key in required
        title = sub.get("title") or ""
        t = sub.get("type")
        if t == "object" and isinstance(sub.get("properties"), dict):
            rows.extend(flatten_fields(sub, path, depth + 1))
        elif t == "array":
            items = sub.get("items") or {}
            if (isinstance(items, dict) and items.get("type") == "object"
                    and isinstance(items.get("properties"), dict)):
                rows.extend(flatten_fields(items, path + "[]", depth + 1))
            else:
                itype = (items.get("meta:xdmType") or items.get("format")
                         or items.get("type") or "any")
                rows.append((path + "[]", f"array<{itype}>", title, req))
        else:
            disp = sub.get("meta:xdmType") or sub.get("format") or t or "object"
            rows.append((path, disp, title, req))
    return rows


# ----------------------------------------------------------------------------
# Sampling -- pull a few real records and grab one example value per field
# ----------------------------------------------------------------------------
def _batch_record_count(b):
    m = b.get("metrics") or {}
    for v in (m.get("outputRecordCount"), m.get("inputRecordCount"),
              b.get("recordCount"), m.get("recordCount")):
        if v is not None:
            try:
                return int(v)
            except (TypeError, ValueError):
                return -1
    return -1


def _read_parquet_rows(raw_bytes, limit):
    """Parse only up to `limit` rows (row-group streaming)."""
    import pyarrow as pa
    import pyarrow.parquet as pq
    pf = pq.ParquetFile(pa.BufferReader(raw_bytes))
    out = []
    for batch in pf.iter_batches(batch_size=max(1, min(limit, 1024))):
        out.extend(batch.to_pylist())
        if len(out) >= limit:
            break
    return out[:limit]


def _download_batch_rows(headers, bid, need, timeout=75):
    """Download a batch's Parquet file(s), return up to `need` row dicts."""
    files = json.loads(http(f"{EXPORT_URL}/batches/{bid}/files?limit=10",
                            headers=headers)[0])
    entries = (files.get("data") if isinstance(files, dict) else files) or []
    rows = []
    for fe in entries:
        if len(rows) >= need:
            break
        fid = fe.get("dataSetFileId")
        if not fid:
            continue
        meta = json.loads(http(f"{EXPORT_URL}/files/{fid}", headers=headers)[0])
        for phys in (meta.get("data") if isinstance(meta, dict) else meta) or []:
            name = phys.get("name") or ""
            if not name.endswith(".parquet"):
                continue
            raw = http(f"{EXPORT_URL}/files/{fid}?path={urllib.parse.quote(name)}",
                       headers=headers, timeout=timeout)[0]
            rows.extend(_read_parquet_rows(raw, need - len(rows)))
            if len(rows) >= need:
                break
    return rows[:need]


def sample_rows(token, conf, sandbox, dsids, target=DEFAULT_ROWS, max_batches=6):
    """Pull up to `target` real records from a schema's datasets (newest
    non-empty SUCCESS batch first). Returns a list of nested row dicts."""
    headers = aep_headers(token, conf, sandbox)
    rows = []
    for di, dsid in enumerate(dsids, 1):
        if len(rows) >= target:
            break
        try:
            url = f"{CATALOG_BATCHES_URL}?dataSet={dsid}&limit=20&orderBy=desc:created"
            batches = json.loads(http(url, headers=headers)[0])
        except Exception as e:
            logger.warning(f"  dataset {di}/{len(dsids)} {dsid}: batch list "
                           f"failed ({e}); skipping.")
            continue
        if not isinstance(batches, dict):
            continue
        succ_items = [(bid, b) for bid, b in batches.items()
                      if isinstance(b, dict) and b.get("status") == "success"
                      and _batch_record_count(b) != 0]
        succ_items.sort(key=lambda kv: _batch_record_count(kv[1]), reverse=True)
        seen = 0
        for bid, _ in succ_items:
            if len(rows) >= target or seen >= max_batches:
                break
            seen += 1
            try:
                chunk = _download_batch_rows(headers, bid, target - len(rows))
            except Exception as e:
                logger.warning(f"    batch {bid[:24]}: skipped ({e})")
                continue
            if chunk:
                rows.extend(chunk)
                logger.info(f"  dataset {di}/{len(dsids)}: batch {bid[:24]} -> "
                            f"{len(chunk)} rows {ANSI['dim']}(total "
                            f"{len(rows)}/{target}){ANSI['reset']}")
    return rows[:target]


def _extract_values(row, path):
    """All scalar/array leaf values at a dot-notation path."""
    cur = [row]
    for seg in path.split("."):
        arr = seg.endswith("[]")
        key = seg[:-2] if arr else seg
        nxt = []
        for c in cur:
            if not isinstance(c, dict):
                continue
            v = c.get(key)
            if v is None:
                continue
            if arr:
                if isinstance(v, list):
                    nxt.extend(x for x in v if x is not None)
            else:
                nxt.append(v)
        cur = nxt
        if not cur:
            break
    return cur


def _fmt_example(v, limit=80):
    return str(v).replace("\n", " ").replace("\r", " ").strip()[:limit]


def build_examples(rows, fields):
    """{field_path: first_example_value} -- the first populated scalar value
    found across the sampled rows (blank if the field is never populated)."""
    out = {}
    for field in fields:
        fpath = field[0]
        example = ""
        for r in rows:
            scal = [v for v in _extract_values(r, fpath)
                    if isinstance(v, (str, int, float, bool))]
            if scal:
                example = _fmt_example(scal[0])
                break
        out[fpath] = example
    return out


# ----------------------------------------------------------------------------
# Schema metadata helpers
# ----------------------------------------------------------------------------
def class_name(class_id: str) -> str:
    if not class_id:
        return "?"
    if class_id in KNOWN_CLASSES:
        return KNOWN_CLASSES[class_id]
    tail = class_id.rstrip("/").rsplit("/", 1)[-1]
    return tail.replace("-", " ").replace("_", " ").title()


def fmt_epoch_ms(ms) -> str:
    try:
        return datetime.fromtimestamp(int(ms) / 1000, tz=timezone.utc).strftime(
            "%Y-%m-%d %H:%M")
    except Exception:
        return ""


def last_modified(raw: dict) -> str:
    meta = raw.get("meta:registryMetadata") or {}
    for k in ("repo:lastModifiedDate", "repo:lastModified", "repo:updatedDate"):
        if meta.get(k):
            return fmt_epoch_ms(meta[k])
    return ""


def _id_tail(sid: str) -> str:
    return sid.rsplit("/", 1)[-1] if sid else ""


# ----------------------------------------------------------------------------
# Field-row builder -- join fields to their descriptor metadata (one schema)
# ----------------------------------------------------------------------------
def build_field_rows(sid, fields, identities, relationships, labels, id_to_title):
    """[(path, type, friendly, required, identity, relationship), ...] for one
    schema, joining the flattened fields to identity/relationship/label
    descriptors -- the same join data_dictionary_v3 does per schema."""
    field_rows = []
    n_identities = n_rels = n_labels = 0
    for path, dtype, title, req in fields:
        mkey = (sid, path.replace("[]", "").lower())
        ident = identities.get(mkey)
        rel = relationships.get(mkey)
        lab = labels.get(mkey)
        if ident:
            n_identities += 1
            id_disp = "PRIMARY" if ident["primary"] else "identity"
            if ident["namespace"]:
                id_disp += f" ({ident['namespace']})"
        else:
            id_disp = ""
        if rel:
            n_rels += 1
            if rel["dest_id"]:
                dest = id_to_title.get(rel["dest_id"]) or _id_tail(rel["dest_id"])
                rel_disp = dest + (f".{rel['dest_prop']}" if rel["dest_prop"] else "")
                if rel["cardinality"]:
                    rel_disp += f"  [{rel['cardinality']}]"
            elif rel.get("namespace"):
                rel_disp = f"ref-identity ({rel['namespace']})"
            else:
                rel_disp = "ref-identity"
        else:
            rel_disp = ""
        friendly = title or (lab["title"] if lab else "")
        if friendly:
            n_labels += 1
        field_rows.append((path, dtype, friendly, "Y" if req else "",
                           id_disp, rel_disp))
    return field_rows, n_identities, n_rels, n_labels


# ----------------------------------------------------------------------------
# Credential menu + sandbox picker  (shared house style)
# ----------------------------------------------------------------------------
def pretty_name(stem: str) -> str:
    return stem.replace("-", " ").replace("_", " ").title()


def client_label(conf: dict, stem: str) -> str:
    c = (conf.get("client") or "").strip()
    if not c:
        parts = stem.split()
        c = " ".join(parts[:-1]) if len(parts) > 1 else stem
    return c.replace("-", " ").replace("_", " ").strip().title() or "Client"


def short_type(raw: str) -> str:
    t = (raw or "").lower()
    if "prod" in t:
        return "prod"
    if "dev" in t:
        return "dev"
    return raw or "?"


def menu(services):
    """House-style numbered picker over keyring service names."""
    print()
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Credential bank{ANSI['reset']}  "
          f"{ANSI['dim']}(OS keyring vault){ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 60 + ANSI["reset"])
    for i, name in enumerate(services, 1):
        print(f"  {ANSI['bold']}[{i}]{ANSI['reset']} "
              f"{ANSI['yellow']}{pretty_name(name):<20}{ANSI['reset']} "
              f"{ANSI['dim']}{name}{ANSI['reset']}")
    print(bar)
    raw = input(
        f"\nChoose a credential set by number "
        f"({ANSI['cyan']}1{ANSI['reset']}-{ANSI['cyan']}{len(services)}{ANSI['reset']}), "
        "blank to quit: "
    ).strip()
    if not raw or not raw.isdigit():
        return None
    idx = int(raw)
    if 1 <= idx <= len(services):
        return services[idx - 1]
    logger.warning(f"Choice {idx} out of range.")
    return None


def _default_prod(sandboxes):
    for sb in sandboxes:
        if sb.get("name") == "prod":
            return sb
    for sb in sandboxes:
        if short_type(sb.get("type", "")) == "prod":
            return sb
    return sandboxes[0] if sandboxes else None


def pick_sandbox(sandboxes):
    """Prompt for ONE sandbox to read. Enter = prod (default)."""
    default = _default_prod(sandboxes)
    print()
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Sandboxes visible to this credential{ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 60 + ANSI["reset"])
    for i, sb in enumerate(sandboxes, 1):
        name = sb.get("name", "?")
        title = sb.get("title") or name
        env = short_type(sb.get("type", ""))
        star = " (default)" if sb is default else ""
        print(f"  {ANSI['bold']}[{i:>2}]{ANSI['reset']} "
              f"{ANSI['yellow']}{title:<22}{ANSI['reset']} "
              f"{ANSI['dim']}{name:<16}{ANSI['reset']} {env}{star}")
    print(bar)
    dlabel = default.get("name", "?") if default else "?"
    raw = input(
        f"\nPick a sandbox ({ANSI['cyan']}1{ANSI['reset']}, or Enter for "
        f"{ANSI['green']}{dlabel}{ANSI['reset']}): "
    ).strip()
    if not raw:
        return default
    if raw.isdigit() and 1 <= int(raw) <= len(sandboxes):
        return sandboxes[int(raw) - 1]
    logger.warning(f"Invalid choice {raw!r}; using {dlabel}.")
    return default


def resolve_sandbox_arg(arg: str, sandboxes):
    """Map a --sandbox=<name> value to a single sandbox dict (first match)."""
    wanted = [t for t in arg.replace(",", " ").split() if t]
    by_name = {sb.get("name"): sb for sb in sandboxes}
    for w in wanted:
        if by_name.get(w):
            return by_name[w]
        logger.warning(f"--sandbox: no sandbox named {w!r}; ignoring.")
    return None


def resolve_sandboxes_arg(arg: str, sandboxes):
    """Map a --sandbox=<names|all> value to a LIST of sandbox dicts. Each token
    matches a sandbox by exact name first, then by substring of name/title."""
    tokens = [t for t in arg.replace(",", " ").split() if t]
    if any(t.lower() == "all" for t in tokens):
        return list(sandboxes)
    out = []
    for t in tokens:
        tl = t.lower()
        m = next((sb for sb in sandboxes
                  if (sb.get("name") or "").lower() == tl), None)
        if not m:
            m = next((sb for sb in sandboxes
                      if tl in (sb.get("name") or "").lower()
                      or tl in (sb.get("title") or "").lower()), None)
        if m and m not in out:
            out.append(m)
        elif not m:
            logger.warning(f"--sandbox: no sandbox matches {t!r}; ignoring.")
    return out


def select_datasets_by_name(datasets, names):
    """Datasets whose name exactly matches (case-insensitive) one of `names`.
    Returns [(dataset, matched_name), ...]. Exact match avoids catching '... -
    Copy' duplicates of a wanted dataset."""
    wanted = {n.strip().lower(): n for n in names if n.strip()}
    out = []
    for d in datasets:
        key = (d["name"] or "").strip().lower()
        if key in wanted:
            out.append((d, wanted[key]))
    return out


# ----------------------------------------------------------------------------
# Schema picker -- the heart of this tool
# ----------------------------------------------------------------------------
def pick_schema(raws, counts):
    """Prompt the user for ONE schema to run. Only schemas with >=1 dataset are
    offered (so there's real data to sample); type part of a name to narrow the
    list. Returns the chosen raw schema dict, or None."""
    pickable = [r for r in raws if counts.get(r.get("$id") or "", 0) >= 1]
    pickable.sort(key=lambda r: (r.get("title") or "").lower())
    if not pickable:
        logger.warning("No schemas with datasets found in this sandbox.")
        return None

    print()
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}{len(pickable)} schema(s) with data in this sandbox"
          f"{ANSI['reset']}")
    print(ANSI["cyan"] + "-" * 60 + ANSI["reset"])
    flt = input("  Type part of a schema name to filter (Enter for all): ").strip().lower()
    matches = [r for r in pickable if flt in (r.get("title") or "").lower()] if flt else pickable
    if not matches:
        logger.warning(f"No schema title matches {flt!r}.")
        return None

    print()
    for i, r in enumerate(matches, 1):
        title = r.get("title") or r.get("meta:altId") or r.get("$id") or "?"
        cls = class_name(r.get("meta:class") or "")
        n_ds = counts.get(r.get("$id") or "", 0)
        print(f"  {ANSI['bold']}[{i:>3}]{ANSI['reset']} "
              f"{ANSI['yellow']}{title[:46]:<46}{ANSI['reset']} "
              f"{ANSI['dim']}{cls[:20]:<20}{ANSI['reset']} "
              f"{n_ds} dataset(s)")
    print(bar)
    raw = input(f"\n  Pick a schema by number "
                f"({ANSI['cyan']}1{ANSI['reset']}-"
                f"{ANSI['cyan']}{len(matches)}{ANSI['reset']}): ").strip()
    if raw.isdigit() and 1 <= int(raw) <= len(matches):
        return matches[int(raw) - 1]
    logger.warning("No valid schema chosen.")
    return None


# ----------------------------------------------------------------------------
# Console one-pager
# ----------------------------------------------------------------------------
def print_one_pager(info):
    bar = ANSI["cyan"] + "=" * 100 + ANSI["reset"]
    print()
    print(bar)
    print(f"  {ANSI['bold']}{info['title']}{ANSI['reset']}  "
          f"{ANSI['dim']}({info['class']}){ANSI['reset']}")
    print(f"  {ANSI['dim']}{info['datasets']} dataset(s)  |  "
          f"{info['n_fields']} fields  |  {info['n_identities']} identity  |  "
          f"{info['n_relationships']} relationship  |  sandbox: {info['sandbox']}"
          f"{ANSI['reset']}")
    if info.get("tables"):
        print(f"  {ANSI['yellow']}SQL table(s): {info['tables']}{ANSI['reset']}")
    if info.get("sampled") is not None:
        print(f"  {ANSI['dim']}example values sampled from {info['sampled']} "
              f"real record(s){ANSI['reset']}")
    print(bar)

    header = (f"{'Field (dot notation)':<46} {'Type':<16} {'Friendly':<22} "
              f"{'Identity':<16} Example")
    print(ANSI["bold"] + header + ANSI["reset"])
    print(ANSI["dim"] + "-" * 130 + ANSI["reset"])
    examples = info.get("examples") or {}
    for (fpath, dtype, friendly, req, ident, rel) in info["fields"]:
        ex = examples.get(fpath, "")
        print(f"{fpath[:46]:<46} {ANSI['dim']}{dtype[:16]:<16}{ANSI['reset']} "
              f"{friendly[:22]:<22} {ANSI['cyan']}{ident[:16]:<16}{ANSI['reset']} "
              f"{ex[:40]}")
    print(ANSI["dim"] + "-" * 130 + ANSI["reset"])


# ----------------------------------------------------------------------------
# XLSX one-pager
# ----------------------------------------------------------------------------
CONFIDENTIAL = "STRICTLY CONFIDENTIAL"
FIELD_COLUMNS = ["Field (dot notation)", "Data Type", "Friendly Name",
                 "Required", "Identity", "Relationship -> target",
                 "Example value"]
_HEADER_BG = "1F4E78"


def print_compact(info):
    """One-line result per dataset -- for a batch run where the full field dump
    lives in the workbook, not the console."""
    populated = sum(1 for v in (info.get("examples") or {}).values() if v)
    if info.get("sampled") is None:
        smp = " (not sampled)"
    else:
        smp = f", {populated}/{info['n_fields']} populated from {info['sampled']} rec(s)"
    tbl = f"  table {info['tables']}" if info["tables"] else ""
    print(f"  {ANSI['green']}OK{ANSI['reset']} "
          f"{ANSI['bold']}{info.get('dataset_name') or info['title']}{ANSI['reset']} "
          f"{ANSI['dim']}[{info['sandbox']}]{ANSI['reset']}  "
          f"{info['schema_title']} - {info['n_fields']} fields{smp}{tbl}")


def _safe_sheet_name(name: str, used: set) -> str:
    s = re.sub(r"[\[\]\:\*\?\/\\]", "-", name).strip()[:31] or "sheet"
    base, i = s, 2
    while s.lower() in used:
        suffix = f"-{i}"
        s = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(s.lower())
    return s


def write_workbook(infos, client: str, datestr: str, label: str = ""):
    """Write ALL extracted datasets into ONE workbook: a Summary tab plus one
    field tab per dataset. Returns the path (or None if openpyxl is missing)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed -- skipping XLSX "
                       "(pip install openpyxl).")
        return None

    OUTPUT_DIR.mkdir(exist_ok=True)
    safe_client = re.sub(r"[^0-9A-Za-z _-]+", "", client).strip() or "Client"
    suffix = f" - {label}" if label else ""
    path = OUTPUT_DIR / f"Schema Extract - {safe_client}{suffix} - {datestr}.xlsx"

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor=_HEADER_BG)
    title_font = Font(bold=True, size=14)
    conf_font = Font(bold=True, size=11, color="C00000")
    center = Alignment(horizontal="center")

    def confidential(ws):
        ws["A1"] = CONFIDENTIAL
        ws["A1"].font = conf_font
        ws.oddHeader.center.text = f'&"-,Bold"&12&KC00000{CONFIDENTIAL}'
        ws.evenHeader.center.text = ws.oddHeader.center.text

    wb = Workbook()

    # Name each dataset's tab up front so the Summary can point at it.
    used = {"summary"}
    tabbed = [(info, _safe_sheet_name(info.get("dataset_name") or info["title"],
                                      used)) for info in infos]

    # ---- Summary tab --------------------------------------------------------
    ws = wb.active
    ws.title = "Summary"
    confidential(ws)
    ws["A2"] = f"Single Schema and Data Extract  -  {client}"
    ws["A2"].font = title_font
    ws["A3"] = f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}"
    ws["A3"].font = Font(italic=True, color="666666")
    ws["A4"] = ("One tab per dataset: its schema fields (dot notation, type, "
                "friendly name, identity, relationship) and one real example "
                "value per field.")
    ws["A4"].font = Font(italic=True, color="666666")
    sum_cols = ["Tab", "Dataset", "Schema", "Class", "Sandbox", "Fields",
                "Populated", "Sampled rows", "SQL table name", "Schema $id"]
    hr = 6
    for c, nm in enumerate(sum_cols, 1):
        ws.cell(hr, c, nm)
    for c in range(1, len(sum_cols) + 1):
        ws.cell(hr, c).font = head_font
        ws.cell(hr, c).fill = head_fill
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    rr = hr + 1
    for info, tab in tabbed:
        populated = sum(1 for v in (info.get("examples") or {}).values() if v)
        row = [tab, info.get("dataset_name") or info["title"],
               info["schema_title"], info["class"], info["sandbox"],
               info["n_fields"], populated,
               "" if info.get("sampled") is None else info["sampled"],
               info["tables"], info["id"]]
        for c, val in enumerate(row, 1):
            ws.cell(rr, c, val)
        rr += 1
    for idx, w in enumerate([26, 34, 30, 22, 16, 7, 10, 12, 40, 50], 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # ---- One tab per dataset ------------------------------------------------
    for info, tab in tabbed:
        sheet = wb.create_sheet(tab)
        confidential(sheet)
        sheet["A2"] = info.get("dataset_name") or info["title"]
        sheet["A2"].font = title_font
        a3 = (f'schema: {info["schema_title"]}   |   {info["class"]}   |   '
              f'{info["n_fields"]} fields   |   {info["n_identities"]} identity'
              f'   |   {info["n_relationships"]} relationship   |   '
              f'sandbox: {info["sandbox"]}')
        if info.get("sampled") is not None:
            a3 += f'   |   examples sampled from {info["sampled"]} record(s)'
        sheet["A3"] = a3
        sheet["A3"].font = Font(italic=True, color="666666")
        sheet["A4"] = info["id"]
        sheet["A4"].font = Font(italic=True, color="999999", size=9)
        if info.get("tables"):
            sheet["A5"] = f"SQL table name(s):  {info['tables']}"
            sheet["A5"].font = Font(italic=True, bold=True, color="C55A11")

        hr = 6
        for c, nm in enumerate(FIELD_COLUMNS, 1):
            sheet.cell(hr, c, nm)
        for c in range(1, len(FIELD_COLUMNS) + 1):
            sheet.cell(hr, c).font = head_font
            sheet.cell(hr, c).fill = head_fill
        sheet.freeze_panes = sheet.cell(row=hr + 1, column=1)

        examples = info.get("examples") or {}
        ridx = hr + 1
        for (fpath, dtype, friendly, req, ident, rel) in info["fields"]:
            sheet.cell(ridx, 1, fpath)
            sheet.cell(ridx, 2, dtype)
            sheet.cell(ridx, 3, friendly)
            sheet.cell(ridx, 4, req).alignment = center
            sheet.cell(ridx, 5, ident)
            sheet.cell(ridx, 6, rel)
            sheet.cell(ridx, 7, examples.get(fpath, ""))
            ridx += 1
        for idx, w in enumerate([50, 20, 28, 9, 22, 42, 46], 1):
            sheet.column_dimensions[get_column_letter(idx)].width = w

    wb.save(path)
    return path


# ----------------------------------------------------------------------------
# Driver
# ----------------------------------------------------------------------------
def _have_pyarrow():
    try:
        import pyarrow  # noqa: F401
        return True
    except ImportError:
        return False


def build_info(token, conf, sandbox, sandbox_title, *, title, schema_title,
               schema_id, schema_raw, dsids_to_sample, tables_list,
               datasets_count, descriptors, id_to_title, rows, have_pyarrow,
               dataset_name=None):
    """Resolve one schema's fields + descriptor joins, sample its dataset(s) for
    one example value per field, and return the info dict the writers consume."""
    identities, relationships, labels = descriptors
    ref = (schema_raw.get("meta:altId") if schema_raw else None) or schema_id
    try:
        full = get_full_schema(token, conf, sandbox, ref)
    except Exception as e:
        logger.warning(f"    full-schema fetch failed for {schema_title!r} "
                       f"({e}); field list will be empty.")
        full = {}
    fields = flatten_fields(full)
    field_rows, n_id, n_rel, n_lab = build_field_rows(
        schema_id, fields, identities, relationships, labels, id_to_title)

    examples, sampled = {}, None
    if have_pyarrow and dsids_to_sample:
        logger.info(f"    sampling up to {rows} record(s) for example values ...")
        sample = sample_rows(token, conf, sandbox, dsids_to_sample, target=rows)
        sampled = len(sample)
        if sample:
            examples = build_examples(sample, fields)
        else:
            logger.warning("    no records sampled (empty or unreadable batches); "
                           "field list still produced.")

    tables_str = ", ".join(t for _, t, _ in tables_list if t)
    return {
        "title": title,
        "schema_title": schema_title,
        "class": class_name((schema_raw or {}).get("meta:class") or ""),
        "id": schema_id,
        "sandbox": sandbox_title,
        "sandbox_name": sandbox,
        "dataset_name": dataset_name,
        "last_mod": last_modified(schema_raw or {}),
        "datasets": datasets_count,
        "tables": tables_str,
        "fields": field_rows,
        "n_fields": len(field_rows),
        "n_identities": n_id,
        "n_relationships": n_rel,
        "examples": examples,
        "sampled": sampled,
    }


def run_datasets(token, conf, sandboxes, names, rows, client, list_only=False):
    """Find the named datasets across the given sandbox(es), extract each one's
    schema + example values, and write them all into ONE workbook."""
    have_pyarrow = _have_pyarrow()
    if not have_pyarrow and not list_only:
        logger.warning("pyarrow not installed -- field lists only, no example "
                       "values (pip install pyarrow).")

    infos = []
    found_names = set()
    for sb in sandboxes:
        sandbox = sb.get("name", "?")
        sandbox_title = sb.get("title") or sandbox
        try:
            datasets = get_all_datasets(token, conf, sandbox)
        except Exception as e:
            logger.warning(f"[{sandbox}] dataset list failed ({e}); skipping.")
            continue
        matches = select_datasets_by_name(datasets, names)
        if not matches:
            logger.info(f"[{sandbox}] none of the {len(names)} named dataset(s) "
                        f"here.")
            continue
        logger.info(f"[{sandbox_title}] {ANSI['bold']}{len(matches)} match(es)"
                    f"{ANSI['reset']}: "
                    f"{', '.join(d['name'] for d, _ in matches)}")
        if list_only:
            for d, _ in matches:
                found_names.add(d["name"].strip().lower())
                print(f"  {ANSI['green']}FOUND{ANSI['reset']} {d['name']}  "
                      f"{ANSI['dim']}[{sandbox_title}]  schema={d['schema_id']}  "
                      f"id={d['id']}{ANSI['reset']}")
            continue

        # Only now (matches exist) pay for the schema list + descriptors.
        raws = get_all_schemas(token, conf, sandbox)
        id_to_title = {r.get("$id"): (r.get("title") or r.get("$id"))
                       for r in raws}
        id_to_raw = {r.get("$id"): r for r in raws}
        descriptors = get_all_descriptors(token, conf, sandbox)
        for d, _ in matches:
            found_names.add(d["name"].strip().lower())
            sid = d["schema_id"]
            schema_raw = id_to_raw.get(sid) or {}
            schema_title = (id_to_title.get(sid)
                            or schema_raw.get("title") or _id_tail(sid) or "?")
            logger.info(f"  {ANSI['bold']}{d['name']}{ANSI['reset']} "
                        f"({schema_title}) ...")
            info = build_info(
                token, conf, sandbox, sandbox_title,
                title=d["name"], schema_title=schema_title, schema_id=sid,
                schema_raw=schema_raw, dsids_to_sample=[d["id"]],
                tables_list=[(d["name"], d["table"], d["id"])],
                datasets_count=1, descriptors=descriptors,
                id_to_title=id_to_title, rows=rows, have_pyarrow=have_pyarrow,
                dataset_name=d["name"])
            print_compact(info)
            infos.append(info)

    missing = [n for n in names if n.strip().lower() not in found_names]
    if missing:
        logger.warning(f"Not found in any sandbox read: {', '.join(missing)}")
    if list_only:
        logger.info("List-only: no data sampled, no workbook written.")
        return
    if not infos:
        logger.warning("No datasets extracted; nothing to write.")
        return

    datestr = datetime.now(timezone.utc).strftime("%Y-%m-%d %H%M%S")
    label = infos[0]["sandbox"] if len({i["sandbox"] for i in infos}) == 1 else ""
    xlsx_path = write_workbook(infos, client, datestr, label=label)
    if xlsx_path:
        logger.info(f"{ANSI['bold']}XLSX written{ANSI['reset']}: {xlsx_path}  "
                    f"({len(infos)} dataset tab(s) + Summary)")


def run_schema_interactive(token, conf, sandbox_arg, rows, client):
    """Original behaviour: pick ONE schema in ONE sandbox, write a one-tab book."""
    ok, result = list_sandboxes(token, conf)
    if not ok or not result:
        logger.error("No sandboxes available.")
        return
    sb = (resolve_sandbox_arg(sandbox_arg, result) if sandbox_arg
          else pick_sandbox(result))
    if not sb:
        logger.info("No sandbox chosen. Exiting.")
        return
    sandbox = sb.get("name", "?")
    sandbox_title = sb.get("title") or sandbox
    logger.info(f"Reading sandbox {sandbox!r} ...")
    raws = get_all_schemas(token, conf, sandbox)
    datasets = get_all_datasets(token, conf, sandbox)
    counts, ids_by_schema, tables_by_schema = derive_schema_maps(datasets)
    logger.info(f"{len(raws)} schema(s), {len(datasets)} dataset(s) seen.")

    chosen = pick_schema(raws, counts)
    if not chosen:
        return
    sid = chosen.get("$id") or ""
    stitle = chosen.get("title") or chosen.get("meta:altId") or sid or "?"
    id_to_title = {r.get("$id"): (r.get("title") or r.get("$id")) for r in raws}
    descriptors = get_all_descriptors(token, conf, sandbox)
    info = build_info(
        token, conf, sandbox, sandbox_title,
        title=stitle, schema_title=stitle, schema_id=sid, schema_raw=chosen,
        dsids_to_sample=ids_by_schema.get(sid, []),
        tables_list=tables_by_schema.get(sid, []),
        datasets_count=counts.get(sid, 0), descriptors=descriptors,
        id_to_title=id_to_title, rows=rows, have_pyarrow=_have_pyarrow())
    print_one_pager(info)
    datestr = datetime.now(timezone.utc).strftime("%Y-%m-%d %H%M%S")
    xlsx_path = write_workbook([info], client, datestr, label=sandbox_title)
    if xlsx_path:
        logger.info(f"XLSX written: {xlsx_path}")
    logger.info("Done.")


def run(service: str, sandbox_arg: str | None, rows: int = DEFAULT_ROWS,
        dataset_names=None, list_only=False):
    bar = ANSI["cyan"] + "=" * 60 + ANSI["reset"]
    print()
    print(bar)
    print(f"  {ANSI['bold']}Single Schema and Data Extractor for "
          f"{ANSI['yellow']}{pretty_name(service)}{ANSI['reset']}  "
          f"{ANSI['dim']}({service}){ANSI['reset']}")
    print(bar)

    try:
        conf = aep_creds.load_creds(service)
    except aep_creds.CredsError as e:
        logger.error(f"Failed to load credentials for {service!r}: {e}")
        return

    try:
        resp = authenticate(conf)
    except urllib.error.HTTPError as e:
        logger.error(f"IMS auth FAILED: HTTP {e.code} "
                     f"{e.read().decode(errors='replace')[:300]}")
        return
    except Exception as e:
        logger.error(f"IMS auth FAILED: {type(e).__name__}: {e}")
        return
    token = resp["access_token"]
    logger.info(f"IMS authenticated (org {conf['org_id']}).")
    client = client_label(conf, service)

    if dataset_names:
        ok, result = list_sandboxes(token, conf)
        if not ok or not result:
            logger.error(f"GET /sandboxes failed or none visible: {result}")
            return
        if sandbox_arg:
            sandboxes = resolve_sandboxes_arg(sandbox_arg, result)
        else:
            # No sandbox given -> scan every visible sandbox to locate them.
            sandboxes = list(result)
            logger.info(f"No --sandbox given; scanning all {len(sandboxes)} "
                        f"visible sandbox(es) for the named dataset(s).")
        if not sandboxes:
            logger.info("No sandbox to read. Exiting.")
            return
        run_datasets(token, conf, sandboxes, dataset_names, rows, client,
                     list_only=list_only)
    else:
        run_schema_interactive(token, conf, sandbox_arg, rows, client)


def print_banner() -> None:
    bar = ANSI["cyan"] + "=" * 72 + ANSI["reset"]
    print(bar)
    print(f"  {ANSI['bold']}Single Schema and Data Extractor "
          f"v{SCRIPT_VERSION}{ANSI['reset']}   ({SCRIPT_DATE})")
    print(f"  by {SCRIPT_AUTHOR}")
    print(f"  {ANSI['dim']}Pulls ONE AEP schema -- fields, friendly names, and a "
          f"real example value per field.{ANSI['reset']}")
    print(bar)


def main():
    print_banner()
    args = sys.argv[1:]
    sandbox_arg = None
    rows = DEFAULT_ROWS
    dataset_names = []
    list_only = False
    positional = []
    for a in args:
        if a.startswith("--sandbox="):
            sandbox_arg = a.split("=", 1)[1]
        elif a.startswith("--dataset=") or a.startswith("--datasets="):
            # Semicolon-separated exact dataset names (names contain commas/
            # spaces). Repeatable. Each is matched case-insensitively by name.
            val = a.split("=", 1)[1]
            dataset_names += [n.strip() for n in val.split(";") if n.strip()]
        elif a == "--list":
            list_only = True
        elif a.startswith("--rows="):
            try:
                rows = max(1, int(a.split("=", 1)[1]))
            except ValueError:
                logger.warning(f"Ignoring bad --rows value in {a!r}.")
        elif a in ("--debug", "-v"):
            logger.setLevel(logging.DEBUG)
        elif a.startswith("-"):
            continue
        else:
            positional.append(a)

    services = aep_creds.list_services()
    if not services:
        logger.error("No credentials found in the keyring vault or the creds/ "
                     "folder. Add one with credential_validator_v2.py store "
                     "(or migrate), or drop a <service>.json in creds/.")
        return

    if positional:
        try:
            chosen = aep_creds.pick_service(positional[0])
        except aep_creds.CredsError as e:
            logger.error(str(e))
            return
    else:
        chosen = menu(services)

    if not chosen:
        logger.info("Nothing chosen. Exiting.")
        return

    run(chosen, sandbox_arg, rows, dataset_names=dataset_names or None,
        list_only=list_only)


if __name__ == "__main__":
    main()
