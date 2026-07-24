#!/usr/bin/env python3
"""
dd_audience_countdown.py
========================
Read-only expiry report for Data Distiller audiences. Lists every external
audience that originates from Data Distiller and reports how long each has left
before its data-expiry date, so you can spot the ones about to lapse.

READ-ONLY: this tool only ever issues GET requests. It never creates, renames,
extends, deletes, or otherwise mutates anything in AEP. No PUT/POST/PATCH/DELETE
anywhere.

It reads the AEP Real-Time Customer Profile audiences API
(GET /core/ups/audiences), paginates the full list, keeps only audiences whose
origin is Data Distiller (originName == "DATA_DISTILLER" / namespace "DDA"), and
for each one captures:
    name, id, createdDate, lastRefresh, ttlDays, dataExpiryDate,
    daysRemaining (calculated), profileCount, tags, lifecycleState
plus a "keepAlive" flag (TRUE when the audience carries the KEEP_ALIVE tag).

Audiences store their tags as tag-ID (UUID) references, so the tool first builds
an org tag ID->name map from the Unified Tags API
(GET https://experience.adobe.io/unifiedtags/tags) and resolves them -- without
that, the KEEP_ALIVE tag (stored as a UUID) is invisible. Pass --keep-alive-only
to report just the keep-alive-tagged audiences (the ones meant to persist).

NOTE ON EXPIRY: the audiences API exposes no absolute expiry date for Data
Distiller audiences -- only a TTL (`ttlInDays`, default 30). Data Distiller
audience data lapses `ttlInDays` after its last refresh, so the report DERIVES
    dataExpiryDate = lastRefresh + ttlInDays
where lastRefresh is the Audience-Portal halo force-refresh timestamp (parsed
from the audience's `audience_portal_halo_force_refresh_timestamp` tag), falling
back to the last-modified / creation time when that tag is absent. The
profile-metrics epochs are deliberately NOT used: they recompute daily for every
audience and would peg everything at a full TTL, hiding the countdown. An
audience that keeps being refreshed never approaches expiry; one whose refresh
stalls counts down.
The original brief's GET /core/ais/external-audiences is not a readable list
endpoint (the AIS service only accepts POST/create there), so /core/ups/audiences
is the authoritative read source -- the same one the other tools in this repo use.

The XLSX export needs openpyxl; without it a CSV fallback is written instead and
the run still completes.

Credentials come from the shared credential layer (aep_creds): the OS keyring
vault (Windows Credential Manager) first, falling back to a plaintext
creds/<service>.json folder only where keyring is unavailable. Manage the vault
with credential_validator_v2.py; run `credential_validator_v2.py list` to see the
service names. On a normal run it prompts you to:
    1. Pick which credential set (service) to use (e.g. acme-insurance).
    2. Pick which sandbox(es) to scan.
Both can be supplied on the command line to run unattended (see Usage).

Before it queries the tenant it always asks for confirmation -- the default is
NO, so nothing hits the tenant unless you explicitly say yes. Pass --yes/-y to
skip that prompt for unattended runs.

Usage:
    python dd_audience_countdown.py                  # interactive menus
    python dd_audience_countdown.py acme-insurance    # cred set by service name
    python dd_audience_countdown.py acme --sandbox prod
    python dd_audience_countdown.py --all             # every stored credential set
    python dd_audience_countdown.py acme --keep-alive-only  # only keep-alive tag
    python dd_audience_countdown.py acme --yes        # skip the confirm prompt

Output (to the standard ./output/ folder, which is gitignored):
    output/dd_audience_countdown_<tenant>_<UTCstamp>.xlsx
        One row per Data Distiller audience, sorted ascending by daysRemaining,
        with conditional formatting: daysRemaining <=7 red, <=14 amber.
    output/dd_audience_countdown_<tenant>_<UTCstamp>.csv
        Only when openpyxl is not installed (fallback, same rows).

Secrets prefer the keyring vault (nothing on disk). Any creds/<service>.json
fallback files are gitignored -- never commit them; they hold the client_secret.
"""

from __future__ import annotations

import csv
import json
import logging
import re
import sys
import urllib.error
import urllib.request
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlencode

import aep_creds  # keyring-backed credential store (replaces creds/*.json)

# ============================================================================
# CONFIG -- shared credential layer (aep_creds: keyring-first, creds/ fallback)
# ----------------------------------------------------------------------------
# Credentials are resolved through aep_creds by SERVICE NAME (the key you stored
# in the keyring vault, or the stem of a creds/<service>.json fallback file),
# picked interactively (or by name on the CLI). aep_creds.load_creds(service)
# returns the same dict shape the old JSON files produced. Recognised keys:
#   client_id          -- Adobe IMS client ID
#   client_secret      -- IMS client_credentials secret (REQUIRED -- mints a
#                         fresh token every run)
#   api_key            -- x-api-key, if different from client_id (optional)
#   org_id             -- Adobe org ID (e.g. "ABC@AdobeOrg")
#   oauth_url          -- IMS token endpoint (optional -- sensible default)
#   scopes             -- IMS scopes, comma-separated (optional -- default)
#   sandbox            -- "all" or a specific sandbox name (optional)
#   sandbox_names      -- fallback list when sandbox-management API is denied
# Optional keys (read for display only):
#   org_labels         -- {org_id: friendly label} for the output filename
#   bearer_token       -- pasted access token; fallback when no client_secret
# Keys "_"-prefixed are treated as comments and ignored (aep_creds strips them).
# ============================================================================

SCRIPT_DIR  = Path(__file__).resolve().parent
OUTPUT_DIR  = SCRIPT_DIR / "output"

DEFAULT_OAUTH_URL = "https://ims-na1.adobelogin.com/ims/token"
DEFAULT_SCOPES    = (
    "openid,AdobeID,read_organizations,"
    "additional_info.projectedProductContext,session"
)


def cred_menu(services: list[str]) -> list[str]:
    """Interactive picker over the available credential service names. Returns
    the chosen service name(s) (possibly several), or [] to quit."""
    bar = f"{ANSI['bold']}{ANSI['cyan']}{'=' * 70}{ANSI['reset']}"
    print()
    print(bar)
    print("  Credential bank   (OS keyring vault)")
    print("-" * 70)
    for i, name in enumerate(services, 1):
        print(f"  {i:>2}  {name}")
    print(bar)
    raw = input("\nPick set(s) by number (e.g. 1, 1,3, or 'all'), "
                "blank to quit: ").strip()
    if not raw:
        return []
    if raw.lower() == "all":
        return list(services)
    chosen: list[str] = []
    for tok in raw.replace(",", " ").split():
        if tok.isdigit() and 1 <= int(tok) <= len(services):
            chosen.append(services[int(tok) - 1])
        else:
            logger.warning(f"Ignoring invalid choice: {tok}")
    return chosen


# These are populated per-run by apply_config() once a credential set has been
# chosen. A run targets exactly one credential set; when several are selected,
# main() loops and re-applies for each.
CRED_SERVICE: str    = ""
_CFG: dict           = {}
BEARER_TOKEN         = ""
CLIENT_ID            = ""
CLIENT_SECRET        = ""
API_KEY              = ""
ORG_ID               = ""
OAUTH_URL            = DEFAULT_OAUTH_URL
SCOPES               = DEFAULT_SCOPES
SANDBOX              = "all"
SANDBOX_NAMES: list[str] = []
TENANT               = ""


def apply_config(conf: dict, service: str) -> None:
    """Load one chosen credential set into the module-level globals every other
    function reads. Also derives the tenant label and clears any cached token
    from a previous credential set."""
    global CRED_SERVICE, _CFG, BEARER_TOKEN, CLIENT_ID, CLIENT_SECRET, API_KEY
    global ORG_ID, OAUTH_URL, SCOPES, SANDBOX, SANDBOX_NAMES, TENANT
    global _CACHED_TOKEN
    CRED_SERVICE  = service
    _CFG          = conf
    BEARER_TOKEN  = conf.get("bearer_token", "")
    CLIENT_ID     = conf["client_id"]
    CLIENT_SECRET = conf.get("client_secret", "")
    API_KEY       = conf.get("api_key") or CLIENT_ID
    ORG_ID        = conf["org_id"]
    OAUTH_URL     = conf.get("oauth_url") or DEFAULT_OAUTH_URL
    SCOPES        = conf.get("scopes") or DEFAULT_SCOPES
    SANDBOX       = conf.get("sandbox") or "all"
    SANDBOX_NAMES = list(conf.get("sandbox_names") or [])
    TENANT        = tenant_for_org(ORG_ID)
    _CACHED_TOKEN = None

# ============================================================================

# Script identity (shown in the startup banner).
SCRIPT_NAME    = "dd_audience_countdown"
SCRIPT_VERSION = "1.1.0"
SCRIPT_DATE    = "2026-07-24"
SCRIPT_AUTHOR  = "Barry Mann (barrymann.com)"

# The Real-Time Customer Profile audiences API (`/core/ups/audiences`) lists
# EVERY audience in the sandbox -- segment-definition audiences, custom uploads,
# CJA, audience orchestration, AND Data Distiller audiences. We page it in full
# and keep only the Data Distiller ones (originName == "DATA_DISTILLER").
#
# The `/core/ais/external-audiences` path in the original brief is not a
# readable list endpoint -- the AIS ("Audience Import Service") routes only
# accept POST (create) and 404/405 on a GET list -- so `/core/ups/audiences` is
# the authoritative read source. The other tools in this repo use it too.
AUDIENCES_URL = ("https://platform.adobe.io/data/core/ups/audiences")
SANDBOX_URL   = ("https://platform.adobe.io/data/foundation/"
                 "sandbox-management/sandboxes")
# Unified Tags API (note the different host). Audiences carry their tags as bare
# tag-ID (UUID) references, NOT names -- e.g. the "KEEP_ALIVE" chip shown in the
# Audience Portal is stored on the audience as a UUID. This org-level API maps
# tag ID -> name so keepAlive detection and the Tags column resolve correctly.
UNIFIED_TAGS_URL = ("https://experience.adobe.io/unifiedtags/tags")
PAGE_LIMIT    = 100

# Data-expiry thresholds for the conditional formatting / summary counts.
RED_THRESHOLD   = 7    # daysRemaining <= 7  -> red
AMBER_THRESHOLD = 14   # daysRemaining <= 14 -> amber

# Data Distiller audiences carry a TTL (`ttlInDays`) but the audiences API does
# NOT expose an absolute expiry date. Their data lapses `ttlInDays` after the
# last refresh, so we DERIVE the expiry as (last refresh + TTL). An audience
# that keeps refreshing daily never approaches expiry; one whose refresh stalls
# counts down -- which is exactly the thing worth flagging. `ttlInDays` is
# frequently absent on the object; DDA's product default is 30 days.
DEFAULT_TTL_DAYS = 30

# Adobe doesn't expose org names via API, so we map org_id -> a friendly label
# to namespace the output filename so two orgs don't collide. Real
# org_id<->label pairs live in the gitignored config.json under "org_labels"
# (keeps real org IDs and client descriptors out of version control). The dict
# below is example-only.
ORG_LABELS: dict[str, str] = {
    # "<your-org-id>@AdobeOrg": "acme-insurance",  # real pairs go in config.json
}


def tenant_for_org(org_id: str) -> str:
    """Friendly label for an Adobe org_id. Consults config.json "org_labels"
    first (local, gitignored), then the example map, then falls back to a
    short prefix so different unknown orgs still get distinct filenames."""
    local_labels = _CFG.get("org_labels") or {}
    if org_id in local_labels:
        return local_labels[org_id]
    if org_id in ORG_LABELS:
        return ORG_LABELS[org_id]
    return f"org-{org_id.split('@')[0][:8]}"


# ----------------------------------------------------------------------------
# ANSI / logging - matches babelfish_query_fetcher.py style
# ----------------------------------------------------------------------------
if sys.platform == "win32":
    try:
        import ctypes
        kernel32 = ctypes.windll.kernel32
        h = kernel32.GetStdHandle(-11)
        mode = ctypes.c_ulong()
        if kernel32.GetConsoleMode(h, ctypes.byref(mode)):
            kernel32.SetConsoleMode(h, mode.value | 0x0004)
    except Exception:
        pass

ANSI = {
    "reset": "\033[0m", "bold": "\033[1m", "dim": "\033[2m",
    "red": "\033[31m", "green": "\033[32m", "yellow": "\033[33m",
    "blue": "\033[34m", "magenta": "\033[35m", "cyan": "\033[36m",
}
LEVEL_COLOR = {
    "DEBUG": ANSI["dim"], "INFO": ANSI["green"],
    "WARNING": ANSI["yellow"],
    "ERROR": ANSI["red"] + ANSI["bold"],
    "CRITICAL": ANSI["red"] + ANSI["bold"],
}


class ColoredFormatter(logging.Formatter):
    def format(self, record):
        color = LEVEL_COLOR.get(record.levelname, "")
        ts = self.formatTime(record, "%H:%M:%S")
        return (
            f"{ANSI['dim']}{ts}{ANSI['reset']} "
            f"{color}[{record.levelname:<7}]{ANSI['reset']} "
            f"{record.getMessage()}"
        )


_handler = logging.StreamHandler(sys.stdout)
_handler.setFormatter(ColoredFormatter())
logging.basicConfig(level=logging.INFO, handlers=[_handler])
logger = logging.getLogger(SCRIPT_NAME)


def print_banner() -> None:
    """Print a short header so each run is self-identifying in the log."""
    bar    = "=" * 72
    head   = f"{ANSI['bold']}{ANSI['cyan']}{bar}{ANSI['reset']}"
    title  = f"{ANSI['bold']}{SCRIPT_NAME} v{SCRIPT_VERSION}{ANSI['reset']}"
    print(head)
    print(f"  {title}   ({SCRIPT_DATE})")
    print(f"  by {SCRIPT_AUTHOR}")
    print( "  READ-ONLY: Data Distiller audience expiry countdown.")
    print( "  Never writes back to AEP. Output goes to ./output/.")
    print(f"  Tenant for this run: {TENANT}    (org_id: {ORG_ID})")
    print(head)


_CACHED_TOKEN: str | None = None


def _mask_secret(s: str, keep: int = 4) -> str:
    """Return a short masked version of a secret -- always 8 stars + last
    `keep` chars + length, so a long bearer token doesn't fill the screen."""
    if not s:
        return "(empty)"
    if len(s) <= keep:
        return "*" * len(s)
    return f"********{s[-keep:]} ({len(s)} chars)"


def fetch_oauth_token() -> str:
    """POST to Adobe IMS to mint a fresh access token via client_credentials."""
    body = urlencode({
        "client_id":     CLIENT_ID,
        "client_secret": CLIENT_SECRET,
        "grant_type":    "client_credentials",
        "scope":         SCOPES,
    }).encode("utf-8")
    req = urllib.request.Request(
        OAUTH_URL,
        data=body,
        headers={"Content-Type": "application/x-www-form-urlencoded"},
        method="POST",
    )
    logger.info(f"POST {OAUTH_URL} (client_credentials, "
                f"client_id={CLIENT_ID}, "
                f"client_secret={_mask_secret(CLIENT_SECRET)})...")
    try:
        with urllib.request.urlopen(req) as resp:
            payload = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        text = e.read().decode("utf-8", errors="replace")[:300]
        logger.error(f"IMS auth failed: HTTP {e.code} {text}")
        sys.exit(1)
    except urllib.error.URLError as e:
        logger.error(f"IMS auth network error: {e.reason}")
        sys.exit(1)
    token = payload.get("access_token")
    if not token:
        logger.error(f"IMS response had no access_token: {payload}")
        sys.exit(1)
    expires_in = payload.get("expires_in", "?")
    logger.info(f"OK - got access token (expires in {expires_in}s).")
    return token


def get_token() -> str:
    """Return a valid bearer token, fetching from IMS if needed (cached per run).

    Prefers minting via OAuth client_credentials so every session picks up
    current permissions/credentials. Only falls back to the pasted
    BEARER_TOKEN when no client_secret is configured."""
    global _CACHED_TOKEN
    if _CACHED_TOKEN is None:
        if CLIENT_SECRET:
            _CACHED_TOKEN = fetch_oauth_token()
        elif BEARER_TOKEN:
            logger.info(f"No client_secret configured; using fallback "
                        f"bearer_token={_mask_secret(BEARER_TOKEN)} "
                        f"(may be expired).")
            _CACHED_TOKEN = BEARER_TOKEN
        else:
            logger.error("Neither client_secret nor bearer_token is set in "
                         "the credential file - cannot authenticate.")
            sys.exit(1)
    return _CACHED_TOKEN


def auth_headers(sandbox: str | None = None) -> dict:
    """Build the standard request headers. `sandbox` overrides SANDBOX for
    requests that need to target a specific sandbox."""
    headers = {
        "Authorization":   f"Bearer {get_token()}",
        "x-api-key":       API_KEY or CLIENT_ID,
        "x-gw-ims-org-id": ORG_ID,
        "Accept":          "application/json",
        "Content-Type":    "application/json",
    }
    sb = sandbox if sandbox is not None else SANDBOX
    if sb and sb != "all":
        headers["x-sandbox-name"] = sb
    return headers


def http_request(method, url, headers, params=None, body=None):
    """Stdlib-only HTTP. Returns (status_code, response_text).

    Never raises on HTTP error codes -- 4xx/5xx are returned as a normal
    (status, text) pair so callers can branch on them."""
    if params:
        url = f"{url}?{urlencode(params)}"
    data = json.dumps(body).encode("utf-8") if body is not None else None
    req = urllib.request.Request(url, data=data, headers=headers, method=method)
    try:
        with urllib.request.urlopen(req) as resp:
            return resp.status, resp.read().decode("utf-8")
    except urllib.error.HTTPError as e:
        return e.code, e.read().decode("utf-8", errors="replace")
    except urllib.error.URLError as e:
        # Network/DNS/proxy failures land here. Surface them as a 0 status.
        return 0, f"URLError: {e.reason}"


def list_sandboxes() -> list[str]:
    """Return the names of every sandbox the token can see."""
    logger.info(f"GET {SANDBOX_URL} (listing all sandboxes)...")
    # Sandbox-management endpoint does NOT take x-sandbox-name itself.
    headers = auth_headers(sandbox="")
    headers.pop("x-sandbox-name", None)
    status, text = http_request("GET", SANDBOX_URL, headers)
    if status == 403:
        logger.warning("Sandbox-management API denied (403) -- token lacks the "
                       "management read scope. Falling back to sandbox_names "
                       "from the credential file.")
        return []
    if status < 200 or status >= 300:
        logger.error(f"Sandbox list failed: HTTP {status} {text[:200]}")
        return []
    body = json.loads(text)
    names = [s.get("name") for s in body.get("sandboxes", []) if s.get("name")]
    logger.info(f"  -> sandboxes available: {names}")
    return names


def discover_sandboxes() -> list[str]:
    """Resolve which sandboxes to scan.

    SANDBOX="all" tries the sandbox-management API. If that fails (typically
    HTTP 403 on tokens without management scope), fall back to SANDBOX_NAMES so
    the script still works on minimally-scoped tokens. Hard-fails only when both
    the API call AND the configured fallback list are empty."""
    if SANDBOX != "all":
        return [SANDBOX]
    sandboxes = list_sandboxes()
    if sandboxes:
        logger.info(f"Using {len(sandboxes)} sandbox(es) from "
                    f"sandbox-management API.")
        return sandboxes
    if SANDBOX_NAMES:
        logger.info(f"Sandbox-management API returned nothing; using "
                    f"configured SANDBOX_NAMES fallback: {SANDBOX_NAMES}")
        return list(SANDBOX_NAMES)
    logger.error("Sandbox listing returned empty AND SANDBOX_NAMES is empty. "
                 "Either get a token with sandbox-management read scope, or "
                 "fill in sandbox_names in the credential file.")
    sys.exit(1)


def _interactive() -> bool:
    """True when we can prompt the user (stdin is a real terminal). When piped
    or scheduled, menus are skipped and sensible defaults are used instead."""
    try:
        return sys.stdin.isatty()
    except Exception:
        return False


def select_sandboxes(flags: dict) -> list[str]:
    """Resolve which sandbox(es) to scan for this run.

    Precedence: an explicit --sandbox on the CLI wins; otherwise we discover
    the sandboxes this credential can see and, when interactive, let the user
    narrow to a subset. Non-interactive runs scan every discovered sandbox."""
    available = discover_sandboxes()
    if flags.get("sandboxes"):
        wanted = flags["sandboxes"]
        known  = {s.lower(): s for s in available}
        chosen = [known.get(w.lower(), w) for w in wanted]
        logger.info(f"Sandbox(es) from --sandbox: {chosen}")
        return chosen
    if not _interactive() or len(available) <= 1:
        return available

    bar = f"{ANSI['bold']}{ANSI['cyan']}{'=' * 70}{ANSI['reset']}"
    print()
    print(bar)
    print(f"  Sandboxes visible to '{TENANT}'")
    print("-" * 70)
    for i, sb in enumerate(available, 1):
        print(f"  {i:>2}  {sb}")
    print(bar)
    raw = input("\nScan which sandbox(es)? number(s) e.g. 1,3, or 'all' "
                "[all]: ").strip()
    if not raw or raw.lower() == "all":
        return available
    chosen: list[str] = []
    for tok in raw.replace(",", " ").split():
        if tok.isdigit() and 1 <= int(tok) <= len(available):
            chosen.append(available[int(tok) - 1])
        else:
            logger.warning(f"Ignoring invalid choice: {tok}")
    return chosen or available


# ----------------------------------------------------------------------------
# External-audience fetch + field extraction
# ----------------------------------------------------------------------------
# The audiences API's field names have drifted across versions, so every getter
# below tolerates a few spellings rather than pinning one. Mirrors the tolerant
# style used elsewhere in this toolkit (e.g. the query fetcher's schedule
# templateId resolver).

# Keys that, across API versions, have listed the audience array in the body.
# `/core/ups/audiences` uses `children`; older deployments used `segments`.
_LIST_KEYS = ("children", "segments", "audiences", "content", "items")
# Origin/source fields consulted to decide "is this a Data Distiller audience".
# On `/core/ups/audiences` the field is `originName` with the exact literal
# "DATA_DISTILLER", and Data Distiller audiences also carry namespace "DDA".
_ORIGIN_KEYS = ("originName", "origin", "source", "sourceName", "dataSource")
_DD_ORIGIN_VALUES = {"data_distiller", "dda", "data distiller"}
_NO_ACCESS_SANDBOXES: list[str] = []  # populated by fetch_audiences_in_sandbox

# {tagId: tagName} resolved once per credential set from the Unified Tags API.
_TAG_NAMES: dict[str, str] = {}
_UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$", re.I)
# Internal housekeeping markers AEP writes into an audience's tag list -- not
# user-applied tags, so we hide them from the Tags column.
_HOUSEKEEPING_TAG_PREFIXES = ("audience_portal_",)


def fetch_tag_names() -> dict:
    """Build an org-wide {tagId: tagName} map from the Unified Tags API.

    Tags are org-level (not sandbox-scoped), so this is fetched once per
    credential set without a sandbox header. On any failure it returns {} and
    tag UUIDs are simply left unresolved -- keepAlive detection then falls back
    to matching literal tag strings."""
    headers = auth_headers(sandbox="")
    headers.pop("x-sandbox-name", None)
    names: dict[str, str] = {}
    start = None
    page = 0
    while True:
        page += 1
        params = {"limit": PAGE_LIMIT}
        if start:
            params["start"] = start
        status, text = http_request("GET", UNIFIED_TAGS_URL, headers,
                                    params=params)
        if status < 200 or status >= 300:
            logger.warning(f"Unified Tags API returned HTTP {status}; tag IDs "
                           f"won't be resolved to names (keep-alive tags stored "
                           f"as IDs may be missed).")
            break
        body = json.loads(text)
        batch = body.get("tags") or []
        for tg in batch:
            if tg.get("id"):
                names[tg["id"]] = tg.get("name") or tg["id"]
        next_cursor = (body.get("_page") or {}).get("next")
        if not batch or len(batch) < PAGE_LIMIT or not next_cursor:
            break
        start = next_cursor
    logger.info(f"Resolved {len(names)} tag name(s) from the Unified Tags API.")
    return names


def _first(d: dict, keys, default=None):
    """Return the first present, non-empty value among `keys` in dict `d`."""
    for k in keys:
        v = d.get(k)
        if v not in (None, "", [], {}):
            return v
    return default


def _audience_list(body: dict) -> list[dict]:
    """Pull the array of audience objects out of a response body, whichever key
    this API version happens to use."""
    for k in _LIST_KEYS:
        v = body.get(k)
        if isinstance(v, list):
            return v
    return []


def _profile_count(aud: dict):
    """Best-effort profile count. On `/core/ups/audiences` the live count lives
    at `metrics.data.totalProfiles`; we check that first, then the record-export
    count, then a few flatter shapes older deployments used."""
    metrics = aud.get("metrics")
    if isinstance(metrics, dict):
        data = metrics.get("data")
        if isinstance(data, dict):
            for key in ("totalProfiles", "profileCount", "count"):
                if data.get(key) is not None:
                    return data[key]
        mv = _first(metrics, ("profileCount", "totalProfiles", "count"))
        if mv is not None:
            return mv
    rec = aud.get("recordMetrics")
    if isinstance(rec, dict):
        data = rec.get("data")
        if isinstance(data, dict) and data.get("recordCount") is not None:
            return data["recordCount"]
    return _first(aud, ("profileCount", "profileEstimate", "totalProfiles",
                        "count"))


def _tags(aud: dict) -> list[str]:
    """Flatten whatever the audience carries as tags/labels into a list of
    display strings, so both display and the keep-alive test have a single shape
    to work with. Handles list-of-strings, list-of-dicts, a dict map, or a
    comma-separated string.

    Bare tag-ID (UUID) references are resolved to their tag name via the
    Unified Tags map (`_TAG_NAMES`) -- this is how "KEEP_ALIVE" is recovered,
    since audiences store it as a UUID. Internal `audience_portal_*`
    housekeeping markers are dropped -- they are not user-applied tags."""
    raw = _first(aud, ("tags", "labels", "audienceTags"), default=[])
    out: list[str] = []
    if isinstance(raw, str):
        out = [t.strip() for t in raw.split(",") if t.strip()]
    elif isinstance(raw, dict):
        for k, v in raw.items():
            out.append(f"{k}={v}" if v not in (None, "", True) else str(k))
    elif isinstance(raw, list):
        for item in raw:
            if isinstance(item, str):
                out.append(item)
            elif isinstance(item, dict):
                out.append(str(_first(item, ("name", "label", "value", "tag"),
                                      default=item)))
            else:
                out.append(str(item))

    resolved: list[str] = []
    for t in (s.strip() for s in out):
        if not t or t.startswith(_HOUSEKEEPING_TAG_PREFIXES):
            continue
        resolved.append(_TAG_NAMES.get(t, t) if _UUID_RE.match(t) else t)
    return resolved


def _has_keep_alive(tags: list[str]) -> bool:
    """TRUE when any tag contains "keep-alive" (case-insensitive). Separator
    variants (space / underscore) are tolerated so 'Keep Alive' and
    'keep_alive' also count."""
    for t in tags:
        norm = re.sub(r"[\s_]+", "-", t.strip().lower())
        if "keep-alive" in norm:
            return True
    return False


def _is_data_distiller(aud: dict) -> bool:
    """TRUE when the audience originates from Data Distiller. Matches the
    `originName`/origin literal ("DATA_DISTILLER") or the "DDA" namespace, and
    falls back to a substring test so any 'distiller' spelling still counts."""
    if str(aud.get("namespace", "")).strip().lower() == "dda":
        return True
    for k in _ORIGIN_KEYS:
        v = aud.get(k)
        if isinstance(v, str):
            norm = v.strip().lower()
            if norm in _DD_ORIGIN_VALUES or "distiller" in norm:
                return True
    return False


def _origin_label(aud: dict) -> str:
    """The origin/source string we matched on -- for the report's Origin column
    (e.g. "DATA_DISTILLER")."""
    for k in _ORIGIN_KEYS:
        v = aud.get(k)
        if isinstance(v, str) and v.strip():
            return v.strip()
    return ""


def _parse_dt(val) -> datetime | None:
    """Parse an epoch (seconds or millis, int or numeric string) or an ISO-8601
    string into a timezone-aware UTC datetime. Returns None when empty/unknown.
    Naive values are assumed to be UTC."""
    if val in (None, "", 0, "0"):
        return None
    if isinstance(val, (int, float)):
        secs = val / 1000 if val > 1e12 else val
        try:
            return datetime.fromtimestamp(secs, tz=timezone.utc)
        except (OverflowError, OSError, ValueError):
            return None
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        if s.isdigit():
            n = int(s)
            secs = n / 1000 if n > 1e12 else n
            try:
                return datetime.fromtimestamp(secs, tz=timezone.utc)
            except (OverflowError, OSError, ValueError):
                return None
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
        except ValueError:
            for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
                try:
                    return datetime.strptime(s[:len(fmt) + 2], fmt).replace(
                        tzinfo=timezone.utc)
                except ValueError:
                    continue
    return None


def _fmt_dt(dt: datetime | None) -> str:
    """Render a parsed datetime as a plain UTC date for the report."""
    return dt.strftime("%Y-%m-%d") if dt else ""


def _days_remaining(expiry: datetime | None):
    """Whole calendar days from today (UTC) until the expiry date. Negative
    when already expired; None when the expiry is unknown."""
    if expiry is None:
        return None
    return (expiry.date() - datetime.now(timezone.utc).date()).days


def _ttl_days(aud: dict) -> int:
    """The audience's effective data TTL in days. Uses `ttlInDays` when it is a
    positive number, otherwise the Data Distiller product default. A 0 (or
    absent) TTL is treated as 'unset' -> default, rather than as a genuine
    0-day retention, so an unconfigured audience is not falsely flagged as
    expiring today."""
    v = _first(aud, ("ttlInDays", "ttl", "dataTtlInDays"))
    try:
        n = int(v) if v is not None else 0
    except (TypeError, ValueError):
        n = 0
    return n if n > 0 else DEFAULT_TTL_DAYS


_HALO_REFRESH_RE = re.compile(
    r"audience_portal_halo_force_refresh_timestamp:(\d+)")


def _halo_refresh(aud: dict) -> datetime | None:
    """The last Audience-Portal force-refresh time, parsed from the
    `audience_portal_halo_force_refresh_timestamp:<ms>` housekeeping tag. This
    is the truest "the audience data was actually rebuilt" signal the object
    carries, and unlike the metrics epochs it genuinely varies per audience."""
    raw = aud.get("tags")
    if isinstance(raw, list):
        for tg in raw:
            match = _HALO_REFRESH_RE.match(str(tg))
            if match:
                return _parse_dt(int(match.group(1)))
    return None


def _last_refresh(aud: dict) -> datetime | None:
    """Best estimate of when the audience DATA was last refreshed, used to
    anchor the derived expiry.

    Priority:
      1. The Audience-Portal halo force-refresh timestamp tag -- the real
         rebuild time (present on portal-managed audiences, incl. keep-alive).
      2. Otherwise the audience's last-modified / creation time.

    The profile-metrics and record-export `updateEpoch`s are deliberately NOT
    used: they are recomputed daily for every audience regardless of whether the
    membership changed, so anchoring on them pegs every audience at a full TTL
    and hides the very countdown this report exists to show.

    Limitation: an audience refreshed purely by a scheduled Data Distiller query
    (new batches into its backing dataset) carries no refresh signal on the
    audience object, so it falls back to (2) and can read as more-expired than
    it really is. Portal/keep-alive audiences -- the ones that matter here -- do
    carry the halo timestamp, so they are anchored correctly."""
    refresh = _halo_refresh(aud)
    if refresh:
        return refresh
    return _parse_dt(_first(aud, ("updateEpoch", "updateTime", "createEpoch",
                                  "creationTime")))


def _derived_expiry(aud: dict) -> tuple[datetime | None, datetime | None, int]:
    """Return (expiry, last_refresh, ttl_days). Expiry = last_refresh + TTL;
    None when there is no refresh timestamp to anchor it to. The audiences API
    exposes no absolute expiry date for Data Distiller audiences, so this is a
    derivation -- see the module docstring."""
    from datetime import timedelta
    ttl = _ttl_days(aud)
    refresh = _last_refresh(aud)
    expiry = refresh + timedelta(days=ttl) if refresh else None
    return expiry, refresh, ttl


def build_audience_row(aud: dict, sandbox: str) -> dict:
    """Extract the reporting fields from one raw audience object into the flat
    row shape the exporter and summary consume."""
    created = _parse_dt(_first(aud, ("createEpoch", "creationTime",
                                     "createdDate", "createTime", "created")))
    expiry, refresh, ttl = _derived_expiry(aud)
    tags = _tags(aud)
    return {
        "sandbox":        sandbox,
        "name":           _first(aud, ("name", "audienceName"),
                                  default="(unnamed)"),
        "id":             _first(aud, ("audienceId", "id", "sid",
                                       "externalAudienceId"), default=""),
        "createdDate":    _fmt_dt(created),
        "lastRefresh":    _fmt_dt(refresh),
        "ttlDays":        ttl,
        "dataExpiryDate": _fmt_dt(expiry),
        "daysRemaining":  _days_remaining(expiry),
        "profileCount":   _profile_count(aud),
        "lifecycleState": _first(aud, ("lifecycleState", "lifecycle", "state",
                                       "status"), default=""),
        "keepAlive":      _has_keep_alive(tags),
        "tags":           ", ".join(tags),
        "origin":         _origin_label(aud),
    }


def fetch_audiences_in_sandbox(sandbox: str) -> list[dict]:
    """Fetch every external audience from a single sandbox, fully paginated.

    A 403 here means the token authenticated fine but lacks audience-read
    permission in *this* sandbox -- expected on narrowly-scoped tokens, not a
    fatal error. Warn, record it, and move on."""
    logger.info(f"Sandbox '{sandbox}': listing external audiences...")
    headers = auth_headers(sandbox=sandbox)
    out: list[dict] = []
    start = None
    page = 0
    while True:
        page += 1
        params = {"limit": PAGE_LIMIT}
        if start:
            params["start"] = start
        status, text = http_request("GET", AUDIENCES_URL, headers,
                                     params=params)
        if status == 401:
            logger.error("401 Unauthorized - token expired/invalid.")
            sys.exit(1)
        if status == 403:
            logger.warning(f"  no audience access in sandbox '{sandbox}' "
                           f"(token lacks the right scope/role here).")
            _NO_ACCESS_SANDBOXES.append(sandbox)
            return []
        if status < 200 or status >= 300:
            logger.error(f"  page {page}: HTTP {status} {text[:200]}")
            break
        body = json.loads(text)
        batch = _audience_list(body)
        for a in batch:
            a["_sandbox"] = sandbox
        out.extend(batch)
        logger.info(f"  page {page}: got {len(batch)} "
                    f"(sandbox total {len(out)}).")
        next_cursor = (body.get("_page") or {}).get("next")
        if not batch or len(batch) < PAGE_LIMIT or not next_cursor:
            break
        start = next_cursor
    logger.info(f"  sandbox '{sandbox}' done - {len(out)} audience(s).")
    return out


def fetch_all_audiences(sandboxes: list[str]) -> list[dict]:
    """Fetch external audiences from each of the given sandboxes."""
    _NO_ACCESS_SANDBOXES.clear()
    all_auds: list[dict] = []
    for sb in sandboxes:
        all_auds.extend(fetch_audiences_in_sandbox(sb))
    accessed = len(sandboxes) - len(_NO_ACCESS_SANDBOXES)
    summary = (f"Done - {len(all_auds)} audience(s) from {accessed} of "
               f"{len(sandboxes)} sandbox(es)")
    if _NO_ACCESS_SANDBOXES:
        summary += f"; no access in: {', '.join(_NO_ACCESS_SANDBOXES)}"
    logger.info(summary + ".")
    return all_auds


# ----------------------------------------------------------------------------
# Output
# ----------------------------------------------------------------------------
# Column order for both the XLSX and the CSV fallback. (key, header, width).
_COLUMNS = [
    ("sandbox",        "Sandbox",           12),
    ("name",           "Name",              42),
    ("id",             "ID",                38),
    ("createdDate",    "Created",           12),
    ("lastRefresh",    "Last Refresh",      13),
    ("ttlDays",        "TTL (days)",        11),
    ("dataExpiryDate", "Data Expiry (est)", 17),
    ("daysRemaining",  "Days Remaining",    15),
    ("profileCount",   "Profile Count",     14),
    ("lifecycleState", "Lifecycle State",   16),
    ("keepAlive",      "Keep-Alive",        11),
    ("tags",           "Tags",              40),
    ("origin",         "Origin",            16),
]

_XLSX_HEADER_BG = "1F4E78"    # dark blue header band (matches sibling tools)
_XLSX_RED_FILL  = "FFC7CE"    # Excel classic "bad" red
_XLSX_RED_FONT  = "9C0006"
_XLSX_AMBER_FILL = "FFEB9C"   # Excel classic "neutral" amber
_XLSX_AMBER_FONT = "9C6500"


def _sort_key(row: dict):
    """Sort ascending by daysRemaining; audiences with no expiry sort last."""
    dr = row["daysRemaining"]
    return (dr is None, dr if dr is not None else 0)


def _timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%SZ")


def _safe_stem(name: str) -> str:
    return re.sub(r"[^0-9A-Za-z._-]+", "-", name).strip("-") or "creds"


def write_xlsx(rows: list[dict]) -> Path | None:
    """Write the sorted audience rows to a single-sheet workbook with the
    daysRemaining conditional formatting. Returns the path, or None if openpyxl
    is not installed (the caller then falls back to CSV)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("openpyxl not installed -- writing a CSV instead "
                       "(pip install openpyxl for the formatted workbook).")
        return None

    OUTPUT_DIR.mkdir(exist_ok=True)
    path = OUTPUT_DIR / (f"dd_audience_countdown_{_safe_stem(TENANT)}_"
                         f"{_timestamp()}.xlsx")

    head_font  = Font(bold=True, color="FFFFFF")
    head_fill  = PatternFill("solid", fgColor=_XLSX_HEADER_BG)
    title_font = Font(bold=True, size=14)
    red_fill   = PatternFill("solid", fgColor=_XLSX_RED_FILL)
    red_font   = Font(color=_XLSX_RED_FONT, bold=True)
    amber_fill = PatternFill("solid", fgColor=_XLSX_AMBER_FILL)
    amber_font = Font(color=_XLSX_AMBER_FONT, bold=True)
    center     = Alignment(horizontal="center")

    wb = Workbook()
    ws = wb.active
    ws.title = "DD Audiences"

    # Header row.
    for c, (_key, header, _w) in enumerate(_COLUMNS, 1):
        cell = ws.cell(1, c, header)
        cell.font = head_font
        cell.fill = head_fill
    ws.freeze_panes = "A2"

    days_col = next(i for i, (k, _, _) in enumerate(_COLUMNS, 1)
                    if k == "daysRemaining")

    for ridx, row in enumerate(rows, 2):
        for cidx, (key, _header, _w) in enumerate(_COLUMNS, 1):
            val = row.get(key)
            if key == "keepAlive":
                val = "TRUE" if val else ""
            cell = ws.cell(ridx, cidx, val)
            if key in ("daysRemaining", "profileCount", "keepAlive",
                       "ttlDays"):
                cell.alignment = center
        # Conditional formatting on the daysRemaining cell.
        dr = row["daysRemaining"]
        if dr is not None:
            dcell = ws.cell(ridx, days_col)
            if dr <= RED_THRESHOLD:
                dcell.fill, dcell.font = red_fill, red_font
            elif dr <= AMBER_THRESHOLD:
                dcell.fill, dcell.font = amber_fill, amber_font

    for idx, (_k, _h, w) in enumerate(_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    # A small legend row under the table so the colours are self-documenting.
    legend_row = len(rows) + 3
    ws.cell(legend_row, 1, "Legend:").font = title_font
    lc = ws.cell(legend_row + 1, 1, f"<= {RED_THRESHOLD} days")
    lc.fill, lc.font = red_fill, red_font
    la = ws.cell(legend_row + 2, 1, f"<= {AMBER_THRESHOLD} days")
    la.fill, la.font = amber_fill, amber_font

    wb.save(path)
    return path


def write_csv(rows: list[dict]) -> Path:
    """Fallback export when openpyxl is unavailable -- same rows/columns, no
    formatting."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    path = OUTPUT_DIR / (f"dd_audience_countdown_{_safe_stem(TENANT)}_"
                         f"{_timestamp()}.csv")
    with path.open("w", newline="", encoding="utf-8-sig") as fh:
        writer = csv.writer(fh)
        writer.writerow([h for _k, h, _w in _COLUMNS])
        for row in rows:
            writer.writerow([
                ("TRUE" if row.get("keepAlive") else "") if k == "keepAlive"
                else ("" if row.get(k) is None else row.get(k))
                for k, _h, _w in _COLUMNS
            ])
    return path


def print_console_summary(rows: list[dict]) -> None:
    """The required console summary: total DD audiences, count expiring within
    7 days, count tagged keep-alive."""
    total = len(rows)
    expiring_7 = sum(1 for r in rows
                     if r["daysRemaining"] is not None
                     and r["daysRemaining"] <= RED_THRESHOLD)
    keep_alive = sum(1 for r in rows if r["keepAlive"])

    bar = f"{ANSI['bold']}{ANSI['cyan']}{'=' * 60}{ANSI['reset']}"
    print()
    print(bar)
    print(f"  Data Distiller audience countdown  -  {TENANT}")
    print("-" * 60)
    print(f"  Total Data Distiller audiences      : {total}")
    print(f"  Expiring within {RED_THRESHOLD} days"
          f"{'':<15}: {ANSI['red']}{ANSI['bold']}{expiring_7}{ANSI['reset']}")
    print(f"  Tagged keep-alive                   : {keep_alive}")
    print(bar)


def confirm_run(flags: dict) -> bool:
    """Safety gate asked once per credential set, BEFORE anything touches the
    tenant. The run is read-only (GETs only, no writes back to AEP), but we
    still ask first and the default is NO -- pressing Enter, or any answer that
    is not an explicit yes, aborts this credential set.

    --yes/-y bypasses the prompt for unattended runs. When not interactive and
    --yes was not given, we default to NO and skip."""
    if flags.get("yes"):
        logger.info(f"--yes given; proceeding with '{TENANT}'.")
        return True
    if not _interactive():
        logger.warning(f"Not interactive and --yes not given -- skipping "
                       f"'{TENANT}' (default is no).")
        return False
    ans = input(f"\nQuery tenant '{TENANT}' now? Read-only -- no changes are "
                f"made to the tenant. [y/N]: ").strip().lower()
    if ans in ("y", "yes"):
        return True
    logger.info(f"Not confirmed -- skipping '{TENANT}'.")
    return False


def run_for_cred(service: str, flags: dict) -> None:
    """Run the full read-only fetch + report pipeline for one credential set."""
    try:
        conf = aep_creds.load_creds(service)
    except aep_creds.CredsError as e:
        logger.error(f"Failed to load credentials for {service!r}: {e}")
        return
    apply_config(conf, service)

    print_banner()

    # 0. Confirm before touching the tenant. Default NO.
    if not confirm_run(flags):
        return

    logger.info(f"{SCRIPT_NAME} starting for '{TENANT}' "
                f"(read-only -- no writes back to AEP).")

    # 1. Resolve the org's tag ID -> name map up front, so the UUID tag
    #    references on audiences (e.g. KEEP_ALIVE) resolve to names.
    global _TAG_NAMES
    _TAG_NAMES = fetch_tag_names()

    # 2. Choose which sandbox(es) to scan (CLI flag, prompt, or all).
    sandboxes = select_sandboxes(flags)

    # 3. Fetch every external audience from every accessible sandbox.
    audiences = fetch_all_audiences(sandboxes)

    # 4. Filter to Data Distiller origin only.
    dd = [a for a in audiences if _is_data_distiller(a)]
    logger.info(f"Data Distiller audiences: {len(dd)} of {len(audiences)} "
                f"external audience(s).")

    # 5. Flatten to report rows and sort ascending by daysRemaining.
    rows = [build_audience_row(a, a.get("_sandbox", "?")) for a in dd]

    # 5b. Optional --keep-alive-only: keep just the audiences meant to persist
    #     (those carrying the keep-alive tag), which are the ones whose expiry
    #     actually matters.
    if flags.get("keep_alive_only"):
        before = len(rows)
        rows = [r for r in rows if r["keepAlive"]]
        logger.info(f"--keep-alive-only: {len(rows)} of {before} Data Distiller "
                    f"audience(s) carry the keep-alive tag.")

    rows.sort(key=_sort_key)

    # 6. Console summary (always).
    print_console_summary(rows)

    if not rows:
        logger.info("No matching Data Distiller audiences -- nothing to export.")
        return

    # 7. Write the XLSX (conditional formatting), or a CSV fallback.
    xlsx_path = write_xlsx(rows)
    if xlsx_path:
        logger.info(f"XLSX: {xlsx_path.relative_to(SCRIPT_DIR)} "
                    f"({len(rows)} audience(s))")
    else:
        csv_path = write_csv(rows)
        logger.info(f"CSV:  {csv_path.relative_to(SCRIPT_DIR)} "
                    f"({len(rows)} audience(s))")


def parse_args(argv: list[str]) -> tuple[dict, list[str]]:
    """Split argv into (flags, names). Positional names are credential service
    names (spaces or hyphens both accepted). Flags:
        --all / -a        select every stored credential set
        --sandbox NAME     scan only NAME (repeatable) -- skips the sandbox menu
        --keep-alive-only  keep only audiences carrying the keep-alive tag
        --yes / -y         skip the "query tenant?" confirmation (default is no)
    """
    flags: dict = {"all": False, "sandboxes": [], "yes": False,
                   "keep_alive_only": False}
    names: list[str] = []
    i = 0
    while i < len(argv):
        a = argv[i]
        if a in ("--all", "-a"):
            flags["all"] = True
        elif a in ("--yes", "-y"):
            flags["yes"] = True
        elif a in ("--keep-alive-only", "--keepalive-only", "--keep-alive"):
            flags["keep_alive_only"] = True
        elif a in ("--sandbox", "--sandboxes", "-s"):
            if i + 1 < len(argv):
                flags["sandboxes"].append(argv[i + 1])
                i += 1
        elif a.startswith("--sandbox="):
            flags["sandboxes"].append(a.split("=", 1)[1])
        elif a.startswith("-"):
            logger.warning(f"Ignoring unknown flag: {a}")
        else:
            names.append(a)
        i += 1
    return flags, names


def _match_cred(name: str, services: list[str]) -> str | None:
    """Resolve a credential name from the CLI to a service name. Matches on the
    service name, tolerating hyphen/space/underscore differences and case."""
    def norm(s: str) -> str:
        return s.lower().replace(" ", "").replace("-", "").replace("_", "")
    target = norm(name)
    for svc in services:
        if norm(svc) == target:
            return svc
    return None


def main() -> None:
    flags, names = parse_args(sys.argv[1:])
    logger.info(aep_creds.source_banner())
    services = aep_creds.list_services()

    if not services:
        logger.error("No credentials found in the keyring vault or the creds/ "
                     "folder. Add one with credential_validator_v2.py store "
                     "(or migrate), or drop a <service>.json in creds/.")
        sys.exit(1)

    if flags["all"]:
        chosen = list(services)
    elif names:
        chosen = []
        for n in names:
            svc = _match_cred(n, services)
            if svc:
                chosen.append(svc)
            else:
                logger.warning(f"No credential set named {n!r}. "
                               f"Known: {', '.join(services)}")
        if not chosen:
            sys.exit(1)
    elif _interactive():
        chosen = cred_menu(services)
    else:
        logger.error("No credential set specified and not running "
                     "interactively. Pass a name or --all.")
        sys.exit(1)

    if not chosen:
        logger.warning("Nothing chosen. Exiting.")
        return

    for service in chosen:
        run_for_cred(service, flags)


if __name__ == "__main__":
    main()
